"""
İcra Yardım Programı v4.0
T.C. Isparta İl Özel İdaresi — Hukuk Müşavirliği
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3, os, sys, base64, tempfile, hashlib, re
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, Image as RLImage, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

BASE_DIR = os.path.dirname(os.path.abspath(sys.argv[0] if getattr(sys,"frozen",False) else __file__))

# Veritabanı her zaman EXE veya script ile aynı klasörde
DB_PATH      = os.path.join(BASE_DIR, "icra_veri.db")
HATIRLA_PATH = os.path.join(BASE_DIR, "hatirla.txt")

# ── PDF Türkçe font ───────────────────────────────────────────────
def kaydet_pdf_font():
    """Windows Segoe UI veya Arial Unicode ile Türkçe PDF"""
    paths = [
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\calibri.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ]
    for p in paths:
        if os.path.exists(p):
            try:
                pdfmetrics.registerFont(TTFont("TurkFont", p))
                pdfmetrics.registerFont(TTFont("TurkFontBold", p.replace(".ttf","-Bold.ttf") if os.path.exists(p.replace(".ttf","-Bold.ttf")) else p))
                return "TurkFont", "TurkFontBold"
            except: pass
    return "Helvetica", "Helvetica-Bold"

PDF_FONT, PDF_FONT_BOLD = kaydet_pdf_font()

# ── Renkler ───────────────────────────────────────────────────────
CLR = {
    "bg":       "#1E2E42",   # Açık lacivert arka plan
    "panel":    "#253548",   # Panel arka planı
    "card":     "#2C4060",   # Kart arka planı
    "accent":   "#F5B731",   # Altın sarı vurgu
    "accent2":  "#3AACF0",   # Mavi vurgu
    "success":  "#28A745",   # Yeşil
    "danger":   "#DC3545",   # Kırmızı
    "warning":  "#FD7E14",   # Turuncu
    "red":      "#CC1A1A",   # Kurumsal kırmızı
    "text":     "#FFFFFF",   # Beyaz metin (daha belirgin)
    "subtext":  "#B8D4F0",   # Açık mavi alt metin
    "border":   "#3A5070",   # Kenarlık
    "hover":    "#2E5080",   # Hover rengi
    "tbl_odd":  "#253548",   # Tablo tek satır
    "tbl_even": "#2C4060",   # Tablo çift satır
    "tbl_head": "#1A2840",   # Tablo başlık
    "select":   "#2E5080",   # Seçili satır
    "menubar":  "#141E2E",   # Menü çubuğu
}
F  = ("Segoe UI", 12)
FB = ("Segoe UI", 12, "bold")
FT = ("Segoe UI", 16, "bold")
FS = ("Segoe UI", 11)
FM = ("Segoe UI", 12)

# ── Sürüm ve Güncelleme ───────────────────────────────────────────
APP_SURUM    = "13.0"
GITHUB_USER  = "celikkismaill"
GITHUB_REPO  = "icra-yardim"
GUNCELLEME_URL = f"https://raw.githubusercontent.com/{GITHUB_USER}/{GITHUB_REPO}/main/icra_app.py"
SURUM_URL      = f"https://raw.githubusercontent.com/{GITHUB_USER}/{GITHUB_REPO}/main/surum.txt"

DURUM_LISTESI = ["Açık","Kesinleşmiş","Haciz","İtirazlı","Kapalı"]
TUR_LISTESI   = ["İdari Para Cezası","Kira","Vekalet Ücreti","Diğer Alacaklar"]
ADMIN_USER    = "ismail"
IMZACILAR     = [{"isim":"Dr. İsmail ÇELİK","unvan":"Şef"},
                 {"isim":"Ayşe BİLGİÇ","unvan":"Avukat"}]

# ══════════════════════════════════════════════════════════════════
# OTOMATİK GÜNCELLEME SİSTEMİ
# ══════════════════════════════════════════════════════════════════
def guncelleme_kontrol(sessiz=False):
    """GitHub'dan yeni sürüm olup olmadığını kontrol et"""
    try:
        import urllib.request, ssl
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        req = urllib.request.urlopen(SURUM_URL, timeout=5, context=ctx)
        uzak_surum = req.read().decode("utf-8").strip()
        return uzak_surum
    except Exception as e:
        if not sessiz:
            messagebox.showwarning("Bağlantı Hatası",
                f"Güncelleme kontrolü yapılamadı.\nİnternet bağlantınızı kontrol edin.\n\nHata: {e}")
        return None

def guncelleme_indir(parent, uzak_surum):
    """Yeni sürümü indir ve uygula"""
    import urllib.request, ssl, shutil, threading, tempfile

    ilerleme_win = tk.Toplevel(parent)
    ilerleme_win.title("Güncelleme İndiriliyor...")
    ilerleme_win.configure(bg=CLR["bg"])
    ilerleme_win.resizable(False, False)
    center_window(ilerleme_win, 420, 220)
    ilerleme_win.grab_set()

    tk.Label(ilerleme_win, text="🔄  Güncelleme İndiriliyor",
             bg=CLR["bg"], fg=CLR["accent"], font=FT).pack(pady=(20,8))
    tk.Label(ilerleme_win, text=f"Sürüm {APP_SURUM} → {uzak_surum}",
             bg=CLR["bg"], fg=CLR["subtext"], font=FS).pack()
    pb = ttk.Progressbar(ilerleme_win, mode="indeterminate", length=300)
    pb.pack(pady=16); pb.start(12)
    durum_lbl = tk.Label(ilerleme_win, text="GitHub'dan indiriliyor...",
                          bg=CLR["bg"], fg=CLR["subtext"], font=FS)
    durum_lbl.pack()

    def indir_thread():
        try:
            durum_lbl.config(text="Yeni sürüm indiriliyor...")
            ctx = ssl.create_default_context()
            ctx.check_hostname = False; ctx.verify_mode = ssl.CERT_NONE
            req = urllib.request.urlopen(GUNCELLEME_URL, timeout=30, context=ctx)
            yeni_kod = req.read()

            durum_lbl.config(text="Dosya kaydediliyor...")

            # EXE yanındaki icra_app.py dosyasını güncelle
            hedef_py = os.path.join(BASE_DIR, "icra_app.py")

            # Yedek al
            if os.path.exists(hedef_py):
                shutil.copy2(hedef_py, hedef_py + ".yedek")

            with open(hedef_py, "wb") as f:
                f.write(yeni_kod)

            # Güncelleme tamamlandı — EXE yeniden oluşturulmalı
            pb.stop()
            try: ilerleme_win.destroy()
            except: pass

            messagebox.showinfo("✅ Güncelleme Tamamlandı",
                f"Sürüm {uzak_surum} başarıyla indirildi!\n\n"
                f"Yeni icra_app.py dosyası şuraya kaydedildi:\n{hedef_py}\n\n"
                "Güncellemenin geçerli olması için:\n"
                "1. CMD açın\n"
                f"2. cd {BASE_DIR}\n"
                "3. python -m PyInstaller --onefile --windowed --name IcraYardim icra_app.py --noconfirm\n"
                "4. Yeni EXE ile setup oluşturun veya direkt dist\\IcraYardim.exe kullanın.")

        except Exception as e:
            pb.stop()
            try: ilerleme_win.destroy()
            except: pass
            messagebox.showerror("Güncelleme Hatası",
                f"Güncelleme sırasında hata oluştu:\n{e}")


def guncelleme_kontrol_ve_goster(parent, sessiz=False):
    """Güncelleme kontrolü yap ve sonucu göster"""
    uzak_surum = guncelleme_kontrol(sessiz=sessiz)
    if not uzak_surum:
        return

    try:
        mevcut_parca = [int(x) for x in APP_SURUM.split(".")]
        uzak_parca   = [int(x) for x in uzak_surum.split(".")]
        yeni_var = uzak_parca > mevcut_parca
    except:
        yeni_var = uzak_surum != APP_SURUM

    if yeni_var:
        if messagebox.askyesno("🔄 Güncelleme Mevcut",
            f"Yeni sürüm bulundu!\n\n"
            f"Mevcut sürüm : {APP_SURUM}\n"
            f"Yeni sürüm   : {uzak_surum}\n\n"
            "Şimdi güncellemek ister misiniz?"):
            guncelleme_indir(parent, uzak_surum)
    elif not sessiz:
        messagebox.showinfo("✅ Güncel",
            f"Program güncel! Sürüm {APP_SURUM}")

# ── İller & ilçeler ──────────────────────────────────────────────
ILLER = {
    "Adana":["Aladağ","Ceyhan","Çukurova","Feke","İmamoğlu","Karaisalı","Karataş","Kozan","Pozantı","Saimbeyli","Sarıçam","Seyhan","Tufanbeyli","Yumurtalık","Yüreğir"],
    "Ankara":["Akyurt","Altındağ","Ayaş","Bala","Beypazarı","Çamlıdere","Çankaya","Çubuk","Elmadağ","Etimesgut","Evren","Gölbaşı","Güdül","Haymana","Kalecik","Kazan","Keçiören","Kızılcahamam","Mamak","Nallıhan","Polatlı","Pursaklar","Sincan","Şereflikoçhisar","Yenimahalle"],
    "Antalya":["Akseki","Aksu","Alanya","Döşemealtı","Elmalı","Finike","Gazipaşa","Gündoğmuş","İbradı","Kaş","Kemer","Kepez","Konyaaltı","Korkuteli","Kumluca","Manavgat","Muratpaşa","Serik"],
    "İstanbul":["Adalar","Arnavutköy","Ataşehir","Avcılar","Bağcılar","Bahçelievler","Bakırköy","Başakşehir","Bayrampaşa","Beşiktaş","Beykoz","Beylikdüzü","Beyoğlu","Büyükçekmece","Çatalca","Çekmeköy","Esenler","Esenyurt","Eyüpsultan","Fatih","Gaziosmanpaşa","Güngören","Kadıköy","Kağıthane","Kartal","Küçükçekmece","Maltepe","Pendik","Sancaktepe","Sarıyer","Silivri","Sultanbeyli","Sultangazi","Şile","Şişli","Tuzla","Ümraniye","Üsküdar","Zeytinburnu"],
    "İzmir":["Aliağa","Balçova","Bayındır","Bayraklı","Bergama","Beydağ","Bornova","Buca","Çeşme","Çiğli","Dikili","Foça","Gaziemir","Güzelbahçe","Karabağlar","Karaburun","Karşıyaka","Kemalpaşa","Kınık","Kiraz","Konak","Menderes","Menemen","Narlıdere","Ödemiş","Seferihisar","Selçuk","Tire","Torbalı","Urla"],
    "Isparta":["Aksu","Atabey","Eğirdir","Gelendost","Gönen","Keçiborlu","Merkez","Senirkent","Sütçüler","Şarkikaraağaç","Uluborlu","Yalvaç","Yenişarbademli"],
    "Konya":["Ahırlı","Akören","Akşehir","Altınekin","Beyşehir","Bozkır","Cihanbeyli","Çeltik","Çumra","Derbent","Derebucak","Doğanhisar","Emirgazi","Ereğli","Güneysinir","Hadim","Halkapınar","Hüyük","Ilgın","Kadınhanı","Karapınar","Karatay","Kulu","Meram","Sarayönü","Selçuklu","Seydişehir","Taşkent","Tuzlukçu","Yalıhüyük","Yunak"],
    "Bursa":["Büyükorhan","Gemlik","Gürsu","Harmancık","İnegöl","İznik","Karacabey","Keles","Kestel","Mudanya","Mustafakemalpaşa","Nilüfer","Orhaneli","Orhangazi","Osmangazi","Yenişehir","Yıldırım"],
    "Burdur":["Ağlasun","Altınyayla","Bucak","Çavdır","Çeltikçi","Gölhisar","Karamanlı","Kemer","Merkez","Tefenni","Yeşilova"],
    "Diğer":["Diğer"],
}

AMBLEM_B64 = "iVBORw0KGgoAAAANSUhEUgAAAHgAAAB4CAIAAAC2BqGFAABJkklEQVR4nO29Z5gcxfE/XtXdM7N5by/fSXfKOSGhhEAgiRxETjbJmBxMNmCDAQMmGjA5m5wxOSMESCAJCQkUUA6XdDlt3pnp7vq92DuhdKcA2N/n//zr1d7tzkzPZ2qqqz5V1Y1EBP+//PYi/tcD6EKyj5+IiGBzVUD8+QcIQAAACEgIiAiY/Qu3Ptv/AcH/QxpNBFoDETAGjG39ZQeqHfASAAJs/aOsaA1aAyIw9n8H9P810ESgNRFphozxTf9W0lWNjbK2XtbUqOZGt66OtcR1KkXpNNMuEQIT2mMZIZ/OyRFFRbyogPfsKUqLjYJiNAzsfB5AhEr9XwD9fwR01iAQAe8A1wZwNqyXy5bJRT+6i39Sq9eq2gaMxUClETQD1KARBAACKABE4AREoAE0AQMAFF4Kh1lhAR/QVwwfaowdI0aP8vbu87NxVAoA/leI/9eBJgKlQIjsVe1ENPP99+6Mr9yvv3WXr6TWVgYugEAwURggOFoGaA3AAJGUBCJEJAKQGjkDzoBxIgVKowaQUksHwCFwCTiEC/iwgebkSdZB+xt7jjXDYZ7VdCmR8/8y3P9FoJUGIhAcAFwpk7Nm2++8m/n0C1i9BkAhGMzwoGUB4yRdIA0EqLR22zWYAIggmRkChqA1cUY+ixJJsjMIDgcveHxAChABGXDOGCOl0XWUk1ZgczB0v77WgdOs448KTN7XMC0AIKmAIW4zGfxG8l8BWmsgTVwoALuqIvXG25mXX1OLlgqwAXzo9QBjpDUC6nScwGWeMBEiKG2Y5o1X+Y86CknFX3pV3nk/goF2hvr1yvnyXdQM0+n02nXJ2+/BeQu4x6+VZIzLdFpDhnMvmh4AAIagATIZopQCIUYNt045wXfSCUZ5HwHANlnw31h+W6BJa9AahSCA2PKlyceelK+8o5s3CvCA1w+MgVagCQCAMXCScPDB4tAp7jV/RzJ0utm6757AuWe1XXKZ1W9Azl+uaTzjHPX88xwsNXpY8aK5sQcfpap1vgsuIr+vdcgeGMswj0clm/mpp3mPOCxxyy24agMaHtAaEJAx4Bw0UDJBkNR5JdbxR/svPCcwcg8EIOkC47+pdv+WfrRSwLliLLN8efu998nX3mGJNi5C3J9PmjqcsE5BxqRKW4MG5F98Se39j7P19QBoDeqP7TH3p5XOx58l33mPV1SjN6TTMVEQIaK2fz3M1v/kxDP5jz3EcvIhVkN2CgcMijz+gOULuAu+s3+6Fy0faA3EtKOUjDJA5gkxM4TRtPv4E60vvJg+/ljrissCo0bxzgH/RmD8Ns9Qa9IaOE/X1zZdflXzhMn66WeESyxQQIKTVJtDnBXSioPf/fwLpbUxZk+lk8wIRK++gdqjPeZ+02NjVe7D9/P+fbW0AYCHQhyx5P2XC+bNynnsgdS8uaq2lptCyUzo+SfsNWvjixdbo0YSEBIA5+S00dGHhN5+gx86TckUxJrAEMyfJ5RpP/9i26T9Gy6+NFldTZyT1jrrnPza8usDTVICYy7DpkcebR432fnXvYaDPFiEGiCdga6megIQhlpfQfGoMXkvBTaaHr5uff2U/RvHTGw940yek+N/9kkUAkBDYQFplfpwhrtgWezGfySOOYUDyHSrcdklnj3Htk7ZHxZ8L0YM04BAhASKRPjS87xHHRF65aWCtYvp7D9KJ01SERL68wyNzsMPN4+b3Hj/g44mxjlI2Rke/WryawJNWmtXaSFiS35oOPDQ5EUX8dpm7i9EjiperwVXkbBKR0kr2BZvIjAFpNrtBYu8kydx9Kpks3XP7QVzZzkVGxLPP5v5cTGEfcCZBmn06KmUSlx9TfJPF8qbb8K2KNlpPXhE7t23K0OE77hDDx+OubncEwEiyiTYiFGefSY3HH1yS1mf+MtvFj75mHnK8eTEgXFwHJLSCBWLlmjmsssbpx0c/X6+EkJnjduvJ78a0CQlMqYM3nz/fS377M9nfGn4Csk0USrppszr/hqpWlJUv9r37FPEN4N5MwcLkQGQ+9UcMXQwFRZzNNL/fomHc0rqK4pbWnxHHpa4+gadSnM0lSZZuV74crivEAOFiOhEfLkfviErK2LnXcx7lJo9e7JQDhYWMKUV2b7TT2KA3oOn8vzC5P0PKgBjyBBNNiNS4YDODatYPVkeESxis2e37Xtwyx13ScaIMVLy18LnV/I6pJRCZOrqYhdd7Lz9tmHmkiGycQGlWo0bb8i56a9tl1zOCop0POk+9AhHi5QCQLAzBC6zggQEnEEqDlOn5s38qGXqoTRrFgHTPXp6Jk9AhnLhQr1yA/oCQpPrEeAxMJZGAEJE11E9SrzXXGq/9Iac9TkCsGmHFH/xcePUA/RXsygnr2D5ovTSZaIgIoYOp1RKxeJtBx8p1ldpN0EXXpBz+83Rcy6Qr79uMD/4g+DY0m5hhx0aeezxYFkZSJX1/X+h/ApAk5QgROybOdE//AHXbWCBfJISiAARlFYGz6tdlfrn/alb/87AAOAILoHJvAGQDgzsr8t7ss+/Qm4QAihFOQE2Zg/8YZlubScglHENCjtYJItAAyCBBiAEyP6bdRhUhRjEcBhcW0aC4X8/knzqOf36K+LCS/Ifvr9+yCgnmSitWtd0zXXw9JO83dGmiVqR32NeeF7o5uuT772X+NNVrLoafTkcmUq26LKy3Oee8k2dylwX+C91/n6Ze0cERCREyzPPJS+61ExLChSQdH8mNhWQl3NTaM4AuREplcmYcell2k6rh58mmTTPONk67ffRfsNNQmIMNUFTk/z0fQZce3OwtAj6lVu9+mLPYpaXxyMR9HnQFAigpaJkWja36ZZmqKrVNVWqplZW17L2BgTiyXj0oKNEOFcLDxQXx76cqdeuZcKAxvpAv97xliYIlKBjg2Gw1trMjwsDpK1hQ62VS9pOP1O98x4zQ+jP5RubWw86KvXg3Xnnn8dcifiLCNhfoNFEpLXivOnmm50bb2FmCE2LUkkgQsvqwJoJlW4KffCOOXFCc59hPB53wS6orFDr17VP3V948ignTKEAVFaDnQFwlD9HjBhq7LOPMX6cd+QQUV4OXl9WYTep0+b3qjfjTlUmpapq1NIV8Tlz9HcL1dIlFGtmYBB4KOjhxHWizvvpJ949xrb26M24BYyBdFQ4ULh+aWbe/OaDDgtcfZVavAI+nwGeAAACY+g4yokZ116bf/stXEkEhnw39Xp3gc7qMmO1f7pEPvSg4SnQglOiVffvh6afrV4NwgQi4AwyKTVyWMmiOfacuW2XX2FOnJzzwD2tZ/xRv/gyBPMh2q4gSd5cc58J4shDfQfuLwYOAUSxxbu21Qg356U3wd7xwe1kq+W61ZnPZ9rvfqjmzodoMzI/mhyHDsU/nOjceh9vaUfLI1NN/uefC518Ut2QPfiGKq1tASb6Q0pLSkcROHgDgKBSTcY55+c/9rCRTTzg7mC9W0ATAZHLWMu5F9pPPsr9xYBAiSYYMzbny08yTz2XufJK9OeBVIDIhKESbXjkkZGH72E9S6m9ve32u/W/HmQOKEhTv8HmqSf6jz/WP3xEJ5FPoCUAKCUbHF2V1lUZVZuxm22IuZCQrgKuQRtIASH8BhYZutRr9vKYZT6j2ETGGRADxqkT8fSKFbE33pEvvAxrVyMoAgO9IUTSqQSffkjhe2+1XnKV/eB9zF8AyIg0JKMKXPP4E6gtqr+ahaYPkOlUIzv9tIJ/P20AIu4OFbXrQBORJuKs8dwL7ScfFf5i0FKlm/kRxxS88WL05TfSF17CmUVaAyI5LukEmhF0EtIX5P17U0MjNlQScBg5ynfRhYGTj+OhHAAwCAAJSK9P2PNaU9/H5E+xTHVGt9rkEGapPwDg2MkmE2kCQK0JBHCTU0hQL485NCz2yvGMz/H18RvAOBC4iBqAEvH462+mH30Cvv+BgckCIcduzZn/DbrUNmEfwxtCZCoZA5A0dXL41hs9kya70fbWURNZdR1YFjCmkw381D8UPP9vQ2tAhmzX7PWuAk2gtOK8/pJL5YMPiEAxSanB8Vxxaejvf4O0k5r7bfyoE7lmIAxwbSou4KeeqJ55BVujjHHKpBQkYeBI358v9v3+FMvn4wCgJYCqTKkPG5MzmtJLok6zAk7MQDIYGuznJCBtGsNm4+94+ABKg0MkiYAoItjgID+4IHBokbev3wRkgFwCKNuJvvxS+q4HaOVSk3vksKEqY5vr1xMJV7WLsRP9V1/uP/7Y1OdfYiRkBMNNo/cyiHeYKsNQiUZ+9nkFTz7CHckMjrsyN+4a0CSlFqLplludG25g/kJQCpV2vUZw5nuZ+x5zVqzosXBu8u132o892fTnUCqBE/cMvv9G7PCT2KJFUmrtNz2XXhS54nKRG1FKmYyA4MvmxMsbE183pxttMjjzCS6IAIGyJmpn7gEAADWS0syWigBMgUAsLWXEgqkRz6llOVMLPEAkgSHnbiwWve++9D8f4olmbuVqAF1W4rvzlsCxxwiA9nv/1Xrl5T1aWxMPPZy54W/cXwSyI2xBQ7iJBuua6wrvuJVJiWIXfLZdAVop4Lzp2edTZ55leCIKCLPOspSSa+FqrVN4/Iklb7zSet+/UldcbngKSTsq6OfRjJYx3HevwL3/9O85jgNw0qDp48b441Xtc1ulozFooImogXYS3M1RlkQuUYSzAgNP7ukPGuLB9e0rUlRgIiFGHSUIx0XEOb28R5bmMmQuMURILF4Sv/Jq+cVMU3hln/LQY/fb/3lHWmb4xuucmbP9RxzSPGYv+GkVeHxbsIxcyFSj7+FHCi48H6SEncZ6Z4EmKZUQiW/ntB9wCM+mPDYdiIAExBgxJpMN1lV/Kbz7tuYbb7RvvYsFclgs4Qrlue7aguv/qoXgjstMPq81/s81bTNbpQD0C8aA9C7i23FlBABQRPcMzplSELAECxkGAPzYHn9oQ3R+m1OdVoUe4WhKS3I1Tc63rukX3ivfB47SpqVJtdxxZ/qm24ST5f8VgsNOOb3wxefS33zbtt+B3AgA6i1tFQOQSrrBz97NmTqN7TSzulNAk9YAmGhqbttrb1FVp00f6O1wiQhA3JCpRs8Tj3Fupi64RDiO7FHof/qxyMGHgtKcQ3PG/ueq9ufrUlKjXwBiB++/G4IAGqHNUScUWk+PKwXgAKgICEggAeDaePrONc3v1im/oRkiAEZdZSD9vjRwzcDcQq+ppAbBYzO/iP3xfFZZizm5lEiokryCdctj115n33uv8BeCdLe+MOPgpp2inMLvvg2UlhHpnXFCdgJoIiWVEqzp6JPUe+9yfx5JZzv022andF1pIAMZU+MmRF59PtS3v3ZdZrBPG2LXL29fm9I5JmNA6hcE/wggCQosum9owdiI12MYSISd05Mm0ASCAZH8/fyar9uUyRAAOZIGaHdUH4v9fWhkemmIXNCGSFZWtJ5yOv/2ezI4239q+PWX2vacyNZWgcezXUVAIXSyDQ44sPCTtw2NTIhu8MjKjh8FaQ2GaHngAfe9N3kgQtLtDmUAILQ8fpJRmn5E4ecf+fr2ByWJw60rGk77obUmo/MsJPqlKJsM2lwqFmy/woBHCA7ANnMCGIJgoAlczV4ZX3Jgnoor4EiKgAjyTdag2B9+aL5hRZPLkUvX7NW76LOP6ejp5Kb1l3PapxzE6ltBmF29biQl8+frGR+333WXNgQptUPDtyON1hoYiy9f0jxhP9PlhAg7fAOEoZNN/Njj8l99nhuWIN1i2xcurvukUUU8ImuOf4lwhLSGdlcflO89q9x7SHGOYF36tIqIgzr+m3mzUj2C7Oeny5AQWJMjp0XYQ6NKe3oth0CRbP/j+e4LLzAeAM53oKQIDNAmN/LNF6Gxe4KSjHc3Me5Ao4nIUW7bRVeZiTQJvkOUURiUbBLHHpv/2guMm4LU6qR91Py6z1qpwCM6oozdEuoIf7HF1mOC7NLevj/3taaX5rJuvVkEAGCHlPaNOi4HzhCz19eEiqjI5LPb6LgFG5fHMiYSB573/NN4+ulaJcg0djggjUzYTvTCy13bxqxD2s1IutNopTTnrU88kTjvAsNXSGqbaWGrKwtTJxvZEUcU/Od1ZMIUbFks9bsFtfUOyxHM/WV8LAcgIAI8q8x/Wd9IwBJaZw1xd6Ralg2JOpnLltR/1ORaKDjbYhwCMaF0xMAXRuePjfhtl4Bj84mn6Lf+g/586H42AkAhKNlg3nt/4eWXdJ/b7RLorKeRbmpoGj1JNLWSYXTvH2jBeTKmxo4umPmxGQhzkkuj6ZMWNba6EOAofyHpTeiSbpf6qZE5vysvIAJJIHaBtiQAfLmq9a8ro51B5s8D4ggpDUEuXxpTNC7H7xAjN91w0HQ2ew76IjtQL0SmpRMI5i2a5S/vxbIVmtuTLk0Haq0Ytt9+D9RVkrX9yXez0zB0Mk5JQfj150UgxEGvS8lTFta3uhDg7JegjAAuQaEX/z4wcHQhP6I4qAgIwdiVCjoClAS/L4/c1M/TYmf4lgcqAj+DuDRPW9S0IpExUTPLn//q87pXOWQSwBl0YxSIyLB4a238H3dq7M48dPGd1oAYX72qZewkw0FirDvrjIDAHDcW+uS93AMOIum2SHncdw0rUiosfqkuM4CUopFB9tmknq7WBjOJ7Q79rokIsDEZO25hc4XDrG0mC4GQkNDLA+9OKC6yLMV5bPbs+EGHc/ACdVuAgMAIXXRz5nwVHr1H1n3Y7o1sR4hII0bvvIclYmAYO5gDOZeZFs9Nf8s54CDtuBrpwiW1S5MqR/wiXd50GwbHhNLtUnFubS9/vlPCEAGoJBC6vBwdOyO24ZRdgoCANRm6cHG9A0SOG5w82br5Jp1pgW7dCSDQgmEmkbzjbg1dOlXbA1ppYiy+crl8/W1hhkl2+zw5h1QcJu9T8JerwZXcxNtXN33WCIUGd35xNhIBCKjJViGDW4gEmu8iObmVEEBTJtquXBdgKwOSjYDyBH7Rom9e2WCYnEuV/+dL8MADdKq9+zibpDI8ufLd92OLFgJn261T2B7QRBox9dBjPNlOhtGdhUJArZXHynngX8QNLtgXDfEHK9K5Jjpa/xJIEEAgMGQAeGm59cSoQq9h4S8rjsjSmif3Hf7Gnj3LTdnmZl/xLYbpEuRb7NHK9Pu1UcYRSOQ8+IAOhVHLbuYEBNCco51KPfg4IW4XsW2GrjVwnqmtcl55mxkh6r4+ihky02JdcYl3j9HoOM2Oc83yVoEMsSNHvXsiEBVBTEJl2t4/37xzVFm53wsAv0ybOyTXEsf0jDw3pmRSDncUMtwaFCLycfzr8ta6dJpJxztooO/aK1UmCqx7pXbRDMu33nUq1gHj2yr1tkATIKRe/Q+11oJpdmedGSM7Qb0HR66+kmsSJr9nZdvaFPkE7nZUwhAQsEXKoIDpReLhkZFbB+Urgl8Sr28lBJBRuk/Qf2EvI+k6uI3NV4AWwxqbbl/dygyDKRW+7CIYNhIyiR1U95omxZpiz70iAbYt4NvqSALOlOs6L7wumJe6rYlCZFqlA9dfLkIhjjS/JflMbTLX5HK3YObIACApIaPlFX1Dn0wofXpMz/N6Fxb7PByR4w7irp0XBDAZEkCpZfi4xs73hAFwhGzziyKImPz1OvubphhDZF5/8PprtXZ2YLuU5Myffvl1J5VAwbcyIFseqTQhxr/9Ri9eDJavm+IzZIwyCRi+h/+UU5TSkty71rZo4LuHB2cs4SoDYJCf/Xt0/t8GFZQFPIqYoixR3dHkBtDJy2X7Xzb/vCvCEDXByLzIkYWeymSaAQrENGG7ROrUcATSgHeuiboghabQCceycRMpE4Vuyg00geWD1asyX32NyEBtgd52uswyb76L5HT/mhAy0o7v0otNj89EmlGfmNkq/XwXODkEIsYMBI5Yl8mcXGp+sXfxjL17HFoUcjVoIp6lDxgSonKlTrtSac2QGGpEjQgMIRt/IxJR9+/fNlcHhuKKAXnXDwwDuPWOHOTFYwsNWylJmFXqoMBv2tyPahPINHJhXX6eJrkDpWbAgOzX31EdF9nsij8HLESAaCcTzSP3oopKsKwuo0GGYDuqd1nx4jlWIKS0e8x3NfPaKSh21pgaiBmtpaYMoYfR3wcFzyrP5dyAbI9VZ8ZVAjiratwlG6G+HQHIEFgcMnrmQX4IGELKBqnIY/Kwj0V8hrF7VVf0fWv0marYNf3zywPmsrbUdauav2snH0cASCgYE8T3J5Qxxhw73bTnZLZqDVjeLt91RHAcVVJU8tN8M5zTURcHAFuUqWRLx+cvUBvWMk+wu6JVzkEnPKeegIEQKPVNS/K7dukXhqId6xQCIFJ1Ru+dy28bXPBuQzzI8dw+hVqjImLYmcXXBAwzP663564yJafWJGrSUlEi47TEyGtCPAMNcZVIc8uAgAVhr7n/SO/AEtC084UAWdszNjdnbG4YCBwNwyPBvrzmC2UEuXCIggIXRuXM5vjBBUHT6/ee+fv01Vdz7uvyRomY5ZUbK1PffisOOxy1xk4HXGz2G0IA+/MZmlzOWHdvoiNdf074dycgEYD74saEqznbbnZrG5QJQLlwTg/vxf3Cg8L+CXk+ILKVNjljm79riEDECkLmYWPV+4uQtO5fZE0bBkEPxjPgSPRbFPCqiiZ33mqoacW6eOrj782yg4TX2qKAqVthCAxRAwAgQxAIUqupRd5PojolgfFsHQN/sTpxcJGfEQVOOj75j3tYyulmMtIMGKjMp18EDzscN5s8NgOaMUXa/XoOA0NT1+EG55SKGVP29w8eyoAq0+rLpkzAYDs0GhyQkKKO86+huaf3yQPgShMgQwBrW7uHAIq8PfLj735PFY18v6HmQaOcuSvkT1XcY1HYB0TkSjF5qP+saekvltC3a0RSqnhGKw2WaRq7UGjLNvtAyI4s7+MxWk9b1OLnQhEFBH7TmlkTlwNCplHe25i6j37nPTRyoKsIQ2sEj/v1HK0k3yye7LwKETKWqapQP61khrc7TxhRgWsec1j2sI8bU80umDvB8yS0qnX0JX2Cp/fJdRRqAs6QZ92r7dX/E2caAIEg6DWmDYe0jUGf//Rp1mlTjHEDxORh5viBmX9/kZm3yrf/SOiVjxmJXis9c4n9/nwNsHs1hQxREw30U9iwFSECMIR2l3/YkABQksA85ggC1V2ygTQall67zl23FhA3WeBOoLUmAPXjMoi2YbcsEjou+HJ906YhABF81pQyGegdeXWaYFyY3Tco9LdBJZqEwbvOPgEAECGk61rd1pgYVUaFfsNrihw/M4Wsb9XRVOaRz5ynv1SNcbNXsZ7xE7nS6FesDY4m53FbLtiQ2tAIO5N1296lGaJWdsZJZ4sZCbTJ6NPGtJZSIPimTqVwIThulxE5ARoGJNvTCxdnbyX77800GsBesJBB1r3pQjgjNy1GDTH6D0CiilR6cVR5hdjhHXGke4cWXtC/kDCbr+v6RomA0E1m7FfnOO8t5L2LvQePBqDkzCXOC7Ocj37kuQGx72BIptQ7C/T6JqWkJtItMdYzIm2pa9pBCLViIxHsXiUDEQjGh4Z8ijALjEew5Ul3VUoKrc2ycjF2lHZT3RAChICg3EULO04HAD8DzRgAuMuWE3SXGERkGhxz772BcdDuvOZUm9Im7lhz2l0VlVoRMMZ2lEQHiWDPXkEJR1e0OjOXWX2L3UTa/WYlCwUx7qTe+NbYd6jnvIOMg0eIkT3EwSNJUWbdRuOgEWrOKp2ymcFZxgGE7oKLrm4QUAOVh4vuHdnDVXFCJAADIebqeS0pQBAAxj6TCNxurQcx8OjFPxHApnAkWyhLwJiWjl69BqFbfgMAgFuTJnIAQJgfV6RRdft7BqAJr+sXHh3xI9EOk8HIUDbH1Y8bjAOGcb+pZq+KffA9ekzvyXtrD0ep9JqG5GOfOPNXsx75nuP3Ck4czGzH97t9VXvanb8WAx6VcbBPoZtxUhsaNMCuGhCOoIiKPTjCjxkFDICAEPmCtkw2A2BNGE8gujOWRIiGXlfhpJPQmTPpBBpANjRCfSOKrg00AjlS+YLG8EECwFV6WTRt8e7UkwPEFBWadNXAPMF2lMCHjlyeu64eomnRu5CGl5FUbEFF6rkvWU7Ae9HBNKgIUq5ZXmgOKQOLZ+MBHvRSVat6az7ToGIpY58h5ug+9n/mud+s3NH1tn+fHDHXEz66JJCSTjZxbjJamnAdTQhgDB1EoTC5TpdmWhMZQjc108Z6gg54f3bv3NpaFY0bhqfL+RoZuWneq9woLweAJldtTGsDuzQcCJAhPcDP/jk0XxGyjjRH9/cIAAANUaZIN7b7Dt0jFkvQ8jpeF888OdOYNtT7u33cPqvdOStdn2mU5MmalanmdlpTB00pCllYFvZMHioG9sh8uAh+qlV98zaPzXZSNFFzOrqhbfUP9S2WGKSJCMBEXpvRdY7qJUD06CF6ldHSlV0rJSHnlEjaVTWe/v2yui+gM1RR1TWgHRA+6Cqlgojgsj5laHkBoCpNrS56+fbfIQQwGMZsvLJvaO/8kNyVPACmHTANe/Zyozw/dNI+qU8Xq3mrmavdT35019RZ0/f0DO7pfrnU/nYFaOQeTv1L+P6FZnkhFQRkRWPyyc/Yxij4PZC0yZFg7ahCA0CRRkCG6GptMPbEioU3r4kWhHr7Ta61BECDQULihqTs5QMmDNG3j7t0CaC/yzMyBmTr2o0agJEGYAKgQ7epoZlBt/2LiARS9OqbrbivSMuMpsD20q8CMaWoXUogp8TK1QQ7W7NNAAja5CA4a7cTz8y0pgz3HjzKHdbD+WIxbGhlK+rTG2dY+w3zHDMRAUgpxrkG0PG0s6FJfvC9rmhkjIPHBKm57szgdRErUucXWZLWJY0AS+rXPlebLincj6tEFmUAAISMVlUpt4MSKi8DUDuqJ9Gqrn7TnLSZ6ajfiB3tT9vX6GwGj5UWZ//cmMoQKQIT4OdgPRvU1mfcA3Ot43vklPmNCblehrCTQXHWs8T8kFYavaZIKvf9H9wfK60DR/rOmOZuaJLz18Cqevet+XLeavOY8UafwthHC2BJNRKjlIOcM48FAKAJATRnIBhs7+IElOX9sn+9tWZWv5yyUQV9AUGhJ4ZlfhXV9PPzQQIg2JhRoDUyjqUlGkh0neXrIBvq6zR0cNyi88Kgo/EdwqEBoLAw+7nZJWJbFEIxxJTSROr4YuuuYUWFXs/OgPvzzROl61t5QY7oXyy/WoGaSCAyA6vaMs9+BaU55she1uFj6AApV27UCyszT3yujpvg3XuoXR/DmigGPPpn8hq1lDo/wAyxrTpnUU5lovXJ5nxf8KMNP562NNk7JO4ZmOjh1c+sWy3YENpS4RCAIzbYLoAm4Ly0BLv1gwkAATGR6ji4A2hEDaASKQDWXeRKBMBEfm522EmJfLOSHwZoKz3Cz/86uHBqQRAAXE0cd2ICzIomZKiWVLveBv++Q+0+BbCmgbwGU6Q9ghFATdSpWARfLhcje4nx/cXeg53FlZmPFuqM6//j/sl3vtPfV2DA+3OWQGmjVyED2Go+JCDSkLLbTvvm3e+Sefks3axD5fkTbU1nr6hE2SD5gJCVo2mLLJcGQsSoqwCIAYhImLonpokQUMXimC176QCaAAEgndkJPBC9ZtZGJ1zFO18IBuASBIV+eVzPAo+lABiAsYvJVAJgliG/XanH9TcPHGFXN3FFmmMH8WIKZglwNc1Zk/l+PR9Uau43xH/lkc6KKp20vYePTaclrKrLTn3kKswLGKP6SEdyYwsPNMuxPLF68YfJoWWh3k3KNQCUSnGAgNUTvL1Qu4q2ziUSAENIKAZEDAAtk3b09iMwlUpTJ9AdJkwDMOkAUPe5KATknV0bacqaruwX6Gg1yGcGDKGyQ9kFhDedHMBvYXsq9clCoyjHPGKMcl1GnQ0U2dwVQ/KZnDFaVm0/8nnm/YVm72Lht0yD+47cUwUspgkQwVX8gGHMYyS+Xa635jwQgMoiwwJmHqi4RRmkDAASoCZXq7Qi2dXwNWnIti8yg+3oFgmAOZlN01PnEg7dVW9scfC2fzGEhNL7RMxXxhZZjP2Sdc7QY5LXhKU19jfLPSN68+ljpOswrX4mFghAEwGA10TLgMWV6Yc+SX61XLmK+T2if4l2JaVtNrLMHN5LVjVhXTvBFkhnmeTDisPDQywhEXErrdgel/jzDVPnh52iUZjCTX7Cz0CDIQh2qIq0iYf1Zr0QQARwNZV6uM+0duIMXQgCAPD8IJgcDcv9Ymly3orAnv2skyYpEyljb90noomI0GcxZGrG0viLX8m0bfQuUMkMleV5Dh8NWjtfLEafyQC2iqk0oc80bxwY0aR2KUDnmF04D5hWO9HcROT1bgIjW6tDAECWtRPEIqlUOvvJx7ki1Zmrx0wH/rurzYhAwPKCLOxDpbhp6U+XpWcuNQf18J57EPQqUMn0dnqDNQEAhry4qkEtrYY8PyvP8588CTxm+r35am1dx3y45TvGERTBtMLQmeW+NkliJ4bMAInAx3T2oaHj7KjuHACIfBYA/Mx1ZCNDEfDvAGVEBJIt7dncj9/K8oGgiUzGF7bH6pNx1sX1NWlFWmVXz+1KSKMpcFRvSjvEGJqm/dXS9KuzmSH8Z04V4/po1wXbBURim81ViKiILJNFgmZxjvfiQ0jr5Cuz+MINojjCB5Zmf7M1cAhKw5/75fbyoK13ytYpolwLAZEAnLY2hO4TpEhAIujfZJMZdA6E5eR1b6E60Ghuzn4o7KD5iIBxdP86sCDf66dtJuOsZ8uQcWQckbrpMUDGAXxj++vyCCRtYMh9flpRH3/8c3dplefI8d7zD2QDCsmWmHTIVZB99IogmWGT+hkDisgQ7vKa1NNfGquaHC74tBHC74Ftx5SdkxDzvd6r+4cSSu3YPyIipFzLAOQA4NY1AujuiPvsrQdD1Gm4BHRCy0rys+uXdaVy2WBWbazVAByg1GNw4AiQJhgdso7vkaORbTZpEQJKrQTjALChdc1bFStaXOsvIyYGfWFN26tyRkAC7jG8x0/KvDYH69rBa7KAF5Ju5q15zuL13qkjrd9PNurb3e/XOxsaWDStlGYRnzhspDWmr1PRaH+9jFbWoTBUrteaNsI7qjf+XL2wtQhERXRyj5z/1MW/bqWw6JbvZcgIygwTgGkAqK/vHuPst7ygADrjfwHQMaHxwiLdve9BxECwisqsS1jutUxOgNx1nTPLggo73D5NGrHj3RaMp9Otl8798L12y2Y9HCNn3pwf3p40JugJ0HbpDwQk8uQF+B+mpGcuU0sqIZYCQwjDhHUtqYqvcECxNaGfecQYE0A3xWQ8Y5XmkMnjMxfrxZVMMNyjN+9VaAwptcJ+pC5R/hlALv42IH/evHpN3eVztSaDsXK/2fHn+goE0d2MRkSArEfRpnYBAZ0WjPcoZWh1V85BmqHlVFQqJ81Nby+vCAlwSAvG57ekTuoJUoOrtcEZaLc13Rwy/V9vXHbbT8u/12Nzwz0CYDPUC9Lex6tarh4cVLR1hXIn1ghEwmcGjhgj9xrgLq1y19dTc4IkkCPpx8rk0ipRGjGG9BBDe3r7FjIADRQaNxD3GUamgE2OVNe6vEk4oqv1HrmBs3r5HqhM5Rt8u5XzCKAA/YL6eDkAgpKqohLA6C6zpEkzD+tZQgDZCVwAdNTDih4lGA5jwu6yzU2DNkxZ26Cra6Ffv0IDyzzip4QKGvzfNale/MdLhg4Bbn1TveC2n5ZWyJwIk6syBvqnFHBTyqhE5IgWo69a4WrSiF1rECICIJGZFzSnDNP7DZXtKd0Uc+vbvLUJ1haPpRJy7krnh3Xg96DfMvYa5O9X0nFstmmfITCkrF51yxwKxjTRJf1zP2qU9Y4yttdDggAuQZnFSz0cAHRdna6q46LbVBRp7vWKsrKfL9R5KhBFJVhUqNvXo9HVS0HAGUtE7WXLeb9+JufDAuaiaDLAMcf03FRhr43Pz+N1D1Zr8O7nsTxtRMEwR5WR2u3IKBNyUEcVBrPO3I46JjGLGkM0In6I+D0DSx5d9iXLyHNHHq6BUIOKJlVDm+6IHLPLDwEBMMo20u6YnUUADZBvef7aP3L24vqIZWzbBIsItqsHBbzZdJK9YhW0taIv2GXxOEPK2FDa0yjtidBhMTpIJdCaGSYb0F+vWgXo63JYjDFQmTlzraOmG8gm5ZgvbkwSIJGKmHkvt3mk8kZCJQwcTWkTUCsiYJ2cL0QlTYqY5/UvdBRwthMFRZuITE0IUBtr/dvcV1zlTuk3clCkhAAMT45RlEMAUmvBstXVnccStWRSbemElxs9QpFuwlWOqAiOKfO91+h9vzGTI/hWxUDZjO2kiAHICMD9bj6C210mH5kmB/v2Nn3+TRask4/WhAzEqGGZD94RuE0V/KbBaw1guHO/Q9KAbHyePyzaJGkGqMgOGiY3i6RKK2QArJNG6zwWQDBKSp2w4wErAMC273ts97oIiNjsJiQ45cH8PCsAHZ4laU2cc8FY3Mmsba37qbV6ebS2Mtpck2xtTrdXtDfdu+8Z5408QJNm21t0ioA0ZbVBXNg79EmzK4m2YpQUQICrvXJ9oMHlkJn9TfccKQIAKHPEcAAArbLZFbHpSwQQY8bsoCyViBk++eNSVVFBffr28YuhQfP7qPQLyDrIkpyuFtHSBF6EFXF96Kz5PXnjAaX9zxo8XpHOuvQEHVQCR9wW/aytXdVWF03Hj++zV74vqImIiDPGOcyrW/3aqjmz6lZWJZvjrg0IApmFPKPlrZNOPHfkAUprvk0VsiJNBIIxjgiggOT4iDnCk1hmhzyoN8HIEJJS7xEyBgZM4sytrdYLljDh64btyNaDmONHA/wcK3UCzRgBWKNHJIJ5YLvQVWMhEZgGxZvSX8zwnH0uY/yQAv+ctvYgsM5ipe40lAAtgev0gKVun5nrkuOLYiMi4Z+/7jxUETDYwv3Lnro60QxKTSjsDwBSK5OLyvaGv81984OahRnpeg3T4qJABLJmvSWTPGfwAZePma6JtkVZd6yDBfFUyydVi16tqkcz7NXR1el+lpEDm1WLMkCH1IGFfiEMCeB8NZva6sGfv2n9n20FXUWeEB81cntAIyJpq7wXGzqIvlsAvkBXpc5EJMC03/lQn30uIzy82Lq3gsmdpmaIyIM6aPImO/jGxpYgNGe0rku2VyXjtlK1yczJA/YYnNsTABUBkEbETW31a9pqheEdXdSHNJlczK5ZdcYXD9Zn2nOMgN/wKNJEJEkjoqNViTf32j2P0tsLCyUpgfzr6u8/rV77Vav7o11keccJZRFJv2kyUJsrqwQKc35oYSAbeWTe/YBBt+2tjEE6zUcONvr37/hzC6ABQGlDCHPy3vZ33/KuzTQoDYZPzZ6bXrvO6te3j9/aNyI+aHRydqKgNCsa0NEUtsRLNc3Pr61wQNhkKhYmQmBFr7RsmBxcdP6A4WMLegD3QIdDjJpoaWt1eaigX7gYGc6pWXXiJ/e4pPKtkKuV0nrzXE+bkzl9yL7FwbCrlbGlH6mBBPJ7fnjv1rVp7R0eMAMlPg+oNAEhCEVSb/ZkOEJM0rQ8MTToBUBZW+18PlMYgW5qmpExBWmx9yTTtDZfQHazdwoRAKyDphCIbivKCQ1Bsdb0a28SIqI4rTwHdrF6QgOAli7mUmCCGRgfDo0sCJQVBHoU+MIZ0fc/8eHHLKj+/eyZr6z69tVVs7JLyzQl42ujDaNyy3MsX02s6eyZjzigvMKSWjNkkn6u8CQgDuyQPqMJtp5stdZIuLBmyb2VGI5MK/QETSQlk4pIEyjajiOkFZzSM4ikJEDytbeprZG6bVYjIgL0H3rgVj2MmwHNGAB4Jk6g3r3QznSXSydizJt5/mVMpwnY1Hz/hLCIu3rnU1fYkRzSqDOgE1qlXWVL7UitEOxcA4W3/4zMkLOWq1s2BBrtBACsitY3JtsnlQxAhMtmvVSZavULi5AcLeNOOmR4O8+MrpYl/pwR+eXZUo0tRg4oVeaONbX1OIhRwtWStgduBx4ISUl7hNlBRX6NQtmZ1DMvctZdD1W2tQJKe1p7T0CAzfuAttRopUQw7D1oqtapbtuPNHq8tHp5/M23FAIHOK9XwNVbe0U7JwgdjvYm4hAVAWk7wKks3G+jjrxblyCA+Q1rSMtJJUM+q1z8btWCAm/YJdmSjuVZ/sennP3KQZdrDQCUNdB9QvkF3gAAbDUkZGAIzw1jpp3R02MhuN3ybwwgQ3R+7xwPN5Cx+Acf6KU/MMvfbdMJI5ngU/e1cgtAbVH4sQWa2aoO87hjFArstrSICDgz0/c9oh1XEh1emjM5YsbkLij1jgQ1gVQZTurHuESgT6uXFYULSny5N3z3eo7X22anlKJzhx78xdF/+92gffIsf7ZIjiHaSg4MlSDitj012UTfiKD49549r+3vc6hLKpohxhWMDYkjS4Nao9I6c+/DHA3qevICANSgAD0nHLOdE27xO8aQKLDvfjhkmLYT3ZEyWoMnoH/4Lvraa8CF4PzqAblsF7JCOyWa0CPYwlb788qVy5vWFvnC/1jw1qr22mTGHpfX773Drr1/v9N7BvM10NK2qnYnKZADACMYW9i3u9MypogdVxIu5Y7TFdQEStO1A3IsAMkw+u7bNGcueENbdQ9ucQRDZadF34He/afBNiukbGkfEFEp4fF4Tj2WdKqjYKzLExPjvsxtd6tkQirau8B3Yqm31Vbi18VbK2SBWfVrYjLTkkm8vn6Oq/RfRh/1ydHXTiztl3YdAHBc+eAPH5ucE4Cr3QJv+OBeI6FTf7cVBsAQvAbv5RWO3s5kJBDaJB1XbB5Q5NeE2k7Hb7qDMdFVDVfHaRlHnbROOtYTCG1lN2A7veCMIYDvlN9TTgk6dndWV2u0/LDyp9i/HgTOlKuuHxQp82FGsW6cw10ShpDSODKEtbENhEyR9grzyann/m3i8chQauU1zNZ04g8zHvquaU1AeDjDpkzy5P6TSgJ5inQ3lJIi4MwYl+N1tNzqeSCAraHEousH5ZJCJnjisafYkh/B02VskT2MXKn9Ef8fT+Od1OgW97LNzTFSylPe2zh+unLj1O3am6Qls3KSd92bXrUSuVHk9dw2OC/puOzX2wyDMVZuJH5qrebccKX77LQLfzd4b1u6HJlg/JVV3059++aPqn4MewIcsSUV37941HVjj9Gkd8SiEAEFoA22seOcYVK6tw7KLfP7JLLEhvXJv99uGCHqvr2Pc3Di5vRDrf4DtrsIzfYQQRREgT+dRx4vuro7pSYiznks3v6nK4mRlHp6cfiivsFmR+1ManmH4pIusgyPbliXbAWtn5p6waF99gAASxjz6tae9MED53z11MZMa64noLWsS8UO6DHqtUMuDnq8m/ZF7UqyFY71jr2VbycQmm11Rpnv2J5hJTUx1n7JldjWrMxu8ykAqLQ0DN8VF3elYtv5f3ZKDI0czY8+guy2HSxyqhT3ReTnHzfe/S8UXCt949CiaTlGmwTjl2EtEJKuHhs2q9qWNycTZw6cnGP5oun43+e+Ne3Nm6a+fdPrVXM5A8d1G5JtjtJXjZz++uGXR3x+nS2+7FoIiCEk7dZFLS0e4dGd+WyBEJM4MWLcMrRYZhQI3v7oo+qD95kvr8ua8axwoTNRcfghwXETQKntttF3uXiVRowtW9I+YT9DGzsoFkFkRC45oRkf5Uzeh1y33pZHLairsinAcbeX2mAAtsa/9lW3fX3b+OLBwwsHPd6U99zw/JdWfFybbA97QwIZMAxwa2Bu6fReew7P75nl3HZIvSoijji7duXR37fm+MqVtgGQI6QUFJj0wYTScsuUgie/X9Ay5SBLiQ5GthsAgEmZjMz7IjR2IimF21PNLkwwYyClf8So1B/O0I8+gv582nbR2U1CpBgzXIqecRabNcMq6VHqZc+OLjz++4aoBC/fnWVNGEJMqn3zIhUNH2fQOLR82G1VFjcHXL286ot9zy0MbKfUvivGeTvjBVCko4qDCAOpLMq2Zj50nxlTUu41XWCyqan91D8aaUkdW4R2KWgIlWgxTjvVN3YidIEydLPuHWNMEEX+8meZW6TddPdcBmqlTZ/YUNV+yunkZhyCYSHfc2PyvczJuGw3dtRAIoF8v0jyP2u/PXfQhI9bKcb7hXiyEvJuX9MEoFytNZEmUqSlVppoZ1AmICIi0hxZ1FGOVoCQRZmj8+/RBXvm+G0NWjmNp57BVq1Fb7dxIGRjbpfCkfDfrxfd9st0PTjGUGurrKfnhj9rJw586yVVtrogKIWmD2d923zq2ZqBo2h8JPDq2BK/KZOyo/B+JwUR4lKOzwnUNs7J84TC3uIv4iW5Jnc0+LhelNBSo8FYlkHNuh87NBfZOh7Mpmq1/dJPn7+1YVXADHHSKUkWuC+MLtyvMKRcLQRv/uMF+NnHzPB1Qzp3DJUL5bRaV13m69OvqxXvstL9Mh8MlQqffwFO3A+S7dTNwpuMIZDMCwlfRP3ntZazzweDZ1wYl+P/z/iinha1ObTz5dIcwCaxf25mRuX3J/Yd81IDD3mLlXYAkBPGJSWkq2gn1qzoFCJCQFSZ2lh1e6L+jK9fO2sFfeMMCiK1SSr0sNfGF+9bEFaStMEbLr1Mvfg88xfo4iIEpK6xIyZ0KgZ7jAtffllXc+Am6RZoRCLyWGb4oTtdy2S6i+JrwVW61c0N5/zwjbvfJIN75TPPNJ59FhOY1myE3/fe+JJ9crAxIzl2GdUzhOzaSQypJm2fVp63vOqTEZGSNY6/mnqb5BJ07DDEEcOmxRF5x4IbHYBvO2NnS3ylVoiYtONHf/7ifl8tOPSrbz5NjyrPHRU2PA2O3DPE3h1fPDbsdRSA4E2XXuY+8BBnXthzj8iPs90ehZRuBS628y4jIijFdfCBewy/D3bEFO/EiuhKac4bb7vDvu4vzF+MW86KKISbbDDOPd+cfkjqjXf51wuwbiMYlk42suNPyH/xGWZ5Da1c0n/7qfGJ6lRQcMG2nh6zdWWu1ITKQOOkHvkR+6vZVT8c1GvUnbWFub5yqe3siq8SwM/x5l6t1dFqxfKuHjURmLHDdLpyk1d99+4TbYMKvGW2Tgc5pF03pfRpPX23Di8IEHOFoaXbfOa56sXnub+AXAfDIT15gufM091Zs+XddwpvoSa9ue+BwpDJevHXG0v+cdPOrPS/U0vPa6UVUP3h0+GzL9Cf22G5EIExnWriZ59f9OQjiQ8+ik0/mgFpT4gRojBUshmmTMl7+TmzpBRtW1jGOxvbr1veUutAxBQIWaIdEEASDPLhsKDwMCoxUj/Vfr2kueIPA8fcVeOXntGCkj83jBBohrFULQAiC11YGju9V9HA/H5A5DjtBvcQt3hH+Q25MrGhvaY6GX107aoZib4F/t6abIa8xdWlJt40KOeEspDOKO2xZHNj46ln0qcfC3+h1ooxjlK6dizw3NPh00+rv+IKuu9f3FektezAWghMtqt9JhV+8ZHFDdyJTd13btcKTYSQqKlumziFN7aQYYHWyLlMNRoXX1rw4L/ann0+ddY53AzglEn6m++4qwgQDUMmmqFf35xnnvZP3odpYExtTGZuW936el0GkYdEx34gtsa7+jlLaz9f0tJSb0cHBIr3Kun9eK0RNcf6uNy29lAwMzu9NaXaB3qjs6eOXtqw9pR580/pP+7GPUZTNmtO6ncz3/qqHSXLAU+PPE9QKSehUJE+utBzw+C8cr9XK0WcJ+bPbzvjbL5yFfNHSGmUriujCJwBV+D67rs397JLW66/wf7HrdybT9k6DemqiD9/7pf+vtsPuHcXaABSijhPfjGz+eDppuEjxnSq2bz6qoI772x5+OH0n65FSpk3/T1y4/VNe0+Fud+BJwBaozAoGVde7r3lpsiVlxOAKV0QYkZj+73ronNbpCVYgIMLvBBap/g2cNXmMz3VDv+gPY97BpvodrXIPmTTxFykSQyC5UmndQnudUAOPj5UB73FIctT21azz9xKaQ70cFdLO+FKG9nYIL+yf/iQ4hAocrhAgLaHH05ec72RlOAPgJLg2irs9918vXfsuNgLL9OTzzlOxnvHTfnXXNN85532tTcyK8g4c9PR4Htv5hxxBOvacd5NoAE6jHXDI486F13KuSn+fHn+7bc03/nP1LU3cnDM08/Mf+6J9rv+mb7uBr55+pJx0FrabcZhh4Tvvcc7aBARCNKS3Ldr409Uti+KEwcmOCOZ9EA8pZTD8iJmiCizw84nAEIGackYcC/TWmfy9LpZ++9nu62/nz3jB9jLb5gxh4hwjyD/Y+/QcaVBC0EzgyEk1q2NXXm1fPcDboWIcyBCKWXQE/7qIyOYY69bn3n1tcxTT3ggIMER119dcMstLffd4151o9Yp7z/vzr3ySnQcbpo7Cd6ubeGkXakN0XDtX907b/eccAoNG2zfdKsAggMPKPzsw8TLr8ZPOU1Y4S33XCBARG7oZFTlRrxXXxq6+ALuD5oEgFpK9WlT4sXq2NxWO6pNC4VfIGeuUooQaed2wWFA2Xp4AADm38ezZH20brEaYwp/ANVeuZ6Te/oOL/KbwgCNDkNtpxIPPZG87W5obeK+XMz2MUtFbpvnzrv9p/6upkcPI5Qf/NP5fMpk56U3My++AjIprriaGhr0S8+KP12a98C/TKlQ7EL72S5uSqZJK6UN0XDOee5TTyCGuCA9qF/RN19lli1r2f8QCyxCtgX3mJ0zXRu0NkzDTkfF8BG+6/4SOvE4ZBwJGElgsDKW+bAh+WlDcmVCxSUiYxYjA5nBsnU/ANv05mX9EIKOXXOUJhfQVrLdhkKfNSoAU/LYESWR4SETNCkUyJgCiL75ZuqOO3HhIi7C2jIRgOyMlmnoN4DW/eS97c7QZX9KvPtu6OSTCcBZs9oYMLDlsOn02UxAU8p243d/KHj53+hKIcRvuCkZQNYJUS5C85ln6xdekjmRvO+/FQWF9b0GmPEEeELgbrbeEOcglXRazD+caxy4b/yU0ywjTK7UYMNek3x/Os9/1HTLF9i0sJN03RUJZ25belG7sywhazMy7oJEprXm2TWOkbCzu5EIFJEGYgQC0W9AT4sPDpp7RcwJOZ7BQVMIK/sgFIBrZxIffZi5/2H6eg6CAL8flEZAnYmqUMR/3z/Mo49Mjt7bScaDjz/k23N04uVXk/c/Lhtri1Px1J1327fcTJrj0UcVvvq8aRi7sS3q7mwcqbVGJFfppjP/KF98UZx4cvjhh5z334udc4EgC6zOanbGIB1XBUXGHoPU8CFq9Tpr2lRoanXvuov58yHeLsHFEaP8p/3ef8LRRu++GkAAMK2BEYDOSKrLuBtSsjJFNRmnKW3HJMY1ukQcQCBaqPM8PGJBmWH29nv7+liJR3g4AXLQqDnLNjCo6qr4W287z7wiFy/iwNEb1qizxXXaTbNDpoUfewQcO3r2hfTlDAZ+Ba4WljF2pKrc4LvsUv9lf2qeuB/88B1OP77o1eeF14NEu7PRL+2WaKWUVBnl1p519gaAynGT4q2N7bO/WV9UWmVGqj351b7CavDUHHZUvLEhTZRMxBvferPiiiuqp06pMgI1/qJqf1GNr6gaApWAGyLFNcee1PD8s4mqCptIEektr0ZaknJI2eQ6JB2SDqkM6Qwpm/QWP9dEiihDFK+pan35peqTf1+Z26MCWBX4anyF1dnriki1mVvJPTVXXGUTNdx331qA9cFI84svVE+eWmF4KvadliayiWKx1ppjTqgE2HjKH9O2rbQirbfGYudkN3dRRsZAaxNY3lNPNhUUuXf8o2XMZO89tzB/LmuqJI9A25Z9ehe9/rK78qeGfQ9kgodff4EXlcTufdiwQpRs08xCnw/9fgZ+lrD1W2/Zb72Zzi0Se+5h7DPRu9d4MWQYLylCbgAgIAfkmF1vpHMCok6r3WFJtNR1tXLlGvu77zKzvtWLllJTLQfGjQD4CoiIADCZdCCB4SL0eFhDG82cpTMpUtr63e/yn37aXr0KYwnm2jD3x5bx+/GBvZ2vv9U166xLLs+7/58mMNJ6t1cL/0X7gpMmLaU2jdZHHrUvugzAEWauAsYMLpON3gceDJ9/Vn3/kbyqgsCFMeOLFn7XNPUQ+upLOOpwtWAx1a4X4EUrAAjEGSKCbWs3TWADWDoSET1LRb9+rE8ZlZSK4iIjJ4SWhaalOUImQ7at2mK6vslt3EjrK9T6SlW1kbW2AzgIHLkXPRYBgJJASFoqN85GjvJdcJ510nGM8diFlzsvP2tddXXu3Xc60fbkbXek77oHQfKjT4IFi2ljJYCrAH333hG6/AoupcF/SfP1L96AnYhAKRSi/fNPY2ddQNV1wp9L2tXpeMGKhe6GmtbDDjby++jWOpg6pWDGx237HuB+t8D/w9e+AUNSr76efPJpmruQcws1ERBaVsfNSAWuS9IhcKgjyZ8tbsfsuukEBKAQdGdjJQcwUFhgCGAdPRkoJQGC4CClygsE7vmH97DD3Kqa5CMPB049gw0d3Np7CGSSObNnsvy8xv4DeFFZ8N+PGgcd1H7okTDjE+rZ1/fofZEjjuRKEet+q6jfHugOkZKESFZuaDv7ApjxOfcVykxb6NMPjOLihhEjDGAO6Pyvv+RDhzb2G+qR3DbAGj/ac/ONoYl7tX89KzX9WBQ+Ipfa6xE4og9NE5CB4NjZ2NuBHRAg09LR0uUeLyEDhpDttiNNSnYUuCCC66qgj2mCjIPIpIfyVy+XG6tax4wjgOBjj3v3mdQ6drKQSo0cXjh/dvLrL61Ro/SG9W1nngfLFov9Dgg/9ZCv/yBwXGaKnfeXu5JfqS5ACFQq0KtP0Sfvi5v/5ugU027873fS4AG5H30i/nh65KMPvPtOiV71F5FIqFScx6I6FLRyc51k3P16FiXarCsvzF/zY/DN14wz/6DLC7PpVZ1opEQCpCIpSSnSSjOU6TY88pDID99IITBjQzxJsRYVq1PxBkilkTFgyBybBvbJXzw39MazkjJocBZtjd/4d9/osd7rritcsTxy3rnRl16iTJwk0KLv2u660zd538TNt7eNmwjLllp/uT7vsw8C/QcxqZhp/HKUAXbX69iuaKWUlC5R9NtvN0zcewNA1fi92+fNzaSTiYrq6rPOqwCzArwbDz44umSpTdT87ruVw0ZVg1nBfO1Ll0R//LHl5VeTrc0p167902VVYG08/+LKkeOrjWC1J7/ayqu28qq9BZXMWzNxWv0/790QKKr0FVQPGlVz8OH1F1zScMXVVUNHV4GvOlBcA0bN9OMyRCmihnvvqQS+0cpfn1OU2FiTIqq96/aWd/5jE8XmLajaZ2q1Fa4q6FVRNqACoHLPCS1fz1JEWktS6lcE59cEOivadRVR2k43/PPuDd6cCoCavJ4V/sJKENXFvZtefiGVsaOzZlVM3X89GJUYqhahmv4j00Q1V1+7HqACoHrfA6qGjqnKLcsQNVx7XSVgtb+oA2hfYRWadcec2L5oYXVBr4qyQfFYWzKVSGxYl2xtTqWSVVMPqhK+CuAN51+UrKupvvSSDFHVSadUgVUJVv0tt6YT8XWBnA2ANSeenGyobfpm1vqismrACm9u/e23J1Nxl4ik/NVh+fWBJiIl3eyH5Pr1NWedtYEHasBbw31VF13sEsXrGyum7F8BUA28OlhSCaz+nPMcokSsNbZ2bdNDD1eNGFMFUDV+n7QrNx52ZBWY1b7CDqD9RVXA6848xyWq7D2w0gxmUvH6f96zHmC9PxJduiz2w6IKHqgG3nDDTbGN1csBWj75NJ3JVA3Zowr4hp4D0lq1PPtCNZrV4K8Kl1aAWYHWxjPOTKxZq4k0af0boEw7XCt096Rjr0qlPH36FD31VGTuTDrxSIUcn3iu6fCj5XdzCz7/MG/9OuOaa4ExDdo6/PDUT0sb956aeeFFc7/J5kHTNCIr78EF17X1sHnXNREBQCpDWjPLh05S1jVYZeUCDJFsS7/4ojlqFPbsKUHx/AKRV1C8erUxcgRTKue151RBMatZH3voUaVsjZwB6VgbP/qI3LkzC5/9t79/P5QSCXeS9txlTH6Lk3YI50haKB0eN67Ha68GZn/MTzjG/uiTlqOOauwzLPOf98TBU8jNsLwenqMOc1etoaWLk3+/qXXESPv+p5FQ9CoH0NTShoYH+CaeDBE4agcZ46ahQasNlXxAH80twf1qbRUiYzlBAoD8iG5vTz36SGzaQXX7H4oD+ltXXQbccC75S/qss7XHRycdGf7ms9K3/xOcMEkoBZpgJ/bo3W3ZzchwJwWRAQfQGkhHJu6tJ+5t3/zX+Auvuc+9Evvz5QhggleVlaa++NI/ZT+rpUUuWui896H9ylvQ3Gb27edu3OhUrzEAmCvACGRX3QcA7SokAs4QwF23zjvsEFIJG8A/cayUijY2MECxxyh32U/uff8yPPncrW4dOAqbm0ApGtDfe8LR3lN/7x0yhGcTGgC/kRZvLr8t0B3CGAIDrTmBr98gz003ZP58ufv116lX37ZnzoZVP8UOODCZX24cPMV3wjGeC85x3/9UtwgcOsDsURZ88kncsNFdt1Z++RVrTwJyAAZSAaISBgNQGyrN4rLQKy/zvr3E+Emx62+A1gYIFSReex2/XWCAqTJxAIBMSkw/1Dr5BHP/ad5QDgfIMl//BYiz8isFLDsvWpNWIIxst7vT0pyZM9edMTMzY5ZavhhAEYAADwBjxx9hTtrLGDbc6F1GuTltBx7Jlq6igB8SbWza1PBbL7ccfJyYPxsOne7761XQ2KTXb0i/+zF8NUODAeAoUACmGD6ST97Lf/AB5uRJntz8jjFICYztTKLvV5T/OtCbJLs0AOcAIAG01u7ate6PizPzFsilP6n1Fbi+woVUJxhZoqGjGFeBQUE/i7czUBK0Bupo8wKT9e4Pfcs8I4aIiWPZqD2Mvv0ty+pY2yt7xZ3IWP8W8r8DOiuU3XqJcMtdMDPJmK5vdjdW6po6WVOvNlazVFrGk5RM87QDRCSEyvEboQCE/FhYJMp6Ymmx2bOHWVIq/MHNgSSlOjrR/hf4bpL/NdCbi9Ydy5pwBp3dGdTZF7eJDt1KaNMCXD//izpaSBD/y/ahG/m/BPTmQtQB+qatRxC3bhvczLmGbIthh3X5X2puV/J/Fej/z8n/A9dOXElYP2QPAAAAAElFTkSuQmCC"

def get_amblem_path():
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    tmp.write(base64.b64decode(AMBLEM_B64))
    tmp.close(); return tmp.name

def hash_sifre(s):
    return hashlib.sha256(s.encode()).hexdigest()

# ══════════════════════════════════════════════════════════════════
# VERİTABANI
# ══════════════════════════════════════════════════════════════════
def init_db():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS kullanici(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        kullanici_adi TEXT UNIQUE, sifre_hash TEXT, ad_soyad TEXT, aktif INTEGER DEFAULT 1)""")
    try:
        c.execute("INSERT INTO kullanici(kullanici_adi,sifre_hash,ad_soyad) VALUES(?,?,?)",
                  ("ismail",hash_sifre("1234"),"Dr. İsmail ÇELİK"))
    except: pass
    c.execute("""CREATE TABLE IF NOT EXISTS faiz_oran(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        gecerlilik_tarihi TEXT, oran REAL)""")
    try:
        c.execute("INSERT INTO faiz_oran(gecerlilik_tarihi,oran) VALUES(?,?)",("31.05.2024",9.0))
        c.execute("INSERT INTO faiz_oran(gecerlilik_tarihi,oran) VALUES(?,?)",("01.06.2024",24.0))
    except: pass
    c.execute("""CREATE TABLE IF NOT EXISTS mukellef(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        no INTEGER UNIQUE, tckn_vkn TEXT UNIQUE,
        mukellef TEXT, il TEXT, ilce TEXT, adres TEXT, iletisim TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS icmal(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mukellef_no INTEGER, tckn_vkn TEXT, mukellef TEXT,
        dosya_no TEXT, dosya_turu TEXT, dosya_durumu TEXT,
        islem_tarihi TEXT, uyari_tarihi TEXT, aciklama TEXT, islem_yapan TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS ayristirma(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mukellef_no INTEGER, tarih TEXT, dosya_no TEXT,
        tckn_vkn TEXT, mukellef TEXT, aciklama TEXT,
        anapara REAL DEFAULT 0, faiz REAL DEFAULT 0, masraf REAL DEFAULT 0,
        harc REAL DEFAULT 0, vekalet REAL DEFAULT 0,
        avans_iadesi REAL DEFAULT 0, iade_edilecek REAL DEFAULT 0,
        toplam REAL DEFAULT 0, islem_yapan TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS parametreler(
        anahtar TEXT PRIMARY KEY, deger TEXT)""")
    conn.commit(); conn.close()

def get_conn(): return sqlite3.connect(DB_PATH)

def migrate_db():
    conn = get_conn(); c = conn.cursor()
    migs = [("mukellef","il","TEXT"),("mukellef","ilce","TEXT"),
            ("icmal","dosya_turu","TEXT"),("icmal","islem_yapan","TEXT"),
            ("ayristirma","islem_yapan","TEXT"),
            ("faiz_oran","gecerlilik_tarihi","TEXT"),("faiz_oran","oran","TEXT")]
    for tbl,col,typ in migs:
        try: c.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} {typ}")
        except: pass
    # parametreler tablosu
    try: c.execute("CREATE TABLE IF NOT EXISTS parametreler(anahtar TEXT PRIMARY KEY, deger TEXT)")
    except: pass
    conn.commit(); conn.close()

def next_mukellef_no():
    conn = get_conn()
    row = conn.execute("SELECT MAX(no) FROM mukellef").fetchone()
    conn.close()
    return (row[0] or 0) + 1

def mukellef_getir_no(no):
    conn = get_conn()
    row = conn.execute("SELECT * FROM mukellef WHERE no=?", (no,)).fetchone()
    conn.close(); return row

def faiz_dilimler_db():
    """DB'den faiz oranlarını oku, liste olarak döndür"""
    conn = get_conn()
    rows = conn.execute("SELECT gecerlilik_tarihi, oran FROM faiz_oran ORDER BY id").fetchall()
    conn.close()
    dilimler = []
    for tarih_str, oran in rows:
        try:
            for fmt in ("%d.%m.%Y","%Y-%m-%d"):
                try: t = datetime.strptime(tarih_str, fmt).date(); break
                except: pass
            dilimler.append((t, Decimal(str(oran))))
        except: pass
    if not dilimler:
        dilimler = [(date(2024,5,31),Decimal("9")), (date(9999,12,31),Decimal("24"))]
    else:
        dilimler.append((date(9999,12,31), dilimler[-1][1]))
    return dilimler

def para_format(val):
    try: return f"{float(val):,.2f} ₺".replace(",","X").replace(".",",").replace("X",".")
    except: return "0,00 ₺"

def tarih_parse(s):
    if not s: return None
    for fmt in ("%d.%m.%Y","%d/%m/%Y","%Y-%m-%d"):
        try: return datetime.strptime(s.strip(), fmt).date()
        except: pass
    return None

def hesapla_faiz(anapara, baslangic, bitis):
    if bitis <= baslangic:
        return {"anapara":anapara,"toplam_gun":0,"dilimler":[],"faiz_tutari":0.0,"toplam":anapara}
    ana = Decimal(str(anapara))
    dilimler_db = faiz_dilimler_db()
    dilimler, toplam_faiz, current = [], Decimal("0"), baslangic
    for i,(kesim,oran) in enumerate(dilimler_db):
        if current >= bitis: break
        if i+1 < len(dilimler_db):
            dilim_bitis = min(bitis, dilimler_db[i+1][0])
        else:
            dilim_bitis = bitis
        if dilim_bitis <= current: continue
        gun = (dilim_bitis - current).days
        if gun <= 0: current = dilim_bitis; continue
        faiz = (ana*oran/Decimal("100")/Decimal("365")*Decimal(str(gun))).quantize(Decimal("0.01"),rounding=ROUND_HALF_UP)
        dilimler.append({"baslangic":current.strftime("%d.%m.%Y"),"bitis":dilim_bitis.strftime("%d.%m.%Y"),
                         "gun":gun,"oran":float(oran),"faiz":float(faiz)})
        toplam_faiz += faiz; current = dilim_bitis
    toplam_faiz = toplam_faiz.quantize(Decimal("0.01"),rounding=ROUND_HALF_UP)
    return {"anapara":float(ana),"toplam_gun":(bitis-baslangic).days,
            "dilimler":dilimler,"faiz_tutari":float(toplam_faiz),"toplam":float(ana+toplam_faiz)}


# ══════════════════════════════════════════════════════════════════
# PDF / EXCEL RAPOR (Türkçe karakter destekli)
# ══════════════════════════════════════════════════════════════════
def yazdir_pdf(baslik, sutun_lbls, satirlar, imzaci=None):
    """Geçici PDF oluştur ve Windows yazıcı iletişim kutusunu aç"""
    import tempfile, subprocess, sys, os
    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False,
                                      prefix="icra_yazdir_")
    tmp.close()
    pdf_rapor_olustur(tmp.name, baslik, sutun_lbls, satirlar, imzaci)
    try:
        if sys.platform == "win32":
            # Windows: PDF'i varsayılan yazıcıya gönder
            import win32api, win32print
            win32api.ShellExecute(0, "print", tmp.name, None, ".", 0)
        else:
            # Linux/Mac fallback
            subprocess.Popen(["xdg-open", tmp.name])
    except Exception:
        # win32api yoksa — PDF'i aç, kullanıcı yazıcıya gönderir
        try:
            if sys.platform == "win32":
                os.startfile(tmp.name, "print")
            else:
                subprocess.Popen(["xdg-open", tmp.name])
        except Exception as e:
            messagebox.showerror("Hata", f"Yazıcı açılamadı:\n{e}")

def pdf_rapor_olustur(filepath, baslik, sutun_lbls, satirlar, imzaci, mukellef_adi=""):
    amblem_path = get_amblem_path()
    doc = SimpleDocTemplate(filepath, pagesize=A4,
        rightMargin=1.2*cm, leftMargin=1.2*cm, topMargin=1.5*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet(); hikaye = []
    try: amb = RLImage(amblem_path, width=2.4*cm, height=2.4*cm)
    except: amb = Paragraph("", styles["Normal"])

    tarih_str = datetime.now().strftime("%d.%m.%Y")
    bp = ParagraphStyle("bh", fontName=PDF_FONT_BOLD, fontSize=11,
        alignment=TA_CENTER, textColor=colors.HexColor("#0F1923"), leading=16)
    tp = ParagraphStyle("th", fontName=PDF_FONT, fontSize=9,
        alignment=TA_RIGHT, textColor=colors.HexColor("#243447"))
    baslik_p = Paragraph(f"<b>T.C.<br/>ISPARTA İL ÖZEL İDARESİ<br/>HUKUK MÜŞAVİRLİĞİ</b><br/>{baslik}", bp)
    tarih_p  = Paragraph(f"<b>Tarih:</b><br/>{tarih_str}", tp)
    bas_t = Table([[amb, baslik_p, tarih_p]], colWidths=[2.6*cm, 11.1*cm, 2.8*cm])
    bas_t.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),("BOTTOMPADDING",(0,0),(-1,-1),8)]))
    hikaye.append(bas_t)
    hikaye.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#cc1a1a"), spaceAfter=8))

    tablo_data = [list(sutun_lbls)]
    cp = ParagraphStyle("cp", fontName=PDF_FONT, fontSize=7, alignment=TA_CENTER, leading=9)
    hp = ParagraphStyle("hp", fontName=PDF_FONT_BOLD, fontSize=7, alignment=TA_CENTER,
        textColor=colors.white, leading=9)
    for satir in satirlar:
        tablo_data.append([Paragraph(str(v) if v else "", cp) for v in satir])
    header_row = [Paragraph(str(h), hp) for h in tablo_data[0]]
    tablo_data[0] = header_row

    avail = A4[0] - 2.4*cm
    n = len(sutun_lbls)
    col_w = [avail/n]*n
    tbl = Table(tablo_data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#cc1a1a")),
        ("FONTSIZE",(0,0),(-1,-1),7),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.HexColor("#F5F5F5"),colors.white]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CCCCCC")),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
    ]))
    hikaye.append(tbl); hikaye.append(Spacer(1, 0.8*cm))
    if imzaci:
        ip = ParagraphStyle("ip", fontName=PDF_FONT, fontSize=9,
            alignment=TA_RIGHT, textColor=colors.HexColor("#0F1923"), leading=14)
        imza_p = Paragraph(f"<b>{imzaci['isim']}</b><br/>{imzaci['unvan']}", ip)
        hikaye.append(Table([["", imza_p]], colWidths=[12*cm, 4.5*cm]))
    doc.build(hikaye)
    try: os.unlink(amblem_path)
    except: pass

def excel_rapor_olustur(filepath, baslik, sutun_lbls, satirlar):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = baslik[:28]
    fill_h = PatternFill("solid", fgColor="cc1a1a")
    fill_e = PatternFill("solid", fgColor="EEF2F5")
    fill_o = PatternFill("solid", fgColor="FFFFFF")
    fnt_h  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    fnt_b  = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    fnt_d  = Font(color="1A1A1A", size=9, name="Calibri")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin",color="CCCCCC"),right=Side(style="thin",color="CCCCCC"),
                    top=Side(style="thin",color="CCCCCC"),bottom=Side(style="thin",color="CCCCCC"))
    n = len(sutun_lbls)
    ws.merge_cells(f"A1:{chr(64+min(n,26))}1")
    tc = ws.cell(1,1, f"T.C. ISPARTA İL ÖZEL İDARESİ HUKUK MÜŞAVİRLİĞİ — {baslik}   |   {datetime.now().strftime('%d.%m.%Y')}")
    tc.font=fnt_b; tc.alignment=center; tc.fill=PatternFill("solid",fgColor="cc1a1a")
    ws.row_dimensions[1].height=26
    for ci,lbl in enumerate(sutun_lbls,1):
        c=ws.cell(2,ci,str(lbl)); c.font=fnt_h; c.fill=fill_h; c.alignment=center; c.border=border
        ws.column_dimensions[chr(64+ci) if ci<=26 else "A"].width=max(12,len(str(lbl))+4)
    for ri,satir in enumerate(satirlar,3):
        fill=fill_e if ri%2==0 else fill_o
        for ci,val in enumerate(satir,1):
            c=ws.cell(ri,ci,str(val) if val is not None else "")
            c.font=fnt_d; c.fill=fill; c.alignment=center; c.border=border
    ws.freeze_panes="A3"
    wb.save(filepath)


# ══════════════════════════════════════════════════════════════════
# ORTAK WİDGETLAR
# ══════════════════════════════════════════════════════════════════
def styled_entry(parent, textvariable=None, width=20, **kw):
    return tk.Entry(parent, textvariable=textvariable, width=width,
        bg=CLR["panel"], fg=CLR["text"], insertbackground=CLR["accent"],
        relief="flat", font=F, highlightthickness=1,
        highlightbackground=CLR["border"], highlightcolor=CLR["accent"], **kw)

def styled_btn(parent, text, command=None, color=None, width=None, **kw):
    c = color or CLR["accent"]
    fg = "#ffffff"
    b = tk.Button(parent, text=text, command=command, bg=c, fg=fg,
        font=("Segoe UI",10,"bold"), relief="flat", cursor="hand2",
        activebackground=CLR["hover"], activeforeground="#ffffff",
        padx=14, pady=7, bd=0, **kw)
    if width: b.config(width=width)
    b.bind("<Enter>", lambda e: b.config(bg=CLR["hover"]))
    b.bind("<Leave>", lambda e: b.config(bg=c))
    return b

def section_frame(parent, title=""):
    frame = tk.Frame(parent, bg=CLR["card"], highlightthickness=1, highlightbackground=CLR["border"])
    if title:
        tk.Label(frame, text=title, bg=CLR["card"], fg=CLR["accent"],
            font=FB, anchor="w", padx=10, pady=5).pack(fill="x")
        tk.Frame(frame, bg=CLR["border"], height=1).pack(fill="x")
    return frame

def styled_tree(parent, columns, heights=15):
    style = ttk.Style(); style.theme_use("clam")
    style.configure("Dark.Treeview", background=CLR["tbl_even"], foreground=CLR["text"],
        fieldbackground=CLR["tbl_even"], rowheight=26, font=FS)
    style.configure("Dark.Treeview.Heading", background=CLR["tbl_head"],
        foreground=CLR["accent"], font=("Segoe UI",9,"bold"), relief="flat")
    style.map("Dark.Treeview", background=[("selected",CLR["select"])], foreground=[("selected",CLR["text"])])
    frame = tk.Frame(parent, bg=CLR["card"])
    tree = ttk.Treeview(frame, columns=columns, show="headings", style="Dark.Treeview", height=heights)
    vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
    tree.pack(side="left", fill="both", expand=True)
    tree.tag_configure("odd", background=CLR["tbl_odd"])
    tree.tag_configure("even", background=CLR["tbl_even"])
    return frame, tree

def center_window(win, w, h):
    win.update_idletasks()
    sw=win.winfo_screenwidth(); sh=win.winfo_screenheight()
    win.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

# ── Takvim widget ─────────────────────────────────────────────────
class Takvim(tk.Toplevel):
    """Tarih seçici popup"""
    def __init__(self, parent, callback, initial=None):
        super().__init__(parent)
        self.callback = callback
        self.overrideredirect(True)
        self.configure(bg=CLR["card"], highlightthickness=1, highlightbackground=CLR["accent"])
        now = initial or date.today()
        self.year = tk.IntVar(value=now.year)
        self.month = tk.IntVar(value=now.month)
        self.sel_day = None
        self._build()
        self._draw()
        center_window(self, 240, 220)
        self.lift(); self.focus_force()
        self.bind("<FocusOut>", lambda e: self.destroy())

    def _build(self):
        nav = tk.Frame(self, bg=CLR["card"]); nav.pack(fill="x", padx=4, pady=4)
        styled_btn(nav,"◀", lambda:self._change(-1), color=CLR["panel"]).pack(side="left")
        self.title_lbl = tk.Label(nav, text="", bg=CLR["card"], fg=CLR["accent"],
            font=FB, width=14); self.title_lbl.pack(side="left", expand=True)
        styled_btn(nav,"▶", lambda:self._change(1), color=CLR["panel"]).pack(side="right")
        tk.Frame(self, bg=CLR["border"], height=1).pack(fill="x")
        self.grid_frame = tk.Frame(self, bg=CLR["card"]); self.grid_frame.pack(padx=4, pady=4)
        for ci, gn in enumerate(["Pzt","Sal","Çar","Per","Cum","Cmt","Paz"]):
            tk.Label(self.grid_frame, text=gn, bg=CLR["tbl_head"], fg=CLR["subtext"],
                font=("Segoe UI",8,"bold"), width=3).grid(row=0, column=ci, padx=1, pady=1)

    def _change(self, d):
        m = self.month.get() + d
        y = self.year.get()
        if m > 12: m=1; y+=1
        if m < 1:  m=12; y-=1
        self.month.set(m); self.year.set(y); self._draw()

    def _draw(self):
        for w in self.grid_frame.grid_slaves():
            if int(w.grid_info()["row"]) > 0: w.destroy()
        import calendar
        y,m = self.year.get(), self.month.get()
        self.title_lbl.config(text=f"{y} / {m:02d}")
        cal = calendar.monthcalendar(y, m)
        today = date.today()
        for ri, week in enumerate(cal):
            for ci, day in enumerate(week):
                if day == 0: tk.Label(self.grid_frame, text="", bg=CLR["card"], width=3).grid(row=ri+1,column=ci,padx=1,pady=1); continue
                is_today = (day==today.day and m==today.month and y==today.year)
                bg = CLR["accent"] if is_today else CLR["panel"]
                fg = "#000" if is_today else CLR["text"]
                b = tk.Button(self.grid_frame, text=str(day), bg=bg, fg=fg,
                    font=("Segoe UI",8), relief="flat", cursor="hand2",
                    width=3, pady=2,
                    command=lambda d2=day: self._pick(d2))
                b.grid(row=ri+1, column=ci, padx=1, pady=1)

    def _pick(self, day):
        d = date(self.year.get(), self.month.get(), day)
        self.callback(d.strftime("%d.%m.%Y"))
        self.destroy()

def tarih_entry_with_cal(parent, textvariable, width=12, row=0, col=0, use_grid=False):
    """Tarih Entry + takvim butonu çifti — .../.../... placeholder"""
    bg = CLR["card"]
    try: bg = parent.cget("bg")
    except: pass
    frame = tk.Frame(parent, bg=bg)

    # Placeholder mantığı
    PLACEHOLDER = "GG/AA/YYYY"
    PLACEHOLDER_RENK = CLR["subtext"]
    NORMAL_RENK = CLR["text"]

    e = tk.Entry(frame, textvariable=textvariable, width=width,
                 bg=CLR["panel"], fg=PLACEHOLDER_RENK,
                 insertbackground=CLR["accent"],
                 relief="flat", font=F, highlightthickness=1,
                 highlightbackground=CLR["border"],
                 highlightcolor=CLR["accent"])
    e.pack(side="left")

    def on_focus_in(event):
        if textvariable.get() == PLACEHOLDER:
            textvariable.set("")
            e.config(fg=NORMAL_RENK)

    def on_focus_out(event):
        val = textvariable.get().strip()
        if not val:
            textvariable.set(PLACEHOLDER)
            e.config(fg=PLACEHOLDER_RENK)
        else:
            e.config(fg=NORMAL_RENK)

    def on_var_change(*args):
        val = textvariable.get()
        if val and val != PLACEHOLDER:
            e.config(fg=NORMAL_RENK)

    # Başlangıçta placeholder göster
    if not textvariable.get():
        textvariable.set(PLACEHOLDER)
        e.config(fg=PLACEHOLDER_RENK)

    e.bind("<FocusIn>", on_focus_in)
    e.bind("<FocusOut>", on_focus_out)
    textvariable.trace_add("write", on_var_change)

    def open_cal():
        val = textvariable.get()
        init = tarih_parse(val) if val != PLACEHOLDER else None
        def set_tarih(s):
            textvariable.set(s)
            e.config(fg=NORMAL_RENK)
        Takvim(parent, set_tarih, initial=init)

    btn = tk.Button(frame, text="📅", command=open_cal,
                    bg=CLR["panel"], fg=CLR["accent"],
                    relief="flat", cursor="hand2",
                    font=("Segoe UI",11), padx=4)
    btn.pack(side="left", padx=(2,0))
    return frame


def tarih_parse_safe(s):
    """Placeholder dahil güvenli tarih parse"""
    if not s or s in ("GG/AA/YYYY", "GG.AA.YYYY"): return None
    return tarih_parse(s)


# ══════════════════════════════════════════════════════════════════
# GİRİŞ (LOGIN)
# ══════════════════════════════════════════════════════════════════
class LoginEkrani(tk.Toplevel):
    def __init__(self, app):
        super().__init__(); self.app = app
        self.title("Giriş"); self.configure(bg=CLR["bg"])
        self.resizable(False,False); self.protocol("WM_DELETE_WINDOW", self._cikis)
        self._build(); center_window(self,420,500); self.grab_set()

    def _build(self):
        logo_f = tk.Frame(self,bg=CLR["bg"],pady=18); logo_f.pack(fill="x")
        try:
            import io; from PIL import Image, ImageTk
            img = Image.open(io.BytesIO(base64.b64decode(AMBLEM_B64))).resize((90,90),Image.LANCZOS)
            self._img = ImageTk.PhotoImage(img)
            tk.Label(logo_f,image=self._img,bg=CLR["bg"]).pack()
        except:
            tk.Label(logo_f,text="⚖",bg=CLR["bg"],fg=CLR["red"],font=("Segoe UI",44)).pack()
        tk.Label(self,text="ISPARTA İL ÖZEL İDARESİ",bg=CLR["bg"],fg=CLR["red"],
                 font=("Segoe UI",13,"bold")).pack()
        tk.Label(self,text="Hukuk Müşavirliği  |  İcra Yardım v6.0",bg=CLR["bg"],
                 fg=CLR["subtext"],font=("Segoe UI",10)).pack(pady=(2,14))
        tk.Frame(self,bg=CLR["border"],height=1).pack(fill="x",padx=24)

        form=tk.Frame(self,bg=CLR["bg"]); form.pack(padx=34,pady=18); form.columnconfigure(1,weight=1)

        tk.Label(form,text="Kullanıcı Adı:",bg=CLR["bg"],fg=CLR["subtext"],
                 font=("Segoe UI",11),anchor="w").grid(row=0,column=0,sticky="w",pady=8)
        self.kadi_var=tk.StringVar()
        e1=tk.Entry(form,textvariable=self.kadi_var,width=22,
                    bg=CLR["panel"],fg=CLR["text"],insertbackground=CLR["accent"],
                    relief="flat",font=("Segoe UI",12),highlightthickness=1,
                    highlightbackground=CLR["border"],highlightcolor=CLR["accent"])
        e1.grid(row=0,column=1,sticky="ew",padx=(12,0),pady=8,ipady=4)

        tk.Label(form,text="Şifre:",bg=CLR["bg"],fg=CLR["subtext"],
                 font=("Segoe UI",11),anchor="w").grid(row=1,column=0,sticky="w",pady=8)
        self.sifre_var=tk.StringVar()
        e2=tk.Entry(form,textvariable=self.sifre_var,width=22,show="•",
                    bg=CLR["panel"],fg=CLR["text"],insertbackground=CLR["accent"],
                    relief="flat",font=("Segoe UI",12),highlightthickness=1,
                    highlightbackground=CLR["border"],highlightcolor=CLR["accent"])
        e2.grid(row=1,column=1,sticky="ew",padx=(12,0),pady=8,ipady=4)

        e1.bind("<Tab>",lambda ev:(e2.focus_set(),"break"))
        e2.bind("<Return>",lambda ev:self.giris())

        # Kullanıcı adını hatırla
        self.hatirla_var=tk.BooleanVar(value=False)
        hatirla_frame=tk.Frame(self,bg=CLR["bg"]); hatirla_frame.pack(padx=34,anchor="w")
        cb=tk.Checkbutton(hatirla_frame,text="Kullanıcı adını hatırla",
                          variable=self.hatirla_var,
                          bg=CLR["bg"],fg=CLR["subtext"],
                          selectcolor=CLR["panel"],activebackground=CLR["bg"],
                          font=("Segoe UI",10),cursor="hand2")
        cb.pack(side="left")

        # Kaydedilmiş kullanıcı adını yükle
        self._yukle_hatirla(e1)

        self.hata_lbl=tk.Label(self,text="",bg=CLR["bg"],fg=CLR["danger"],
                               font=("Segoe UI",10)); self.hata_lbl.pack(pady=(8,0))

        btn=tk.Button(self,text="  🔓  Giriş Yap  ",command=self.giris,
                      bg=CLR["red"],fg="#ffffff",font=("Segoe UI",12,"bold"),
                      relief="flat",cursor="hand2",padx=20,pady=10)
        btn.pack(pady=14)
        btn.bind("<Enter>",lambda e:btn.config(bg="#a01010"))
        btn.bind("<Leave>",lambda e:btn.config(bg=CLR["red"]))

        # Güncelleme butonu
        tk.Frame(self,bg=CLR["border"],height=1).pack(fill="x",padx=30)
        guncelle_btn=tk.Button(self,text="🔄  Güncellemeleri Kontrol Et",
                               command=lambda:guncelleme_kontrol_ve_goster(self),
                               bg=CLR["bg"],fg=CLR["subtext"],
                               font=("Segoe UI",9),relief="flat",cursor="hand2",
                               pady=8)
        guncelle_btn.pack()
        guncelle_btn.bind("<Enter>",lambda e:guncelle_btn.config(fg=CLR["accent"]))
        guncelle_btn.bind("<Leave>",lambda e:guncelle_btn.config(fg=CLR["subtext"]))

        # Sürüm bilgisi
        tk.Label(self,text=f"Sürüm {APP_SURUM}",bg=CLR["bg"],
                 fg=CLR["subtext"],font=("Segoe UI",8)).pack(pady=(0,8))
        e1.focus_set()

    def _yukle_hatirla(self, entry):
        """Kaydedilmiş kullanıcı adını yükle"""
        try:
            if os.path.exists(HATIRLA_PATH):
                with open(HATIRLA_PATH,"r",encoding="utf-8") as f:
                    saved=f.read().strip()
                if saved:
                    self.kadi_var.set(saved)
                    self.hatirla_var.set(True)
        except: pass

    def _kaydet_hatirla(self):
        try:
            if self.hatirla_var.get():
                with open(HATIRLA_PATH,"w",encoding="utf-8") as f:
                    f.write(self.kadi_var.get().strip())
            else:
                if os.path.exists(HATIRLA_PATH): os.remove(HATIRLA_PATH)
        except: pass

    def giris(self):
        kadi=self.kadi_var.get().strip().lower(); sifre=self.sifre_var.get()
        if not kadi or not sifre:
            self.hata_lbl.config(text="Kullanıcı adı ve şifre gereklidir."); return
        conn=get_conn()
        row=conn.execute("SELECT id,ad_soyad,aktif FROM kullanici WHERE kullanici_adi=? AND sifre_hash=?",
                         (kadi,hash_sifre(sifre))).fetchone()
        conn.close()
        if not row:
            self.hata_lbl.config(text="❌  Hatalı kullanıcı adı veya şifre!")
            self.sifre_var.set(""); return
        if not row[2]:
            self.hata_lbl.config(text="Bu hesap devre dışı."); return
        self._kaydet_hatirla()
        self.app.aktif_kullanici={"id":row[0],"kadi":kadi,"ad_soyad":row[1]}
        self.destroy(); self.app.after_login()

    def _cikis(self): self.app.destroy()

# ══════════════════════════════════════════════════════════════════
# KULLANICI YÖNETİMİ
# ══════════════════════════════════════════════════════════════════
class KullaniciEkrani(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=CLR["bg"]); self.app=app; self._build(); self.load()

    def _build(self):
        hdr=tk.Frame(self,bg=CLR["bg"]); hdr.pack(fill="x",padx=16,pady=(12,4))
        tk.Label(hdr,text="🔐  Kullanıcı Yönetimi",bg=CLR["bg"],fg=CLR["accent"],font=FT).pack(side="left")
        warn=tk.Frame(self,bg="#2a1a0a",highlightthickness=1,highlightbackground=CLR["warning"]); warn.pack(fill="x",padx=16,pady=4)
        tk.Label(warn,text="⚠  Yalnızca yönetici kullanıcı ekleyip yönetebilir.",bg="#2a1a0a",fg=CLR["warning"],font=FS,padx=12,pady=7).pack(fill="x")
        bf=tk.Frame(self,bg=CLR["bg"]); bf.pack(fill="x",padx=16,pady=4)
        if self.app.aktif_kullanici["kadi"]==ADMIN_USER:
            styled_btn(bf,"＋ Yeni Kullanıcı",self.yeni,color=CLR["success"]).pack(side="left",padx=4)
            styled_btn(bf,"🚫 Devre Dışı",self.sil,color=CLR["danger"]).pack(side="left",padx=4)
            styled_btn(bf,"🔑 Şifre Sıfırla",self.sifre_sifirla,color=CLR["warning"]).pack(side="left",padx=4)
        else:
            tk.Label(bf,text="Bu ekrana yalnızca yönetici erişebilir.",bg=CLR["bg"],fg=CLR["danger"],font=FS).pack(side="left")
        cols=("kullanici_adi","ad_soyad","aktif"); lbls=("Kullanıcı Adı","Ad Soyad","Durum"); widths=(130,200,80)
        tf,self.tree=styled_tree(self,cols,heights=12)
        for c,l,w in zip(cols,lbls,widths): self.tree.heading(c,text=l); self.tree.column(c,width=w,anchor="center")
        tf.pack(fill="both",expand=True,padx=16,pady=(4,12))

    def load(self):
        conn=get_conn(); rows=conn.execute("SELECT kullanici_adi,ad_soyad,aktif FROM kullanici ORDER BY id").fetchall(); conn.close()
        self.tree.delete(*self.tree.get_children())
        for i,row in enumerate(rows):
            self.tree.insert("","end",iid=str(row[0]),values=(row[0],row[1],"✅ Aktif" if row[2] else "🚫 Pasif"),tags=("odd" if i%2 else "even",))

    def yeni(self): KullaniciForm(self, self.load)
    def sil(self):
        sel=self.tree.focus()
        if not sel: messagebox.showinfo("Bilgi","Kullanıcı seçin."); return
        if sel==ADMIN_USER: messagebox.showerror("Hata","Yönetici silinemez."); return
        if messagebox.askyesno("Onayla",f"'{sel}' devre dışı bırakılsın mı?"):
            conn=get_conn(); conn.execute("UPDATE kullanici SET aktif=0 WHERE kullanici_adi=?",(sel,)); conn.commit(); conn.close(); self.load()
    def sifre_sifirla(self):
        import tkinter.simpledialog
        sel=self.tree.focus()
        if not sel: messagebox.showinfo("Bilgi","Kullanıcı seçin."); return
        yeni=tkinter.simpledialog.askstring("Şifre Sıfırla",f"'{sel}' için yeni şifre:",show="•")
        if yeni and yeni.strip():
            conn=get_conn(); conn.execute("UPDATE kullanici SET sifre_hash=? WHERE kullanici_adi=?",(hash_sifre(yeni.strip()),sel)); conn.commit(); conn.close()
            messagebox.showinfo("Başarılı","Şifre güncellendi.")

class KullaniciForm(tk.Toplevel):
    def __init__(self, parent, callback):
        super().__init__(parent); self.callback=callback
        self.title("Yeni Kullanıcı"); self.configure(bg=CLR["card"]); self.resizable(False,False)
        self._build(); center_window(self,360,280); self.grab_set()

    def _build(self):
        tk.Label(self,text="Yeni Kullanıcı",bg=CLR["card"],fg=CLR["accent"],font=FT,pady=10).pack(fill="x",padx=16)
        tk.Frame(self,bg=CLR["border"],height=1).pack(fill="x")
        form=tk.Frame(self,bg=CLR["card"]); form.pack(padx=20,pady=12); form.columnconfigure(1,weight=1)
        self.vars={}; entries=[]
        for i,(lbl,key,show) in enumerate([("Kullanıcı Adı","kadi",False),("Ad Soyad","ad_soyad",False),("Şifre","sifre",True),("Şifre Tekrar","sifre2",True)]):
            tk.Label(form,text=lbl+":",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=i,column=0,sticky="w",padx=(0,12),pady=6)
            v=tk.StringVar(); e=styled_entry(form,textvariable=v,width=22,show="•" if show else "")
            e.grid(row=i,column=1,sticky="ew",pady=6); self.vars[key]=v; entries.append(e)
        for i in range(len(entries)-1): entries[i].bind("<Tab>",lambda ev,n=i+1:(entries[n].focus_set(),"break"))
        bf=tk.Frame(self,bg=CLR["card"],pady=10); bf.pack(fill="x",padx=16)
        styled_btn(bf,"💾 Kaydet",self.kaydet,color=CLR["success"]).pack(side="left")
        styled_btn(bf,"İptal",self.destroy,color=CLR["danger"]).pack(side="right")

    def kaydet(self):
        kadi=self.vars["kadi"].get().strip().lower(); ad=self.vars["ad_soyad"].get().strip()
        s1=self.vars["sifre"].get(); s2=self.vars["sifre2"].get()
        if not kadi or not ad: messagebox.showerror("Hata","Tüm alanlar zorunludur."); return
        if s1!=s2: messagebox.showerror("Hata","Şifreler eşleşmiyor."); return
        if len(s1)<4: messagebox.showerror("Hata","Şifre en az 4 karakter."); return
        conn=get_conn()
        try: conn.execute("INSERT INTO kullanici(kullanici_adi,sifre_hash,ad_soyad) VALUES(?,?,?)",(kadi,hash_sifre(s1),ad)); conn.commit()
        except sqlite3.IntegrityError: messagebox.showerror("Hata","Bu kullanıcı adı zaten var."); conn.close(); return
        conn.close(); self.callback(); self.destroy()


# ══════════════════════════════════════════════════════════════════
# FAİZ ORAN TANIMLAMA
# ══════════════════════════════════════════════════════════════════
def vekalet_param_getir():
    """Vekalet ücreti parametrelerini DB'den oku"""
    conn=get_conn()
    def get(k, default):
        row=conn.execute("SELECT deger FROM parametreler WHERE anahtar=?",(k,)).fetchone()
        return row[0] if row else default
    katsayi = float(get("vekalet_katsayi","1.387871"))
    carpan   = float(get("vekalet_carpan","20000"))
    ay       = float(get("vekalet_ay","12"))
    damga    = float(get("vekalet_damga","0.00759"))
    import json
    dilimleri_raw = get("vekalet_gv_dilimleri", None)
    if dilimleri_raw:
        try: dilimleri = json.loads(dilimleri_raw)
        except: dilimleri = None
    else: dilimleri = None
    if not dilimleri:
        dilimleri=[
            {"limit":190000,"oran":15},
            {"limit":400000,"oran":20},
            {"limit":1500000,"oran":27},
            {"limit":5300000,"oran":35},
            {"limit":None,"oran":40},
        ]
    conn.close()
    return {"katsayi":katsayi,"carpan":carpan,"ay":ay,"damga":damga,"dilimleri":dilimleri}

def vekalet_param_kaydet(katsayi,carpan,ay,damga,dilimleri):
    import json
    conn=get_conn()
    for anahtar,deger in [
        ("vekalet_katsayi",str(katsayi)),
        ("vekalet_carpan",str(carpan)),
        ("vekalet_ay",str(ay)),
        ("vekalet_damga",str(damga)),
        ("vekalet_gv_dilimleri",json.dumps(dilimleri))
    ]:
        conn.execute("INSERT OR REPLACE INTO parametreler(anahtar,deger) VALUES(?,?)",(anahtar,deger))
    conn.commit(); conn.close()

# ── Parametreler ana penceresi ────────────────────────────────────
class ParametrelerPencere(tk.Toplevel):
    """Ana parametreler menüsü — sekmeli"""
    def __init__(self, parent, acilis="faiz"):
        super().__init__(parent)
        self.title("⚙  Parametreler"); self.configure(bg=CLR["bg"]); self.resizable(True,True)
        center_window(self,640,520); self.grab_set()
        self._build(acilis)

    def _build(self, acilis):
        tk.Label(self,text="⚙  Parametreler",bg=CLR["bg"],fg=CLR["accent"],font=FT,pady=10,padx=16,anchor="w").pack(fill="x")
        tk.Frame(self,bg=CLR["red"],height=2).pack(fill="x")
        nb=ttk.Notebook(self); nb.pack(fill="both",expand=True,padx=12,pady=10)
        style=ttk.Style(); style.configure("TNotebook",background=CLR["bg"]); style.configure("TNotebook.Tab",font=FB,padding=[10,4])

        # Sekme 1: Faiz Oranı
        tab1=tk.Frame(nb,bg=CLR["card"]); nb.add(tab1,text="💰  Faiz Oranı")
        self._build_faiz(tab1)
        # Sekme 2: Vekalet Ücreti
        tab2=tk.Frame(nb,bg=CLR["card"]); nb.add(tab2,text="⚖  Vekalet Ücreti")
        self._build_vekalet(tab2)

        if acilis=="vekalet": nb.select(1)

    # ── Faiz Oranı sekmesi ──────────────────────────────────────
    def _build_faiz(self, parent):
        tk.Label(parent,text="Yasal Faiz Oranları",bg=CLR["card"],fg=CLR["accent"],font=FB,padx=12,pady=8,anchor="w").pack(fill="x")
        tk.Frame(parent,bg=CLR["border"],height=1).pack(fill="x")
        cols=("tarih","oran"); tf,self.faiz_tree=styled_tree(parent,cols,heights=6)
        self.faiz_tree.heading("tarih",text="Geçerlilik Tarihi"); self.faiz_tree.column("tarih",width=180,anchor="center")
        self.faiz_tree.heading("oran",text="Oran (%)"); self.faiz_tree.column("oran",width=120,anchor="center")
        tf.pack(fill="x",padx=12,pady=(8,4))
        self._faiz_load()
        tk.Frame(parent,bg=CLR["border"],height=1).pack(fill="x",padx=12,pady=4)
        tk.Label(parent,text="Yeni Oran Ekle:",bg=CLR["card"],fg=CLR["accent"],font=FB,anchor="w",padx=12).pack(fill="x")
        form=tk.Frame(parent,bg=CLR["card"]); form.pack(fill="x",padx=12,pady=8)
        tk.Label(form,text="Geçerlilik Tarihi:",bg=CLR["card"],fg=CLR["subtext"],font=FS).grid(row=0,column=0,sticky="w",padx=(0,8),pady=4)
        self.faiz_tarih_var=tk.StringVar()
        tarih_entry_with_cal(form,self.faiz_tarih_var,width=12).grid(row=0,column=1,sticky="w",pady=4)
        tk.Label(form,text="Oran (%):",bg=CLR["card"],fg=CLR["subtext"],font=FS).grid(row=1,column=0,sticky="w",padx=(0,8),pady=4)
        self.faiz_oran_var=tk.StringVar()
        styled_entry(form,textvariable=self.faiz_oran_var,width=12).grid(row=1,column=1,sticky="w",pady=4)
        bf=tk.Frame(parent,bg=CLR["card"],pady=8); bf.pack(fill="x",padx=12)
        styled_btn(bf,"➕ Ekle",self._faiz_ekle,color=CLR["success"]).pack(side="left",padx=4)
        styled_btn(bf,"✕ Seçiliyi Sil",self._faiz_sil,color=CLR["danger"]).pack(side="left",padx=4)

    def _faiz_load(self):
        conn=get_conn(); rows=conn.execute("SELECT id,gecerlilik_tarihi,oran FROM faiz_oran ORDER BY id").fetchall(); conn.close()
        self.faiz_tree.delete(*self.faiz_tree.get_children())
        for i,row in enumerate(rows):
            self.faiz_tree.insert("","end",iid=str(row[0]),values=(row[1],f"%{float(row[2]):.2f}"),tags=("odd" if i%2 else "even",))

    def _faiz_ekle(self):
        tarih=self.faiz_tarih_var.get().strip(); oran_s=self.faiz_oran_var.get().strip()
        if not tarih or not oran_s: messagebox.showerror("Hata","Tüm alanlar zorunludur."); return
        try: oran=float(oran_s.replace(",","."))
        except: messagebox.showerror("Hata","Geçerli oran giriniz."); return
        conn=get_conn(); conn.execute("INSERT INTO faiz_oran(gecerlilik_tarihi,oran) VALUES(?,?)",(tarih,oran)); conn.commit(); conn.close()
        self._faiz_load(); self.faiz_tarih_var.set(""); self.faiz_oran_var.set("")

    def _faiz_sil(self):
        sel=self.faiz_tree.focus()
        if not sel: messagebox.showinfo("Bilgi","Oran seçin."); return
        if messagebox.askyesno("Sil","Seçili oran silinecek?"):
            conn=get_conn(); conn.execute("DELETE FROM faiz_oran WHERE id=?",(int(sel),)); conn.commit(); conn.close(); self._faiz_load()

    # ── Vekalet Ücreti sekmesi ──────────────────────────────────
    def _build_vekalet(self, parent):
        p=vekalet_param_getir()
        tk.Label(parent,text="Vekalet Ücreti Parametreleri",bg=CLR["card"],fg=CLR["accent"],font=FB,padx=12,pady=8,anchor="w").pack(fill="x")
        tk.Frame(parent,bg=CLR["border"],height=1).pack(fill="x")

        # Formül göstergesi
        self._vk_formul_lbl=tk.Label(parent,text="",bg="#1a2a1a",fg="#27AE60",font=("Segoe UI",9),padx=12,pady=6,anchor="w")
        self._vk_formul_lbl.pack(fill="x")

        form=tk.Frame(parent,bg=CLR["card"]); form.pack(fill="x",padx=16,pady=12); form.columnconfigure(1,weight=1)

        def row(r,lbl,var,hint=""):
            tk.Label(form,text=lbl,bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=r,column=0,sticky="w",padx=(0,10),pady=6)
            e=styled_entry(form,textvariable=var,width=18); e.grid(row=r,column=1,sticky="w",pady=6)
            if hint: tk.Label(form,text=hint,bg=CLR["card"],fg=CLR["subtext"],font=("Segoe UI",8),anchor="w").grid(row=r,column=2,sticky="w",padx=8)
            return e

        self.v_katsayi=tk.StringVar(value=str(p["katsayi"]).replace(".",","))
        self.v_carpan  =tk.StringVar(value=str(int(p["carpan"])))
        self.v_ay      =tk.StringVar(value=str(int(p["ay"])))
        self.v_damga   =tk.StringVar(value=str(p["damga"]).replace(".",","))

        e1=row(0,"Memur Maaş Katsayısı:",self.v_katsayi,"(örn: 1,387871)")
        e2=row(1,"Çarpan:",               self.v_carpan,  "(örn: 20000)")
        e3=row(2,"Ay:",                   self.v_ay,      "(örn: 12)")
        e4=row(3,"Damga Vergisi Oranı:",  self.v_damga,   "(örn: 0,00759)")

        # Formülü otomatik güncelle
        def guncelle_formul(*_):
            try:
                k=float(self.v_katsayi.get().replace(",",".")); c=float(self.v_carpan.get()); a=float(self.v_ay.get())
                max_v=k*c*a
                self._vk_formul_lbl.config(text=f"  ✓  Azami Vekalet Ücreti = {k} × {int(c):,} × {int(a)} = {para_format(max_v)}")
            except: self._vk_formul_lbl.config(text="  Değerleri doldurun...")
        self.v_katsayi.trace_add("write",guncelle_formul)
        self.v_carpan.trace_add("write",guncelle_formul)
        self.v_ay.trace_add("write",guncelle_formul)
        guncelle_formul()

        tk.Frame(form,bg=CLR["border"],height=1).grid(row=4,column=0,columnspan=3,sticky="ew",pady=8)
        tk.Label(form,text="Gelir Vergisi Dilimleri:",bg=CLR["card"],fg=CLR["accent"],font=FB,anchor="w").grid(row=5,column=0,columnspan=3,sticky="w",pady=(0,6))

        self.gv_vars=[]
        for i,d in enumerate(p["dilimleri"]):
            limit_v=tk.StringVar(value=str(int(d["limit"])) if d["limit"] else "")
            oran_v=tk.StringVar(value=str(d["oran"]))
            lbl_limit=f"Dilim {i+1} — {'Üzeri için' if not d['limit'] else 'Sınır (₺)'}:"
            tk.Label(form,text=lbl_limit,bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=6+i,column=0,sticky="w",padx=(0,10),pady=3)
            f2=tk.Frame(form,bg=CLR["card"]); f2.grid(row=6+i,column=1,columnspan=2,sticky="w",pady=3)
            styled_entry(f2,textvariable=limit_v,width=12).pack(side="left",padx=(0,8))
            tk.Label(f2,text="Oran %:",bg=CLR["card"],fg=CLR["subtext"],font=FS).pack(side="left",padx=(0,4))
            styled_entry(f2,textvariable=oran_v,width=6).pack(side="left")
            self.gv_vars.append((limit_v,oran_v))

        bf=tk.Frame(parent,bg=CLR["card"],pady=10); bf.pack(fill="x",padx=16)
        styled_btn(bf,"💾 Parametreleri Kaydet",self._vekalet_kaydet,color=CLR["success"]).pack(side="left")

    def _sf(self,s,default=0.0):
        try: return float(str(s).replace(",",".").strip())
        except: return default

    def _vekalet_kaydet(self):
        katsayi=self._sf(self.v_katsayi.get(),1.387871)
        carpan =self._sf(self.v_carpan.get(),20000)
        ay     =self._sf(self.v_ay.get(),12)
        damga  =self._sf(self.v_damga.get(),0.00759)
        dilimleri=[]
        for limit_v,oran_v in self.gv_vars:
            limit_s=limit_v.get().strip()
            limit=self._sf(limit_s) if limit_s else None
            oran=self._sf(oran_v.get(),0)
            dilimleri.append({"limit":limit,"oran":oran})
        vekalet_param_kaydet(katsayi,carpan,ay,damga,dilimleri)
        messagebox.showinfo("Başarılı","Vekalet ücreti parametreleri kaydedildi.")

# Eski isim uyumu için alias
OranTanimlama = ParametrelerPencere


# ══════════════════════════════════════════════════════════════════
# MÜKELLEF EKRANI
# ══════════════════════════════════════════════════════════════════
class GirisEkrani(tk.Frame):
    COLS=("no","tckn_vkn","mukellef","il","ilce","adres","iletisim")
    LBLS=("No","TCKN/VKN","Mükellef","İl","İlçe","Adres","İletişim")
    WIDTHS=(55,110,160,80,90,180,130)

    def __init__(self, parent, app):
        super().__init__(parent,bg=CLR["bg"]); self.app=app; self._build(); self.load()

    def _build(self):
        hdr=tk.Frame(self,bg=CLR["bg"]); hdr.pack(fill="x",padx=16,pady=(12,4))
        tk.Label(hdr,text="👤  Mükellef Kayıtları",bg=CLR["bg"],fg=CLR["accent"],font=FT).pack(side="left")
        sf=tk.Frame(self,bg=CLR["panel"],pady=7,padx=12); sf.pack(fill="x",padx=16,pady=4)
        tk.Label(sf,text="Ara (No / TCKN / Ad):",bg=CLR["panel"],fg=CLR["subtext"],font=FS).pack(side="left",padx=(0,6))
        self.ara_var=tk.StringVar(); self.ara_var.trace_add("write",lambda *a:self.load())
        styled_entry(sf,textvariable=self.ara_var,width=26).pack(side="left")
        tf,self.tree=styled_tree(self,self.COLS,heights=17)
        for c,l,w in zip(self.COLS,self.LBLS,self.WIDTHS): self.tree.heading(c,text=l); self.tree.column(c,width=w,minwidth=40,anchor="center")
        tf.pack(fill="both",expand=True,padx=16,pady=(4,4))
        self.tree.bind("<Double-1>",lambda e:self._duzenle_sec())
        self.sayac=tk.Label(self,text="",bg=CLR["bg"],fg=CLR["subtext"],font=FS); self.sayac.pack(anchor="e",padx=20,pady=(0,6))

    def load(self):
        ara=self.ara_var.get().lower()
        conn=get_conn(); rows=conn.execute("SELECT no,tckn_vkn,mukellef,il,ilce,adres,iletisim FROM mukellef ORDER BY no").fetchall(); conn.close()
        self.tree.delete(*self.tree.get_children()); n=0
        for row in rows:
            if ara and not any(ara in str(v).lower() for v in row): continue
            self.tree.insert("","end",iid=str(row[0]),values=row,tags=("odd" if n%2 else "even",)); n+=1
        self.sayac.config(text=f"Toplam: {n} mükellef")

    def yeni_kayit(self):    MukellefForm(self,None,self.load)
    def _duzenle_sec(self):
        sel=self.tree.focus()
        if not sel: return
        conn=get_conn(); row=conn.execute("SELECT * FROM mukellef WHERE no=?",(int(sel),)).fetchone(); conn.close()
        MukellefForm(self,row,self.load)
    def guncelleme(self):
        # Düzenleme: No sor, sonra formu aç
        DuzenleNoSor(self, self.load)
    def kisi_sil(self):
        sel=self.tree.focus()
        if not sel: messagebox.showinfo("Bilgi","Kayıt seçin."); return
        if messagebox.askyesno("Sil","Mükellef silinecek. Emin misiniz?"):
            conn=get_conn(); conn.execute("DELETE FROM mukellef WHERE no=?",(int(sel),)); conn.commit(); conn.close(); self.load()

class DuzenleNoSor(tk.Toplevel):
    """Güncelleme için mükellef no sorma penceresi"""
    def __init__(self, parent, callback):
        super().__init__(parent); self.callback=callback
        self.title("Mükellef Güncelle"); self.configure(bg=CLR["card"]); self.resizable(False,False)
        self._build(); center_window(self,340,160); self.grab_set()

    def _build(self):
        tk.Label(self,text="Güncellenecek Mükellef No:",bg=CLR["card"],fg=CLR["accent"],font=FB,pady=14).pack()
        self.no_var=tk.StringVar()
        e=styled_entry(self,textvariable=self.no_var,width=14); e.pack(pady=4)
        e.bind("<Return>",lambda ev:self.ara())
        bf=tk.Frame(self,bg=CLR["card"],pady=10); bf.pack()
        styled_btn(bf,"🔍 Getir",self.ara,color=CLR["success"]).pack(side="left",padx=8)
        styled_btn(bf,"İptal",self.destroy,color=CLR["danger"]).pack(side="left")
        e.focus_set()

    def ara(self):
        no=self.no_var.get().strip()
        if not no: return
        try: no=int(no)
        except: messagebox.showerror("Hata","Geçerli numara girin."); return
        row=mukellef_getir_no(no)
        if not row: messagebox.showwarning("Bulunamadı",f"No={no} kayıtlı mükellef yok."); return
        self.destroy()
        MukellefForm(self.master, row, self.callback)

class MukellefForm(tk.Toplevel):
    def __init__(self, parent, data, callback):
        super().__init__(parent); self.callback=callback; self.data=data
        self.title("Yeni Mükellef" if not data else "Mükellef Düzenle")
        self.configure(bg=CLR["card"]); self.resizable(False,False)
        self._build()
        if data: self._fill(data)
        center_window(self,480,420); self.grab_set()

    def _build(self):
        tk.Label(self,text="Mükellef Kaydı",bg=CLR["card"],fg=CLR["accent"],font=FT,pady=10).pack(fill="x",padx=16)
        tk.Frame(self,bg=CLR["border"],height=1).pack(fill="x")
        form=tk.Frame(self,bg=CLR["card"]); form.pack(fill="both",padx=16,pady=12); form.columnconfigure(1,weight=1)
        self.entries=[]; self.vars={}

        # Otomatik No
        tk.Label(form,text="No (Otomatik):",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=0,column=0,sticky="w",padx=(0,10),pady=5)
        self.vars["no"]=tk.StringVar(value=str(next_mukellef_no()) if not self.data else "")
        no_e=styled_entry(form,textvariable=self.vars["no"],width=10,state="readonly" if not self.data else "readonly")
        no_e.grid(row=0,column=1,sticky="w",pady=5)

        # TCKN/VKN
        tk.Label(form,text="TCKN/VKN:",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=1,column=0,sticky="w",padx=(0,10),pady=5)
        self.vars["tckn"]=tk.StringVar()
        e=styled_entry(form,textvariable=self.vars["tckn"],width=20); e.grid(row=1,column=1,sticky="ew",pady=5); self.entries.append(e)

        # Mükellef adı
        tk.Label(form,text="Mükellef Adı:",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=2,column=0,sticky="w",padx=(0,10),pady=5)
        self.vars["mukellef"]=tk.StringVar()
        e=styled_entry(form,textvariable=self.vars["mukellef"],width=28); e.grid(row=2,column=1,sticky="ew",pady=5); self.entries.append(e)

        # İl seçimi
        tk.Label(form,text="İl:",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=3,column=0,sticky="w",padx=(0,10),pady=5)
        self.vars["il"]=tk.StringVar()
        il_cb=ttk.Combobox(form,textvariable=self.vars["il"],values=sorted(ILLER.keys()),state="readonly",width=18)
        il_cb.grid(row=3,column=1,sticky="w",pady=5)
        il_cb.bind("<<ComboboxSelected>>",self._il_degis)

        # İlçe seçimi
        tk.Label(form,text="İlçe:",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=4,column=0,sticky="w",padx=(0,10),pady=5)
        self.vars["ilce"]=tk.StringVar()
        self.ilce_cb=ttk.Combobox(form,textvariable=self.vars["ilce"],values=[],state="readonly",width=18)
        self.ilce_cb.grid(row=4,column=1,sticky="w",pady=5)

        # Adres
        tk.Label(form,text="Adres:",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=5,column=0,sticky="w",padx=(0,10),pady=5)
        self.vars["adres"]=tk.StringVar()
        e=styled_entry(form,textvariable=self.vars["adres"],width=32); e.grid(row=5,column=1,sticky="ew",pady=5); self.entries.append(e)

        # İletişim
        tk.Label(form,text="İletişim:",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=6,column=0,sticky="w",padx=(0,10),pady=5)
        self.vars["iletisim"]=tk.StringVar()
        e=styled_entry(form,textvariable=self.vars["iletisim"],width=20); e.grid(row=6,column=1,sticky="ew",pady=5); self.entries.append(e)

        for i in range(len(self.entries)-1):
            self.entries[i].bind("<Tab>",lambda ev,n=i+1:(self.entries[n].focus_set(),"break"))
        self.entries[-1].bind("<Return>",lambda ev:self.kaydet())

        bf=tk.Frame(self,bg=CLR["card"],pady=10); bf.pack(fill="x",padx=16)
        styled_btn(bf,"💾 Kaydet",self.kaydet,color=CLR["success"]).pack(side="left")
        styled_btn(bf,"İptal",self.destroy,color=CLR["danger"]).pack(side="right")

    def _il_degis(self,event=None):
        il=self.vars["il"].get()
        ilceler=ILLER.get(il,["Diğer"])
        self.ilce_cb.config(values=ilceler)
        if ilceler: self.vars["ilce"].set(ilceler[0])

    def _fill(self, row):
        # id,no,tckn_vkn,mukellef,il,ilce,adres,iletisim
        self.vars["no"].set(str(row[1]) if row[1] else "")
        keys=["tckn","mukellef","il","ilce","adres","iletisim"]
        for k,v in zip(keys,row[2:]): self.vars[k].set(str(v) if v else "")
        self._il_degis()
        self.vars["ilce"].set(str(row[5]) if row[5] else "")

    def kaydet(self):
        no_val=int(self.vars["no"].get()) if self.vars["no"].get().isdigit() else next_mukellef_no()
        tckn=self.vars["tckn"].get().strip()
        mukellef=self.vars["mukellef"].get().strip()
        il=self.vars["il"].get().strip(); ilce=self.vars["ilce"].get().strip()
        adres=self.vars["adres"].get().strip(); iletisim=self.vars["iletisim"].get().strip()
        if not tckn or not mukellef: messagebox.showerror("Hata","TCKN/VKN ve Mükellef adı zorunludur."); return
        conn=get_conn()
        try:
            if self.data:
                conn.execute("UPDATE mukellef SET tckn_vkn=?,mukellef=?,il=?,ilce=?,adres=?,iletisim=? WHERE no=?",
                             (tckn,mukellef,il,ilce,adres,iletisim,no_val))
            else:
                conn.execute("INSERT INTO mukellef(no,tckn_vkn,mukellef,il,ilce,adres,iletisim) VALUES(?,?,?,?,?,?,?)",
                             (no_val,tckn,mukellef,il,ilce,adres,iletisim))
            conn.commit()
        except sqlite3.IntegrityError as ex:
            msg="Bu TCKN/VKN zaten kayıtlı." if "tckn" in str(ex) else "Bu numara zaten kayıtlı."
            messagebox.showerror("Hata",msg); conn.close(); return
        conn.close(); self.callback(); self.destroy()


# ══════════════════════════════════════════════════════════════════
# İCMAL EKRANI
# ══════════════════════════════════════════════════════════════════
class IcmalEkrani(tk.Frame):
    COLS=("mukellef_no","tckn_vkn","mukellef","dosya_no","dosya_turu","dosya_durumu","islem_tarihi","uyari_tarihi","aciklama","islem_yapan")
    LBLS=("No","TCKN/VKN","Mükellef","Dosya No","Tür","Durum","İşlem Tarihi","Uyarı Tarihi","Açıklama","İşlem Yapan")
    WIDTHS=(50,100,140,80,110,90,90,90,200,110)

    def __init__(self, parent, app):
        super().__init__(parent,bg=CLR["bg"]); self.app=app; self._build(); self.load()

    def _build(self):
        hdr=tk.Frame(self,bg=CLR["bg"]); hdr.pack(fill="x",padx=16,pady=(12,4))
        tk.Label(hdr,text="📋  İcra İcmal",bg=CLR["bg"],fg=CLR["accent"],font=FT).pack(side="left")
        sf=tk.Frame(self,bg=CLR["panel"],pady=7,padx=12); sf.pack(fill="x",padx=16,pady=4)
        tk.Label(sf,text="Ara:",bg=CLR["panel"],fg=CLR["subtext"],font=FS).pack(side="left",padx=(0,4))
        self.ara_var=tk.StringVar(); self.ara_var.trace_add("write",lambda *a:self.load())
        styled_entry(sf,textvariable=self.ara_var,width=20).pack(side="left",padx=(0,12))
        tk.Label(sf,text="Durum:",bg=CLR["panel"],fg=CLR["subtext"],font=FS).pack(side="left",padx=(0,4))
        self.durum_var=tk.StringVar(value="Tümü")
        cb1=ttk.Combobox(sf,textvariable=self.durum_var,values=["Tümü"]+DURUM_LISTESI,state="readonly",width=11)
        cb1.pack(side="left",padx=(0,10)); cb1.bind("<<ComboboxSelected>>",lambda e:self.load())
        tk.Label(sf,text="Tür:",bg=CLR["panel"],fg=CLR["subtext"],font=FS).pack(side="left",padx=(0,4))
        self.tur_var=tk.StringVar(value="Tümü")
        cb2=ttk.Combobox(sf,textvariable=self.tur_var,values=["Tümü"]+TUR_LISTESI,state="readonly",width=14)
        cb2.pack(side="left"); cb2.bind("<<ComboboxSelected>>",lambda e:self.load())
        bf=tk.Frame(self,bg=CLR["bg"]); bf.pack(fill="x",padx=16,pady=4)
        styled_btn(bf,"⬇ Excel",self.excel_export,color=CLR["accent2"]).pack(side="right",padx=4)
        tf,self.tree=styled_tree(self,self.COLS,heights=16)
        for c,l,w in zip(self.COLS,self.LBLS,self.WIDTHS): self.tree.heading(c,text=l); self.tree.column(c,width=w,minwidth=40,anchor="center")
        tf.pack(fill="both",expand=True,padx=16,pady=(4,4))
        self.tree.bind("<Double-1>",lambda e:self.duzenle())
        self.sayac=tk.Label(self,text="",bg=CLR["bg"],fg=CLR["subtext"],font=FS); self.sayac.pack(anchor="e",padx=20,pady=(0,6))

    def load(self):
        ara=self.ara_var.get().lower(); durum=self.durum_var.get(); tur=self.tur_var.get()
        conn=get_conn()
        rows=conn.execute("SELECT id,mukellef_no,tckn_vkn,mukellef,dosya_no,dosya_turu,dosya_durumu,islem_tarihi,uyari_tarihi,aciklama,islem_yapan FROM icmal ORDER BY id DESC").fetchall()
        conn.close()
        self.tree.delete(*self.tree.get_children()); n=0
        for row in rows:
            vals=row[1:]
            if ara and not any(ara in str(v).lower() for v in vals): continue
            if durum!="Tümü" and vals[5]!=durum: continue
            if tur!="Tümü" and vals[4]!=tur: continue
            self.tree.insert("","end",iid=str(row[0]),values=vals,tags=("odd" if n%2 else "even",)); n+=1
        self.sayac.config(text=f"Toplam: {n} kayıt")

    def yeni_kayit(self):
        conn=get_conn(); mk_count=conn.execute("SELECT COUNT(*) FROM mukellef").fetchone()[0]; conn.close()
        if mk_count==0: messagebox.showwarning("Uyarı","Önce Mükellef ekranından mükellef kaydı yapınız."); return
        IcmalForm(self,None,self.load,self.app)

    def duzenle(self):
        sel=self.tree.focus()
        if not sel: messagebox.showinfo("Bilgi","Kayıt seçin."); return
        conn=get_conn(); row=conn.execute("SELECT * FROM icmal WHERE id=?",(int(sel),)).fetchone(); conn.close()
        IcmalForm(self,row,self.load,self.app)

    def ara_pencere(self):
        IcmalAra(self)

    def sil(self):
        sel=self.tree.focus()
        if not sel: messagebox.showinfo("Bilgi","Kayıt seçin."); return
        if messagebox.askyesno("Sil","Kayıt silinecek. Emin misiniz?"):
            conn=get_conn(); conn.execute("DELETE FROM icmal WHERE id=?",(int(sel),)); conn.commit(); conn.close(); self.load()

    def excel_export(self):
        rows=[self.tree.item(i)["values"] for i in self.tree.get_children()]
        if not rows: messagebox.showinfo("Bilgi","Veri yok."); return
        fp=filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel","*.xlsx")],initialfile="icmal_raporu")
        if not fp: return
        excel_rapor_olustur(fp,"İcmal Raporu",self.LBLS,rows); messagebox.showinfo("Başarılı",f"Excel kaydedildi:\n{fp}")

class IcmalAra(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent); self.parent_ekran=parent
        self.title("İcmal Ara"); self.configure(bg=CLR["card"]); self.resizable(False,False)
        self._build(); center_window(self,420,220); self.grab_set()

    def _build(self):
        tk.Label(self,text="İcmal Ara",bg=CLR["card"],fg=CLR["accent"],font=FT,pady=10).pack(fill="x",padx=16)
        tk.Frame(self,bg=CLR["border"],height=1).pack(fill="x")
        form=tk.Frame(self,bg=CLR["card"]); form.pack(padx=20,pady=14)
        tk.Label(form,text="Arama Kriteri:",bg=CLR["card"],fg=CLR["subtext"],font=FS).grid(row=0,column=0,sticky="w",pady=6)
        self.kriter_var=tk.StringVar(value="No")
        ttk.Combobox(form,textvariable=self.kriter_var,values=["No","Mükellef","TCKN/VKN","Dosya No"],state="readonly",width=14).grid(row=0,column=1,sticky="w",padx=10,pady=6)
        tk.Label(form,text="Değer:",bg=CLR["card"],fg=CLR["subtext"],font=FS).grid(row=1,column=0,sticky="w",pady=6)
        self.deger_var=tk.StringVar()
        e=styled_entry(form,textvariable=self.deger_var,width=22); e.grid(row=1,column=1,padx=10,pady=6)
        e.bind("<Return>",lambda ev:self.ara()); e.focus_set()
        bf=tk.Frame(self,bg=CLR["card"],pady=8); bf.pack()
        styled_btn(bf,"🔍 Ara",self.ara,color=CLR["success"]).pack(side="left",padx=8)
        styled_btn(bf,"Kapat",self.destroy,color=CLR["danger"]).pack(side="left")

    def ara(self):
        deger=self.deger_var.get().strip().lower()
        self.parent_ekran.ara_var.set(deger)
        self.destroy()

class IcmalForm(tk.Toplevel):
    def __init__(self, parent, data, callback, app):
        super().__init__(parent); self.callback=callback; self.data=data; self.app=app
        self._kayit_yapildi=False
        self.title("Yeni İcmal Kaydı" if not data else "İcmal Düzenle")
        self.configure(bg=CLR["card"]); self.resizable(False,False)
        self._build()
        if data: self._fill(data)
        center_window(self,540,480); self.grab_set()

    def _build(self):
        tk.Label(self,text="İcmal Kaydı",bg=CLR["card"],fg=CLR["accent"],font=FT,pady=10).pack(fill="x",padx=16)
        tk.Frame(self,bg=CLR["border"],height=1).pack(fill="x")
        form=tk.Frame(self,bg=CLR["card"]); form.pack(fill="both",padx=16,pady=10)
        form.columnconfigure(1,weight=1); form.columnconfigure(3,weight=1)
        # Mükellef No
        top=tk.Frame(form,bg=CLR["card"]); top.grid(row=0,column=0,columnspan=4,sticky="ew",pady=(0,6))
        tk.Label(top,text="Mükellef No:",bg=CLR["card"],fg=CLR["accent"],font=FB).pack(side="left",padx=(0,8))
        self.mno_var=tk.StringVar()
        e_mno=styled_entry(top,textvariable=self.mno_var,width=10); e_mno.pack(side="left")
        e_mno.bind("<Return>",self._auto_fill); e_mno.bind("<Tab>",self._auto_fill)
        tk.Frame(form,bg=CLR["border"],height=1).grid(row=1,column=0,columnspan=4,sticky="ew",pady=4)
        self.vars={}; self.entries=[]
        fields=[("TCKN/VKN","tckn",None),("Mükellef","mukellef",None),
                ("Dosya No","dosya_no",None),("Dosya Türü","dosya_turu","tur"),
                ("Dosya Durumu","dosya_durumu","durum"),("İşlem Tarihi","islem_tarihi","tarih"),
                ("Uyarı Tarihi","uyari_tarihi","tarih"),("Açıklama","aciklama",None)]
        for i,(lbl,key,tip) in enumerate(fields):
            r,c=(i//2)+2,(i%2)*2
            tk.Label(form,text=lbl+":",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=r,column=c,sticky="w",padx=(4,8),pady=5)
            v=tk.StringVar()
            if tip=="tur": w=ttk.Combobox(form,textvariable=v,values=TUR_LISTESI,state="readonly",width=18)
            elif tip=="durum": w=ttk.Combobox(form,textvariable=v,values=DURUM_LISTESI,state="readonly",width=18)
            elif tip=="tarih":
                w=tarih_entry_with_cal(form,v,width=12)
            else:
                ew=28 if key=="aciklama" else 18
                w=styled_entry(form,textvariable=v,width=ew); self.entries.append(w)
            if isinstance(w,tk.Frame): w.grid(row=r,column=c+1,sticky="w",padx=(0,16),pady=5)
            else: w.grid(row=r,column=c+1,sticky="ew",padx=(0,16),pady=5)
            self.vars[key]=v
        # İşlem yapan
        r_son=2+len(fields)//2
        tk.Label(form,text="İşlem Yapan:",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=r_son,column=0,sticky="w",padx=(4,8),pady=5)
        tk.Label(form,text=self.app.aktif_kullanici["ad_soyad"],bg=CLR["panel"],fg=CLR["accent"],font=FS,anchor="w",padx=8).grid(row=r_son,column=1,sticky="ew",padx=(0,16),pady=5)
        bf=tk.Frame(self,bg=CLR["card"],pady=10); bf.pack(fill="x",padx=16)
        styled_btn(bf,"💾 Kaydet ve Kapat",self.kaydet,color=CLR["success"]).pack(side="left")
        styled_btn(bf,"İptal",self.destroy,color=CLR["danger"]).pack(side="right")

    def _auto_fill(self,event=None):
        no=self.mno_var.get().strip()
        if not no: return
        try: no=int(no)
        except: return
        row=mukellef_getir_no(no)
        if row:
            self.vars["tckn"].set(str(row[2]) if row[2] else "")
            self.vars["mukellef"].set(str(row[3]) if row[3] else "")
        else:
            messagebox.showwarning("Bulunamadı",f"No={no} mükellef kayıtlı değil.")

    def _fill(self,row):
        self.mno_var.set(str(row[1]) if row[1] else "")
        key_order=["tckn","mukellef","dosya_no","dosya_turu","dosya_durumu","islem_tarihi","uyari_tarihi","aciklama"]
        for k,v in zip(key_order,row[2:10]):
            if k in self.vars: self.vars[k].set(str(v) if v else "")

    def kaydet(self):
        mno=self.mno_var.get().strip()
        if not mno: messagebox.showerror("Hata","Mükellef No zorunludur."); return
        try: mno=int(mno)
        except: messagebox.showerror("Hata","Geçerli mükellef no girin."); return
        if not mukellef_getir_no(mno): messagebox.showerror("Hata","Bu no ile kayıtlı mükellef yok.\nÖnce Mükellef ekranından kayıt yapın."); return
        tckn=self.vars["tckn"].get().strip(); mukellef=self.vars["mukellef"].get().strip()
        dosya=self.vars["dosya_no"].get().strip(); dosya_turu=self.vars["dosya_turu"].get().strip()
        durum=self.vars["dosya_durumu"].get().strip()
        islem=self.vars["islem_tarihi"].get().strip(); uyari=self.vars["uyari_tarihi"].get().strip()
        aciklama=self.vars["aciklama"].get().strip()
        yapan=self.app.aktif_kullanici["ad_soyad"]
        if not dosya: messagebox.showerror("Hata","Dosya No zorunludur."); return
        conn=get_conn()
        if self.data:
            conn.execute("UPDATE icmal SET mukellef_no=?,tckn_vkn=?,mukellef=?,dosya_no=?,dosya_turu=?,dosya_durumu=?,islem_tarihi=?,uyari_tarihi=?,aciklama=?,islem_yapan=? WHERE id=?",
                         (mno,tckn,mukellef,dosya,dosya_turu,durum,islem,uyari,aciklama,yapan,self.data[0]))
        else:
            conn.execute("INSERT INTO icmal(mukellef_no,tckn_vkn,mukellef,dosya_no,dosya_turu,dosya_durumu,islem_tarihi,uyari_tarihi,aciklama,islem_yapan) VALUES(?,?,?,?,?,?,?,?,?,?)",
                         (mno,tckn,mukellef,dosya,dosya_turu,durum,islem,uyari,aciklama,yapan))
        conn.commit(); conn.close(); self._kayit_yapildi=True; self.callback(); self.destroy()


# ══════════════════════════════════════════════════════════════════
# AYRIŞTIRAMA, FAİZ, RAPORLAR (kompakt)
# ══════════════════════════════════════════════════════════════════
class AyristirmaEkrani(tk.Frame):
    COLS=("mukellef_no","tarih","dosya_no","tckn_vkn","mukellef","aciklama","anapara","faiz","masraf","harc","vekalet","avans_iadesi","iade_edilecek","toplam","islem_yapan")
    LBLS=("No","Tarih","Dosya No","TCKN/VKN","Mükellef","Açıklama","Anapara","Faiz","Masraf","Harç","Vekalet","Avans İadesi","İade Ed.","Toplam","İşlem Yapan")
    WIDTHS=(50,85,85,100,130,160,80,70,70,70,70,85,80,85,110)

    def __init__(self, parent, app):
        super().__init__(parent,bg=CLR["bg"]); self.app=app; self._build()
        # Açılışta veri gösterilmez

    def _build(self):
        hdr=tk.Frame(self,bg=CLR["bg"]); hdr.pack(fill="x",padx=16,pady=(12,4))
        tk.Label(hdr,text="🔀  Ayrıştırma",bg=CLR["bg"],fg=CLR["accent"],font=FT).pack(side="left")
        sf=tk.Frame(self,bg=CLR["panel"],pady=7,padx=12); sf.pack(fill="x",padx=16,pady=4)
        tk.Label(sf,text="Ara:",bg=CLR["panel"],fg=CLR["subtext"],font=FS).pack(side="left",padx=(0,4))
        self.ara_var=tk.StringVar()
        ara_e=styled_entry(sf,textvariable=self.ara_var,width=22); ara_e.pack(side="left",padx=(0,8))
        ara_e.bind("<Return>",lambda e:self.load())
        styled_btn(sf,"🔍 Ara",self.load,color=CLR["success"]).pack(side="left",padx=4)
        styled_btn(sf,"Tümü",lambda:self._load_all(),color=CLR["panel"]).pack(side="left",padx=4)
        bf=tk.Frame(self,bg=CLR["bg"]); bf.pack(fill="x",padx=16,pady=4)
        styled_btn(bf,"✕ Sil",self.sil,color=CLR["danger"]).pack(side="left",padx=4)
        styled_btn(bf,"⬇ Excel",self.excel_export,color=CLR["accent2"]).pack(side="right",padx=4)
        tf,self.tree=styled_tree(self,self.COLS,heights=14)
        for c,l,w in zip(self.COLS,self.LBLS,self.WIDTHS): self.tree.heading(c,text=l); self.tree.column(c,width=w,minwidth=40,anchor="center")
        tf.pack(fill="both",expand=True,padx=16,pady=(4,4))
        self.tree.bind("<Double-1>",lambda e:self.duzenle())
        self.ozet=tk.Label(self,text="  Arama yapmak için Ara butonuna basın veya Tümü'ne tıklayın.",
                           bg=CLR["panel"],fg=CLR["subtext"],font=FS,pady=5)
        self.ozet.pack(fill="x",padx=16,pady=(2,8))

    def _load_all(self):
        self.ara_var.set(""); self.load()

    def load(self):
        ara=self.ara_var.get().lower()
        conn=get_conn(); rows=conn.execute("SELECT id,mukellef_no,tarih,dosya_no,tckn_vkn,mukellef,aciklama,anapara,faiz,masraf,harc,vekalet,avans_iadesi,iade_edilecek,toplam,islem_yapan FROM ayristirma ORDER BY id DESC").fetchall(); conn.close()
        self.tree.delete(*self.tree.get_children()); count=0; t_ana=t_faiz=t_top=0.0
        for row in rows:
            vals=row[1:]
            if ara and not any(ara in str(v).lower() for v in vals): continue
            disp=list(vals)
            for idx in range(6,14): disp[idx]=para_format(vals[idx])
            self.tree.insert("","end",iid=str(row[0]),values=disp,tags=("odd" if count%2 else "even",))
            t_ana+=float(row[8] or 0); t_faiz+=float(row[9] or 0); t_top+=float(row[15] or 0); count+=1
        self.ozet.config(text=f"  Kayıt: {count}   |   Anapara: {para_format(t_ana)}   Faiz: {para_format(t_faiz)}   Toplam: {para_format(t_top)}",fg=CLR["accent"])

    def yeni_kayit(self):
        conn=get_conn(); mk=conn.execute("SELECT COUNT(*) FROM mukellef").fetchone()[0]; conn.close()
        if mk==0: messagebox.showwarning("Uyarı","Önce mükellef kaydı yapınız."); return
        AyristirmaForm(self,None,self.load,self.app)

    def duzenle(self):
        sel=self.tree.focus()
        if not sel: return
        conn=get_conn(); row=conn.execute("SELECT * FROM ayristirma WHERE id=?",(int(sel),)).fetchone(); conn.close()
        AyristirmaForm(self,row,self.load,self.app)

    def sil(self):
        sel=self.tree.focus()
        if not sel: messagebox.showinfo("Bilgi","Kayıt seçin."); return
        if messagebox.askyesno("Sil","Kayıt silinecek. Emin misiniz?"):
            conn=get_conn(); conn.execute("DELETE FROM ayristirma WHERE id=?",(int(sel),)); conn.commit(); conn.close(); self.load()

    def excel_export(self):
        conn=get_conn(); rows=conn.execute("SELECT mukellef_no,tarih,dosya_no,tckn_vkn,mukellef,aciklama,anapara,faiz,masraf,harc,vekalet,avans_iadesi,iade_edilecek,toplam,islem_yapan FROM ayristirma ORDER BY id DESC").fetchall(); conn.close()
        if not rows: messagebox.showinfo("Bilgi","Veri yok."); return
        fp=filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel","*.xlsx")],initialfile="ayristirma_raporu")
        if not fp: return
        excel_rapor_olustur(fp,"Ayrıştırma Raporu",self.LBLS,rows); messagebox.showinfo("Başarılı",f"Excel kaydedildi:\n{fp}")

class AyristirmaForm(tk.Toplevel):
    def __init__(self,parent,data,callback,app):
        super().__init__(parent); self.callback=callback; self.data=data; self.app=app
        self.title("Yeni Ayrıştırma" if not data else "Ayrıştırma Düzenle")
        self.configure(bg=CLR["card"]); self.resizable(False,False)
        self._build()
        if data: self._fill(data)
        center_window(self,560,520); self.grab_set()

    def _sf(self,s):
        try: return float(str(s).replace(",",".").replace(" ","").replace("₺",""))
        except: return 0.0

    def _build(self):
        tk.Label(self,text="Ayrıştırma Kaydı",bg=CLR["card"],fg=CLR["accent"],font=FT,pady=10).pack(fill="x",padx=16)
        tk.Frame(self,bg=CLR["border"],height=1).pack(fill="x")
        form=tk.Frame(self,bg=CLR["card"]); form.pack(fill="both",padx=16,pady=10)
        form.columnconfigure(1,weight=1); form.columnconfigure(3,weight=1)
        top=tk.Frame(form,bg=CLR["card"]); top.grid(row=0,column=0,columnspan=4,sticky="ew",pady=(0,6))
        tk.Label(top,text="Mükellef No:",bg=CLR["card"],fg=CLR["accent"],font=FB).pack(side="left",padx=(0,8))
        self.mno_var=tk.StringVar()
        e_mno=styled_entry(top,textvariable=self.mno_var,width=10); e_mno.pack(side="left")
        e_mno.bind("<Return>",self._auto_fill); e_mno.bind("<Tab>",self._auto_fill)
        tk.Frame(form,bg=CLR["border"],height=1).grid(row=1,column=0,columnspan=4,sticky="ew",pady=4)
        self.vars={}; self.entries=[]
        for i,(lbl,key,tip) in enumerate([("Tarih","tarih","tarih"),("Dosya No","dosya_no",None),("TCKN/VKN","tckn",None),("Mükellef","mukellef",None)]):
            r,c=(i//2)+2,(i%2)*2
            tk.Label(form,text=lbl+":",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=r,column=c,sticky="w",padx=(4,8),pady=4)
            v=tk.StringVar()
            if tip=="tarih": w=tarih_entry_with_cal(form,v,width=12); w.grid(row=r,column=c+1,sticky="w",padx=(0,16),pady=4)
            else: w=styled_entry(form,textvariable=v,width=18); w.grid(row=r,column=c+1,sticky="ew",padx=(0,16),pady=4); self.entries.append(w)
            self.vars[key]=v
        tk.Label(form,text="Açıklama:",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=4,column=0,sticky="w",padx=(4,8),pady=4)
        self.vars["aciklama"]=tk.StringVar()
        e_ac=styled_entry(form,textvariable=self.vars["aciklama"],width=44)
        e_ac.grid(row=4,column=1,columnspan=3,sticky="ew",padx=(0,16),pady=4); self.entries.append(e_ac)
        tk.Frame(form,bg=CLR["border"],height=1).grid(row=5,column=0,columnspan=4,sticky="ew",pady=6)
        for i,(lbl,key) in enumerate([("Anapara (₺)","anapara"),("Faiz (₺)","faiz"),("Masraf (₺)","masraf"),("Harç (₺)","harc"),("Vekalet (₺)","vekalet"),("Avans İadesi (₺)","avans_iadesi"),("İade Edilecek (₺)","iade_edilecek")]):
            r,c=(i//2)+6,(i%2)*2
            tk.Label(form,text=lbl,bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=r,column=c,sticky="w",padx=(4,8),pady=4)
            v=tk.StringVar(value="0"); v.trace_add("write",self._auto_toplam)
            e=styled_entry(form,textvariable=v,width=16); e.grid(row=r,column=c+1,sticky="ew",padx=(0,16),pady=4)
            self.vars[key]=v; self.entries.append(e)
        r_top=6+(7+1)//2
        tk.Label(form,text="TOPLAM (₺)",bg=CLR["card"],fg=CLR["accent"],font=FB,anchor="w").grid(row=r_top,column=0,sticky="w",padx=(4,8),pady=6)
        self.vars["toplam"]=tk.StringVar(value="0")
        self.toplam_lbl=tk.Label(form,text="0,00 ₺",bg=CLR["panel"],fg=CLR["accent"],font=FB,padx=10,pady=4,anchor="w")
        self.toplam_lbl.grid(row=r_top,column=1,sticky="ew",padx=(0,16),pady=6)
        tk.Label(form,text="İşlem Yapan:",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=r_top,column=2,sticky="w",padx=(4,8),pady=6)
        tk.Label(form,text=self.app.aktif_kullanici["ad_soyad"],bg=CLR["panel"],fg=CLR["accent"],font=FS,padx=8,pady=4,anchor="w").grid(row=r_top,column=3,sticky="ew",padx=(0,16),pady=6)
        bf=tk.Frame(self,bg=CLR["card"],pady=10); bf.pack(fill="x",padx=16)
        styled_btn(bf,"💾 Kaydet",self.kaydet,color=CLR["success"]).pack(side="left")
        styled_btn(bf,"İptal",self.destroy,color=CLR["danger"]).pack(side="right")

    def _auto_toplam(self,*_):
        keys=["anapara","faiz","masraf","harc","vekalet","avans_iadesi","iade_edilecek"]
        top=sum(self._sf(self.vars[k].get()) for k in keys)
        self.vars["toplam"].set(f"{top:.2f}"); self.toplam_lbl.config(text=para_format(top))

    def _auto_fill(self,event=None):
        no=self.mno_var.get().strip()
        if not no: return
        try: no=int(no)
        except: return
        row=mukellef_getir_no(no)
        if row: self.vars["tckn"].set(str(row[2]) if row[2] else ""); self.vars["mukellef"].set(str(row[3]) if row[3] else "")
        else: messagebox.showwarning("Bulunamadı",f"No={no} mükellef kayıtlı değil.")

    def _fill(self,row):
        self.mno_var.set(str(row[1]) if row[1] else "")
        key_order=["tarih","dosya_no","tckn","mukellef","aciklama","anapara","faiz","masraf","harc","vekalet","avans_iadesi","iade_edilecek","toplam"]
        for k,v in zip(key_order,row[2:15]):
            if k in self.vars: self.vars[k].set(str(v) if v is not None else "0")
        self.toplam_lbl.config(text=para_format(self.vars["toplam"].get()))

    def kaydet(self):
        mno=self.mno_var.get().strip()
        if not mno: messagebox.showerror("Hata","Mükellef No zorunludur."); return
        try: mno=int(mno)
        except: messagebox.showerror("Hata","Geçerli mükellef no girin."); return
        if not mukellef_getir_no(mno): messagebox.showerror("Hata","Bu no ile kayıtlı mükellef yok."); return
        tarih=self.vars["tarih"].get().strip(); dosya=self.vars["dosya_no"].get().strip()
        tckn=self.vars["tckn"].get().strip(); mukellef=self.vars["mukellef"].get().strip()
        aciklama=self.vars["aciklama"].get().strip()
        if not tarih: messagebox.showerror("Hata","Tarih zorunludur."); return
        para_keys=["anapara","faiz","masraf","harc","vekalet","avans_iadesi","iade_edilecek"]
        para_vals=[self._sf(self.vars[k].get()) for k in para_keys]; toplam=sum(para_vals)
        yapan=self.app.aktif_kullanici["ad_soyad"]
        conn=get_conn()
        if self.data:
            conn.execute("UPDATE ayristirma SET mukellef_no=?,tarih=?,dosya_no=?,tckn_vkn=?,mukellef=?,aciklama=?,anapara=?,faiz=?,masraf=?,harc=?,vekalet=?,avans_iadesi=?,iade_edilecek=?,toplam=?,islem_yapan=? WHERE id=?",
                         (mno,tarih,dosya,tckn,mukellef,aciklama,*para_vals,toplam,yapan,self.data[0]))
        else:
            conn.execute("INSERT INTO ayristirma(mukellef_no,tarih,dosya_no,tckn_vkn,mukellef,aciklama,anapara,faiz,masraf,harc,vekalet,avans_iadesi,iade_edilecek,toplam,islem_yapan) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                         (mno,tarih,dosya,tckn,mukellef,aciklama,*para_vals,toplam,yapan))
        conn.commit(); conn.close(); self.callback(); self.destroy()


# ══════════════════════════════════════════════════════════════════
# HESAPLAMA (FAİZ) EKRANI
# ══════════════════════════════════════════════════════════════════
class HesaplamaEkrani(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=CLR["bg"]); self.app=app; self._build()

    def _build(self):
        hdr=tk.Frame(self,bg=CLR["bg"]); hdr.pack(fill="x",padx=16,pady=(10,2))
        tk.Label(hdr,text="🧮  Hesaplama",bg=CLR["bg"],fg=CLR["accent"],font=FT).pack(side="left")

        # Kaydırılabilir içerik
        canvas=tk.Canvas(self,bg=CLR["bg"],highlightthickness=0)
        vsb=ttk.Scrollbar(self,orient="vertical",command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right",fill="y"); canvas.pack(side="left",fill="both",expand=True)
        self.inner=tk.Frame(canvas,bg=CLR["bg"])
        self._cwin=canvas.create_window((0,0),window=self.inner,anchor="nw")
        self.inner.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",lambda e:canvas.itemconfig(self._cwin,width=e.width))
        canvas.bind_all("<MouseWheel>",lambda e:canvas.yview_scroll(int(-1*(e.delta/120)),"units"))

        # Faiz oranları bilgi
        info=section_frame(self.inner); info.pack(fill="x",padx=16,pady=(6,4))
        self.oran_lbl=tk.Label(info,text="",bg=CLR["card"],fg=CLR["accent2"],font=FS,padx=12,pady=8)
        self.oran_lbl.pack(fill="x"); self._guncelle_oran_lbl()

        # ══ Yasal Faiz bölümü ══
        yf=section_frame(self.inner,"💰  Yasal Faiz"); yf.pack(fill="x",padx=16,pady=6)
        grid=tk.Frame(yf,bg=CLR["card"]); grid.pack(fill="x",padx=12,pady=10)
        grid.columnconfigure(1,weight=1); grid.columnconfigure(3,weight=1)
        self.ana_var=tk.StringVar(); self.bas_var=tk.StringVar(); self.bit_var=tk.StringVar()
        tk.Label(grid,text="Anapara (₺):",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=0,column=0,sticky="w",padx=(8,6),pady=6)
        e1=styled_entry(grid,textvariable=self.ana_var,width=16); e1.grid(row=0,column=1,sticky="ew",padx=(0,20),pady=6)
        tk.Label(grid,text="Başlangıç:",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=0,column=2,sticky="w",padx=(8,6),pady=6)
        tarih_entry_with_cal(grid,self.bas_var,width=12).grid(row=0,column=3,sticky="w",padx=(0,20),pady=6)
        tk.Label(grid,text="Bitiş:",bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=1,column=0,sticky="w",padx=(8,6),pady=6)
        tarih_entry_with_cal(grid,self.bit_var,width=12).grid(row=1,column=1,sticky="w",padx=(0,20),pady=6)

        # Ek alanlar: Harç, Vekalet Ücreti, Masraf
        tk.Frame(yf,bg=CLR["border"],height=1).pack(fill="x",padx=12,pady=2)
        ek_lbl=tk.Label(yf,text="  Faiz dışı ek kalemler (rapora dahil edilir):",bg=CLR["card"],fg=CLR["subtext"],font=("Segoe UI",9,"italic"),anchor="w")
        ek_lbl.pack(fill="x",padx=12,pady=(4,0))
        ek_grid=tk.Frame(yf,bg=CLR["card"]); ek_grid.pack(fill="x",padx=12,pady=6)
        ek_grid.columnconfigure(1,weight=1); ek_grid.columnconfigure(3,weight=1); ek_grid.columnconfigure(5,weight=1)
        self.harc_var=tk.StringVar(value="0")
        self.vekalet_ek_var=tk.StringVar(value="0")
        self.masraf_var=tk.StringVar(value="0")
        for ci,(lbl,var) in enumerate([("Harç (₺):",self.harc_var),("Vekalet Ücreti (₺):",self.vekalet_ek_var),("Masraf (₺):",self.masraf_var)]):
            tk.Label(ek_grid,text=lbl,bg=CLR["card"],fg=CLR["subtext"],font=FS,anchor="w").grid(row=0,column=ci*2,sticky="w",padx=(8,4),pady=4)
            e=styled_entry(ek_grid,textvariable=var,width=14); e.grid(row=0,column=ci*2+1,sticky="ew",padx=(0,16),pady=4)

        bf_f=tk.Frame(yf,bg=CLR["card"],pady=8); bf_f.pack(fill="x",padx=12)
        styled_btn(bf_f,"🧮 Hesapla",self.faiz_hesapla,color=CLR["success"]).pack(side="left")
        styled_btn(bf_f,"Temizle",self.faiz_temizle,color=CLR["warning"]).pack(side="left",padx=8)
        styled_btn(bf_f,"🖨 Yazdır",self.faiz_yazdir,color=CLR["red"]).pack(side="left",padx=4)
        styled_btn(bf_f,"⬇ PDF İndir",self.faiz_pdf,color="#8B0000").pack(side="left",padx=4)
        styled_btn(bf_f,"⬇ Excel",self.faiz_excel,color=CLR["accent2"]).pack(side="left",padx=4)
        self.faiz_sonuc_frame=section_frame(self.inner,"Yasal Faiz Sonucu"); self.faiz_sonuc_frame.pack(fill="x",padx=16,pady=(0,6))
        self.faiz_sonuc=tk.Frame(self.faiz_sonuc_frame,bg=CLR["card"]); self.faiz_sonuc.pack(fill="x",padx=12,pady=8)
        self._faiz_data=None

        # ══ Vekalet Ödemesi bölümü ══
        tk.Frame(self.inner,bg=CLR["red"],height=2).pack(fill="x",padx=16,pady=4)
        vk=section_frame(self.inner,"⚖  Vekalet Ödemesi"); vk.pack(fill="x",padx=16,pady=6)

        # Parametre bilgi satırı
        self.vk_info_lbl=tk.Label(vk,text="",bg=CLR["card"],fg=CLR["subtext"],font=("Segoe UI",8),padx=12,pady=4,anchor="w")
        self.vk_info_lbl.pack(fill="x"); self._guncelle_vk_info()

        # Max vekalet göstergesi
        self.max_vek_frame=tk.Frame(vk,bg=CLR["card"]); self.max_vek_frame.pack(fill="x",padx=12,pady=(4,0))
        self.max_vek_lbl=tk.Label(self.max_vek_frame,text="",bg=CLR["panel"],fg=CLR["accent"],font=FB,padx=12,pady=6,anchor="w")
        self.max_vek_lbl.pack(fill="x"); self._guncelle_max_vek()

        # Giriş
        gform=tk.Frame(vk,bg=CLR["card"]); gform.pack(fill="x",padx=12,pady=10)
        tk.Label(gform,text="Emanet Toplamı (₺):",bg=CLR["card"],fg=CLR["subtext"],font=FB,anchor="w").pack(side="left",padx=(0,10))
        self.emanet_var=tk.StringVar()
        e_em=styled_entry(gform,textvariable=self.emanet_var,width=18); e_em.pack(side="left")
        e_em.bind("<Return>",lambda ev:self.vekalet_hesapla())
        tk.Label(gform,text="₺",bg=CLR["card"],fg=CLR["subtext"],font=FS).pack(side="left",padx=4)

        bf_v=tk.Frame(vk,bg=CLR["card"],pady=8); bf_v.pack(fill="x",padx=12)
        styled_btn(bf_v,"🧮 Hesapla",self.vekalet_hesapla,color=CLR["success"]).pack(side="left")
        styled_btn(bf_v,"Temizle",self.vekalet_temizle,color=CLR["warning"]).pack(side="left",padx=8)
        styled_btn(bf_v,"🖨 Yazdır",self.vekalet_yazdir,color=CLR["red"]).pack(side="left",padx=4)
        styled_btn(bf_v,"⬇ PDF İndir",self.vekalet_pdf,color="#8B0000").pack(side="left",padx=4)
        styled_btn(bf_v,"⬇ Excel",self.vekalet_excel,color=CLR["accent2"]).pack(side="left",padx=4)

        self.vek_sonuc_frame=section_frame(self.inner,"Vekalet Ödemesi Sonucu"); self.vek_sonuc_frame.pack(fill="x",padx=16,pady=(0,16))
        self.vek_sonuc=tk.Frame(self.vek_sonuc_frame,bg=CLR["card"]); self.vek_sonuc.pack(fill="x",padx=12,pady=8)
        self._vek_data=None

    # ── Yasal Faiz ────────────────────────────────────────────────
    def _guncelle_oran_lbl(self):
        dilimler=faiz_dilimler_db(); parts=[]
        for i,(kesim,oran) in enumerate(dilimler):
            if kesim.year==9999: break
            parts.append(f"{kesim.strftime('%d.%m.%Y')}'e kadar %{float(oran):.0f}")
        self.oran_lbl.config(text="ℹ  Yasal Faiz Oranları:  "+"   |   ".join(parts) if parts else "")

    def _sf_float(self,s,default=0.0):
        try: return float(str(s).replace(",",".").replace(" ","").replace("₺",""))
        except: return default

    def faiz_hesapla(self):
        try: ana=float(self.ana_var.get().replace(",",".").replace(" ",""))
        except: messagebox.showerror("Hata","Geçerli anapara girin."); return
        bas=tarih_parse_safe(self.bas_var.get()); bit=tarih_parse_safe(self.bit_var.get())
        if not bas or not bit: messagebox.showerror("Hata","Tarihleri GG/AA/YYYY formatında girin veya takvimden seçin."); return
        if bit<=bas: messagebox.showerror("Hata","Bitiş > başlangıç olmalı."); return
        res=hesapla_faiz(ana,bas,bit); self._faiz_data=res
        harc    = self._sf_float(self.harc_var.get())
        vekalet = self._sf_float(self.vekalet_ek_var.get())
        masraf  = self._sf_float(self.masraf_var.get())
        genel_toplam = res["toplam"] + harc + vekalet + masraf
        self._faiz_data["harc"]=harc; self._faiz_data["vekalet_ek"]=vekalet
        self._faiz_data["masraf"]=masraf; self._faiz_data["genel_toplam"]=genel_toplam

        for w in self.faiz_sonuc.winfo_children(): w.destroy()
        # Özet kartlar — tek toplam borç
        cards=tk.Frame(self.faiz_sonuc,bg=CLR["card"]); cards.pack(fill="x",pady=(0,8))
        for lbl,val,clr in [
            ("Anapara",       para_format(res["anapara"]),      CLR["subtext"]),
            ("Faiz Tutarı",   para_format(res["faiz_tutari"]),  CLR["warning"]),
            ("Toplam Gün",    f'{res["toplam_gun"]} gün',       CLR["accent2"]),
            ("Anapara + Faiz",para_format(res["toplam"]),       CLR["accent"]),
        ]:
            kf=tk.Frame(cards,bg=CLR["panel"],padx=16,pady=10); kf.pack(side="left",padx=5)
            tk.Label(kf,text=lbl,bg=CLR["panel"],fg=CLR["subtext"],font=FS).pack()
            tk.Label(kf,text=val,bg=CLR["panel"],fg=clr,font=FB).pack()

        # Faiz dilim tablosu
        if res["dilimler"]:
            tk.Label(self.faiz_sonuc,text="Faiz Dilimleri:",bg=CLR["card"],fg=CLR["subtext"],font=FS).pack(anchor="w",pady=(4,2))
            cols=("bas","bit","gun","oran","faiz")
            tf,tree=styled_tree(self.faiz_sonuc,cols,heights=len(res["dilimler"])+1)
            for c,l,w in zip(cols,("Başlangıç","Bitiş","Gün","Oran","Faiz (₺)"),(110,110,65,80,130)):
                tree.heading(c,text=l); tree.column(c,width=w,anchor="center")
            for i,d in enumerate(res["dilimler"]):
                tree.insert("","end",values=(d["baslangic"],d["bitis"],d["gun"],
                    f'%{d["oran"]:.0f}',para_format(d["faiz"])),tags=("odd" if i%2 else "even",))
            tf.pack(fill="x",pady=4)

        # Ek kalemler ve genel toplam
        if harc>0 or vekalet>0 or masraf>0:
            tk.Frame(self.faiz_sonuc,bg=CLR["border"],height=1).pack(fill="x",pady=6)
            ek_frame=tk.Frame(self.faiz_sonuc,bg=CLR["card"]); ek_frame.pack(fill="x",pady=4)
            for lbl,val in [("Harç",harc),("Vekalet Ücreti",vekalet),("Masraf",masraf)]:
                if val>0:
                    row=tk.Frame(ek_frame,bg=CLR["card"]); row.pack(fill="x",padx=4,pady=2)
                    tk.Label(row,text=f"  + {lbl}:",bg=CLR["card"],fg=CLR["subtext"],font=FS,width=20,anchor="w").pack(side="left")
                    tk.Label(row,text=para_format(val),bg=CLR["card"],fg=CLR["text"],font=FB).pack(side="left")
        # Genel toplam
        gt_frame=tk.Frame(self.faiz_sonuc,bg=CLR["panel"],padx=16,pady=10)
        gt_frame.pack(fill="x",pady=(4,0))
        tk.Label(gt_frame,text="GENEL TOPLAM:",bg=CLR["panel"],fg=CLR["subtext"],font=FB).pack(side="left")
        tk.Label(gt_frame,text=para_format(genel_toplam),bg=CLR["panel"],fg=CLR["accent"],font=("Segoe UI",14,"bold")).pack(side="left",padx=16)

    def faiz_temizle(self):
        self.ana_var.set(""); self.bas_var.set(""); self.bit_var.set("")
        self.harc_var.set("0"); self.vekalet_ek_var.set("0"); self.masraf_var.set("0")
        self._faiz_data=None
        for w in self.faiz_sonuc.winfo_children(): w.destroy()

    def _faiz_satirlar(self):
        """Faiz PDF satırlarını hazırla"""
        res=self._faiz_data
        lbls=("Başlangıç","Bitiş","Gün","Oran (%)","Faiz (₺)")
        satirlar=[(d["baslangic"],d["bitis"],d["gun"],f'%{d["oran"]:.0f}',para_format(d["faiz"])) for d in res["dilimler"]]
        satirlar.append(("","","","Anapara",para_format(res["anapara"])))
        satirlar.append(("","","","Faiz Tutarı",para_format(res["faiz_tutari"])))
        satirlar.append(("","","","Anapara + Faiz",para_format(res["toplam"])))
        if res.get("harc",0)>0:       satirlar.append(("","","","Harç",para_format(res["harc"])))
        if res.get("vekalet_ek",0)>0: satirlar.append(("","","","Vekalet Ücreti",para_format(res["vekalet_ek"])))
        if res.get("masraf",0)>0:     satirlar.append(("","","","Masraf",para_format(res["masraf"])))
        satirlar.append(("","","","GENEL TOPLAM",para_format(res.get("genel_toplam",res["toplam"]))))
        return lbls, satirlar

    def faiz_yazdir(self):
        if not self._faiz_data: messagebox.showinfo("Bilgi","Önce hesaplama yapın."); return
        lbls, satirlar = self._faiz_satirlar()
        yazdir_pdf("Yasal Faiz Hesaplama Raporu", lbls, satirlar)

    def faiz_pdf(self):
        if not self._faiz_data: messagebox.showinfo("Bilgi","Önce hesaplama yapın."); return
        fp=filedialog.asksaveasfilename(defaultextension=".pdf",filetypes=[("PDF","*.pdf")],initialfile="yasal_faiz_hesaplama")
        if not fp: return
        lbls, satirlar = self._faiz_satirlar()
        pdf_rapor_olustur(fp,"Yasal Faiz Hesaplama Raporu",lbls,satirlar,None)
        messagebox.showinfo("Başarılı",f"PDF kaydedildi:\n{fp}")

    def faiz_excel(self):
        if not self._faiz_data: messagebox.showinfo("Bilgi","Önce hesaplama yapın."); return
        fp=filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel","*.xlsx")],initialfile="yasal_faiz_hesaplama")
        if not fp: return
        res=self._faiz_data
        lbls=("Başlangıç","Bitiş","Gün","Oran (%)","Faiz (₺)")
        satirlar=[(d["baslangic"],d["bitis"],d["gun"],f'%{d["oran"]:.0f}',para_format(d["faiz"])) for d in res["dilimler"]]
        satirlar.append(("","","","Anapara + Faiz",para_format(res["toplam"])))
        if res.get("harc",0)>0:    satirlar.append(("","","","Harç",para_format(res["harc"])))
        if res.get("vekalet_ek",0)>0: satirlar.append(("","","","Vekalet Ücreti",para_format(res["vekalet_ek"])))
        if res.get("masraf",0)>0:  satirlar.append(("","","","Masraf",para_format(res["masraf"])))
        satirlar.append(("","","","GENEL TOPLAM",para_format(res.get("genel_toplam",res["toplam"]))))
        excel_rapor_olustur(fp,"Yasal Faiz Hesaplama",lbls,satirlar)
        messagebox.showinfo("Başarılı",f"Excel kaydedildi:\n{fp}")

    # ── Vekalet Ödemesi ───────────────────────────────────────────
    def _guncelle_vk_info(self):
        p=vekalet_param_getir()
        self.vk_info_lbl.config(text=f"  Katsayı: {p['katsayi']}   |   Çarpan: {int(p['carpan']):,}   |   Damga Vergisi: %{p['damga']:.5f}   |   Parametreler → Giriş menüsü > Parametreler > Vekalet Ücreti")

    def _guncelle_max_vek(self):
        p=vekalet_param_getir()
        max_v=p["katsayi"]*p["carpan"]*p["ay"]
        self.max_vek_lbl.config(text=f"  Yıllık Azami Vekalet Ücreti:  {p['katsayi']} × {int(p['carpan']):,} × {int(p['ay'])} ay =  {para_format(max_v)}")

    def _sf(self,s):
        try: return Decimal(str(s).replace(",",".").replace(" ","").replace("₺",""))
        except: return Decimal("0")

    def vekalet_hesapla(self):
        try:
            emanet=Decimal(self.emanet_var.get().replace(",",".").replace(" ","").replace("₺",""))
            if emanet<=0: raise ValueError
        except: messagebox.showerror("Hata","Geçerli emanet tutarı girin."); return

        p=vekalet_param_getir()
        max_vek=Decimal(str(p["katsayi"]))*Decimal(str(p["carpan"]))*Decimal(str(p["ay"]))
        damga_oran=Decimal(str(p["damga"]))
        dilimleri=p["dilimleri"]

        # Gelir vergisi hesapla
        kalan=emanet; toplam_gv=Decimal("0"); gv_satirlar=[]
        onceki_limit=Decimal("0")
        for i,d in enumerate(dilimleri):
            limit=Decimal(str(d["limit"])) if d["limit"] else None
            oran=Decimal(str(d["oran"]))/Decimal("100")
            if kalan<=0: break
            if limit:
                dilim_matrah=min(kalan, limit-onceki_limit)
            else:
                dilim_matrah=kalan
            if dilim_matrah<=0:
                onceki_limit=limit if limit else onceki_limit
                continue
            vergi=(dilim_matrah*oran).quantize(Decimal("0.01"),rounding=ROUND_HALF_UP)
            toplam_gv+=vergi
            gv_satirlar.append({
                "dilim":i+1,
                "limit":f"{float(limit):,.0f} ₺" if limit else "Üzeri",
                "matrah":dilim_matrah,
                "oran":f"%{d['oran']}",
                "vergi":vergi,
            })
            kalan-=dilim_matrah
            onceki_limit=limit if limit else onceki_limit

        damga_vergisi=(emanet*damga_oran).quantize(Decimal("0.01"),rounding=ROUND_HALF_UP)
        toplam_kesinti=(toplam_gv+damga_vergisi).quantize(Decimal("0.01"),rounding=ROUND_HALF_UP)
        net_odeme=(emanet-toplam_kesinti).quantize(Decimal("0.01"),rounding=ROUND_HALF_UP)

        self._vek_data={
            "emanet":float(emanet),"max_vekalet":float(max_vek),
            "gelir_vergisi":float(toplam_gv),"damga_vergisi":float(damga_vergisi),
            "toplam_kesinti":float(toplam_kesinti),"net_odeme":float(net_odeme),
            "gv_satirlar":gv_satirlar,
        }

        for w in self.vek_sonuc.winfo_children(): w.destroy()

        # Özet kartlar
        cards=tk.Frame(self.vek_sonuc,bg=CLR["card"]); cards.pack(fill="x",pady=(0,10))
        for lbl,val,clr in [
            ("Emanet Toplamı",para_format(emanet),CLR["subtext"]),
            ("Gelir Vergisi",para_format(toplam_gv),CLR["warning"]),
            ("Damga Vergisi",para_format(damga_vergisi),CLR["warning"]),
            ("Toplam Kesinti",para_format(toplam_kesinti),CLR["danger"]),
            ("Net Ödeme",para_format(net_odeme),CLR["success"]),
        ]:
            kf=tk.Frame(cards,bg=CLR["panel"],padx=14,pady=10); kf.pack(side="left",padx=4)
            tk.Label(kf,text=lbl,bg=CLR["panel"],fg=CLR["subtext"],font=FS).pack()
            tk.Label(kf,text=val,bg=CLR["panel"],fg=clr,font=FB).pack()

        # GV Dilim tablosu
        tk.Label(self.vek_sonuc,text="Gelir Vergisi Dilim Detayı:",bg=CLR["card"],fg=CLR["subtext"],font=FS).pack(anchor="w",pady=(4,2))
        cols=("dilim","limit","matrah","oran","vergi")
        tf,tree=styled_tree(self.vek_sonuc,cols,heights=len(gv_satirlar)+1)
        for c,l,w in zip(cols,("Dilim","Sınır","Matrah (₺)","Oran","Vergi (₺)"),(60,120,120,70,120)):
            tree.heading(c,text=l); tree.column(c,width=w,anchor="center")
        for i,d in enumerate(gv_satirlar):
            tree.insert("","end",values=(
                f"Dilim {d['dilim']}",d["limit"],
                para_format(d["matrah"]),d["oran"],para_format(d["vergi"])
            ),tags=("odd" if i%2 else "even",))
        tf.pack(fill="x",pady=4)

        # Uyarı: max vekalet
        if float(emanet)>float(max_vek):
            tk.Label(self.vek_sonuc,
                text=f"⚠  Not: Emanet toplamı ({para_format(emanet)}), yıllık azami vekalet ücretini ({para_format(max_vek)}) aşmaktadır.",
                bg="#2a1a0a",fg=CLR["warning"],font=FS,padx=10,pady=6,anchor="w").pack(fill="x",pady=4)

    def vekalet_temizle(self):
        self.emanet_var.set(""); self._vek_data=None
        for w in self.vek_sonuc.winfo_children(): w.destroy()

    def _vekalet_satirlar(self):
        d=self._vek_data; p=vekalet_param_getir()
        max_vek=p["katsayi"]*p["carpan"]*p["ay"]
        lbls=("Kalem","Tutar (₺)")
        satirlar=[
            ("Üst Limit Tutarı (Azami Vekalet Ücreti)", para_format(max_vek)),
            ("─"*30, "─"*15),
            ("Emanet Toplamı", para_format(d["emanet"])),
            ("Gelir Vergisi (Toplam)", para_format(d["gelir_vergisi"])),
        ]
        for gs in d["gv_satirlar"]:
            satirlar.append((f"   • {gs['oran']} dilim ({gs['limit']})", para_format(gs["vergi"])))
        satirlar+=[
            ("Damga Vergisi", para_format(d["damga_vergisi"])),
            ("Toplam Kesinti", para_format(d["toplam_kesinti"])),
            ("NET ÖDEME", para_format(d["net_odeme"])),
        ]
        return lbls, satirlar

    def vekalet_yazdir(self):
        if not self._vek_data: messagebox.showinfo("Bilgi","Önce hesaplama yapın."); return
        lbls, satirlar = self._vekalet_satirlar()
        yazdir_pdf("Vekalet Ücreti Ödeme Hesaplama", lbls, satirlar)

    def vekalet_pdf(self):
        if not self._vek_data: messagebox.showinfo("Bilgi","Önce hesaplama yapın."); return
        fp=filedialog.asksaveasfilename(defaultextension=".pdf",filetypes=[("PDF","*.pdf")],initialfile="vekalet_ucret_hesaplama")
        if not fp: return
        lbls, satirlar = self._vekalet_satirlar()
        pdf_rapor_olustur(fp,"Vekalet Ücreti Ödeme Hesaplama",lbls,satirlar,None)
        messagebox.showinfo("Başarılı",f"PDF kaydedildi:\n{fp}")

    def vekalet_excel(self):
        if not self._vek_data: messagebox.showinfo("Bilgi","Önce hesaplama yapın."); return
        fp=filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel","*.xlsx")],initialfile="vekalet_ucret_hesaplama")
        if not fp: return
        d=self._vek_data
        lbls=("Kalem","Tutar (₺)")
        satirlar=[
            ("Emanet Toplamı",para_format(d["emanet"])),
            ("Gelir Vergisi",para_format(d["gelir_vergisi"])),
        ]
        for gs in d["gv_satirlar"]:
            satirlar.append((f"GV {gs['oran']} — {gs['limit']}",para_format(gs["vergi"])))
        satirlar+=[
            ("Damga Vergisi",para_format(d["damga_vergisi"])),
            ("Toplam Kesinti",para_format(d["toplam_kesinti"])),
            ("NET ÖDEME",para_format(d["net_odeme"])),
        ]
        excel_rapor_olustur(fp,"Vekalet Ücreti Hesaplama",lbls,satirlar)
        messagebox.showinfo("Başarılı",f"Excel kaydedildi:\n{fp}")

# ══════════════════════════════════════════════════════════════════
# RAPORLAR EKRANI
# ══════════════════════════════════════════════════════════════════
class RaporlarEkrani(tk.Frame):
    ICMAL_LBLS=("No","TCKN/VKN","Mükellef","Dosya No","Tür","Durum","İşlem Tarihi","Uyarı Tarihi","Açıklama","İşlem Yapan")
    AYRIST_LBLS=("No","Tarih","Dosya No","TCKN/VKN","Mükellef","Açıklama","Anapara","Faiz","Masraf","Harç","Vekalet","Avans İadesi","İade Ed.","Toplam","İşlem Yapan")
    MUKELLEF_LBLS=("No","TCKN/VKN","Mükellef","İl","İlçe","Adres","İletişim")

    def __init__(self, parent):
        super().__init__(parent,bg=CLR["bg"]); self._rows_cache=[]; self._tur="İcmal"; self._build()

    def _build(self):
        hdr=tk.Frame(self,bg=CLR["bg"]); hdr.pack(fill="x",padx=16,pady=(12,4))
        tk.Label(hdr,text="📊  Raporlar",bg=CLR["bg"],fg=CLR["accent"],font=FT).pack(side="left")
        flt=section_frame(self,"Rapor Filtreleri"); flt.pack(fill="x",padx=16,pady=6)
        fp=tk.Frame(flt,bg=CLR["card"]); fp.pack(fill="x",padx=12,pady=8)
        tk.Label(fp,text="Rapor Türü:",bg=CLR["card"],fg=CLR["subtext"],font=FS).pack(side="left",padx=(0,8))
        self.tur_var=tk.StringVar(value="İcmal")
        for t in ["İcmal","Ayrıştırma","Mükellef"]:
            tk.Radiobutton(fp,text=t,variable=self.tur_var,value=t,
                bg=CLR["card"],fg=CLR["text"],selectcolor=CLR["panel"],
                activebackground=CLR["card"],font=F,command=self._tur_degis).pack(side="left",padx=6)
        self.alt=tk.Frame(flt,bg=CLR["card"]); self.alt.pack(fill="x",padx=12,pady=(0,6))
        self._tur_degis()
        imza_f=tk.Frame(flt,bg=CLR["card"]); imza_f.pack(fill="x",padx=12,pady=(0,10))
        tk.Label(imza_f,text="İmzacı:",bg=CLR["card"],fg=CLR["subtext"],font=FS).pack(side="left",padx=(0,8))
        self.imzaci_var=tk.StringVar()
        secs=["(İmzasız)"]+[f"{i['isim']} — {i['unvan']}" for i in IMZACILAR]
        self.imzaci_var.set(secs[0])
        ttk.Combobox(imza_f,textvariable=self.imzaci_var,values=secs,state="readonly",width=30).pack(side="left")
        bf=tk.Frame(self,bg=CLR["bg"]); bf.pack(fill="x",padx=16,pady=4)
        styled_btn(bf,"🔍 Raporu Göster",self.rapor_goster,color=CLR["success"]).pack(side="left",padx=4)
        styled_btn(bf,"🖨 Yazdır",self.yazdir,color=CLR["red"]).pack(side="left",padx=4)
        styled_btn(bf,"⬇ PDF İndir",self.pdf_export,color="#8B0000").pack(side="left",padx=4)
        styled_btn(bf,"⬇ Excel İndir",self.excel_export,color=CLR["accent2"]).pack(side="left",padx=4)
        res=section_frame(self,"Sonuçlar"); res.pack(fill="both",expand=True,padx=16,pady=(4,12))
        self.res_inner=tk.Frame(res,bg=CLR["card"]); self.res_inner.pack(fill="both",expand=True)

    def _tur_degis(self):
        for w in self.alt.winfo_children(): w.destroy()
        self._tur=self.tur_var.get()
        # Sonuçları temizle
        for w in self.res_inner.winfo_children() if hasattr(self,"res_inner") else []: w.destroy()
        self._rows_cache=[]
        if self._tur=="Mükellef":
            tk.Label(self.alt,text="Ara (boş=tümü):",bg=CLR["card"],fg=CLR["subtext"],font=FS).pack(side="left",padx=(0,6))
            self.no_var=tk.StringVar(); styled_entry(self.alt,textvariable=self.no_var,width=20).pack(side="left")
            return
        tk.Label(self.alt,text="Mükellef No (boş=tümü):",bg=CLR["card"],fg=CLR["subtext"],font=FS).pack(side="left",padx=(0,6))
        self.no_var=tk.StringVar()
        no_e=styled_entry(self.alt,textvariable=self.no_var,width=10)
        no_e.pack(side="left",padx=(0,16))
        no_e.bind("<Return>",lambda ev:self.rapor_goster())
        if self._tur=="İcmal":
            tk.Label(self.alt,text="Durum:",bg=CLR["card"],fg=CLR["subtext"],font=FS).pack(side="left",padx=(0,6))
            self.durum_var=tk.StringVar(value="Tümü")
            ttk.Combobox(self.alt,textvariable=self.durum_var,values=["Tümü"]+DURUM_LISTESI,state="readonly",width=12).pack(side="left",padx=(0,10))
            tk.Label(self.alt,text="Tür:",bg=CLR["card"],fg=CLR["subtext"],font=FS).pack(side="left",padx=(0,6))
            self.tur2_var=tk.StringVar(value="Tümü")
            ttk.Combobox(self.alt,textvariable=self.tur2_var,values=["Tümü"]+TUR_LISTESI,state="readonly",width=14).pack(side="left")

    def rapor_goster(self):
        no=self.no_var.get().strip(); conn=get_conn()
        for w in self.res_inner.winfo_children(): w.destroy()
        if self._tur=="İcmal":
            durum=getattr(self,"durum_var",None); tur2=getattr(self,"tur2_var",None)
            q="SELECT mukellef_no,tckn_vkn,mukellef,dosya_no,dosya_turu,dosya_durumu,islem_tarihi,uyari_tarihi,aciklama,islem_yapan FROM icmal WHERE 1=1"
            p=[]
            if no: q+=" AND mukellef_no=?"; p.append(no)
            if durum and durum.get()!="Tümü": q+=" AND dosya_durumu=?"; p.append(durum.get())
            if tur2 and tur2.get()!="Tümü": q+=" AND dosya_turu=?"; p.append(tur2.get())
            rows=conn.execute(q,p).fetchall(); self._rows_cache=rows
            self._show_table(rows,self.ICMAL_LBLS,(50,95,130,75,100,80,85,85,185,105))
        elif self._tur=="Ayrıştırma":
            q="SELECT mukellef_no,tarih,dosya_no,tckn_vkn,mukellef,aciklama,anapara,faiz,masraf,harc,vekalet,avans_iadesi,iade_edilecek,toplam,islem_yapan FROM ayristirma WHERE 1=1"
            p=[]
            if no: q+=" AND mukellef_no=?"; p.append(no)
            rows=conn.execute(q,p).fetchall(); self._rows_cache=rows
            disp=[]
            for row in rows:
                r=list(row)
                for idx in range(6,14): r[idx]=para_format(r[idx])
                disp.append(r)
            self._show_table(disp,self.AYRIST_LBLS,(50,75,75,88,110,120,68,58,58,58,58,72,68,72,105))
        else:  # Mükellef
            q="SELECT no,tckn_vkn,mukellef,il,ilce,adres,iletisim FROM mukellef WHERE 1=1"
            p=[]
            if no: q+=" AND (CAST(no AS TEXT) LIKE ? OR mukellef LIKE ? OR tckn_vkn LIKE ?)"; p+=[f"%{no}%"]*3
            rows=conn.execute(q,p).fetchall(); self._rows_cache=rows
            self._show_table(rows,self.MUKELLEF_LBLS,(60,110,160,80,90,190,130))
        conn.close()

    def _show_table(self,rows,lbls,widths):
        cols=tuple(f"c{i}" for i in range(len(lbls)))
        tf,tree=styled_tree(self.res_inner,cols,heights=13)
        for col,lbl,w in zip(cols,lbls,widths): tree.heading(col,text=lbl); tree.column(col,width=w,anchor="center")
        for i,row in enumerate(rows):
            tree.insert("","end",values=[str(v) if v else "" for v in row],tags=("odd" if i%2 else "even",))
        tf.pack(fill="both",expand=True,padx=4,pady=4)
        tk.Label(self.res_inner,text=f"Toplam {len(rows)} kayıt",bg=CLR["card"],fg=CLR["subtext"],font=FS).pack(anchor="e",padx=8,pady=2)

    def _get_imzaci(self):
        val=self.imzaci_var.get()
        if val=="(İmzasız)": return None
        for im in IMZACILAR:
            if im["isim"] in val: return im
        return None

    def _mukellef_adi(self):
        no=self.no_var.get().strip()
        if no:
            row=mukellef_getir_no(int(no)) if no.isdigit() else None
            if row: return row[3]
        return self._tur

    def _hazirla_rapor(self):
        """Rapor satırlarını hazırla"""
        if self._tur=="İcmal": lbls=self.ICMAL_LBLS; satirlar=self._rows_cache
        elif self._tur=="Ayrıştırma":
            lbls=self.AYRIST_LBLS; satirlar=[]
            for row in self._rows_cache:
                r=list(row)
                for idx in range(6,14): r[idx]=para_format(r[idx])
                satirlar.append(r)
        else: lbls=self.MUKELLEF_LBLS; satirlar=self._rows_cache
        return lbls, satirlar

    def yazdir(self):
        if not self._rows_cache: messagebox.showinfo("Bilgi","Önce raporu gösterin."); return
        lbls, satirlar = self._hazirla_rapor()
        yazdir_pdf(f"{self._tur} Raporu", lbls, satirlar, self._get_imzaci())

    def pdf_export(self):
        if not self._rows_cache: messagebox.showinfo("Bilgi","Önce raporu gösterin."); return
        adi=self._mukellef_adi().replace(" ","_")
        fp=filedialog.asksaveasfilename(defaultextension=".pdf",filetypes=[("PDF","*.pdf")],initialfile=f"{adi}_raporu")
        if not fp: return
        lbls, satirlar = self._hazirla_rapor()
        pdf_rapor_olustur(fp,f"{self._tur} Raporu",lbls,satirlar,self._get_imzaci())
        messagebox.showinfo("Başarılı",f"PDF oluşturuldu:\n{fp}")

    def excel_export(self):
        if not self._rows_cache: messagebox.showinfo("Bilgi","Önce raporu gösterin."); return
        adi=self._mukellef_adi().replace(" ","_")
        fp=filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel","*.xlsx")],initialfile=f"{adi}_raporu")
        if not fp: return
        if self._tur=="İcmal": lbls=self.ICMAL_LBLS
        elif self._tur=="Ayrıştırma": lbls=self.AYRIST_LBLS
        else: lbls=self.MUKELLEF_LBLS
        excel_rapor_olustur(fp,f"{self._tur} Raporu",lbls,self._rows_cache)
        messagebox.showinfo("Başarılı",f"Excel kaydedildi:\n{fp}")


# ══════════════════════════════════════════════════════════════════
# UYARI HATIRLATICI (açılışta geçmiş uyarı tarihleri)
# ══════════════════════════════════════════════════════════════════
def uyari_kontrol(parent):
    """Bugün veya geçmiş uyarı tarihli kayıtları göster — işlem yapıldı takibi ile"""
    bugun = date.today()
    conn = get_conn()

    # Tablo yoksa oluştur
    conn.execute("""CREATE TABLE IF NOT EXISTS uyari_tamamlandi(
        icmal_dosya_no TEXT, uyari_tarihi TEXT,
        PRIMARY KEY (icmal_dosya_no, uyari_tarihi))""")
    conn.commit()

    rows = conn.execute("""SELECT mukellef_no, mukellef, dosya_no, dosya_durumu, uyari_tarihi
        FROM icmal WHERE uyari_tarihi!='' AND uyari_tarihi IS NOT NULL
        ORDER BY uyari_tarihi""").fetchall()

    # Tamamlananları çek
    tamamlananlar = set()
    for r in conn.execute("SELECT icmal_dosya_no, uyari_tarihi FROM uyari_tamamlandi").fetchall():
        tamamlananlar.add((r[0], r[1]))
    conn.close()

    gecmis = []
    for row in rows:
        t = tarih_parse(row[4])
        if t and t <= bugun:
            anahtar = (str(row[2]), str(row[4]))
            if anahtar not in tamamlananlar:
                gecmis.append(row)

    if not gecmis: return

    win = tk.Toplevel(parent)
    win.title("⚠  Uyarı — Tarihi Geçmiş / Bugün Olan Kayıtlar")
    win.configure(bg=CLR["bg"])
    win.resizable(True, True)

    # Başlık
    tk.Label(win, text=f"⚠  {len(gecmis)} adet işlem bekleyen kayıt var!",
             bg="#2a1a0a", fg=CLR["warning"], font=FT,
             pady=10, padx=14).pack(fill="x")
    tk.Label(win, text="'İşlem Yapıldı' butonuna basılan kayıtlar bir daha hatırlatılmaz.",
             bg="#1a1a0a", fg=CLR["subtext"], font=FS, pady=4).pack(fill="x")
    tk.Frame(win, bg=CLR["red"], height=2).pack(fill="x")

    # Scrollable liste
    liste_frame = tk.Frame(win, bg=CLR["bg"])
    liste_frame.pack(fill="both", expand=True, padx=12, pady=10)

    canvas = tk.Canvas(liste_frame, bg=CLR["bg"], highlightthickness=0)
    vsb = ttk.Scrollbar(liste_frame, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=vsb.set)
    vsb.pack(side="right", fill="y"); canvas.pack(side="left", fill="both", expand=True)

    ic = tk.Frame(canvas, bg=CLR["bg"])
    cwin = canvas.create_window((0, 0), window=ic, anchor="nw")
    ic.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(cwin, width=e.width))

    # Sütun başlıkları
    baslik = tk.Frame(ic, bg=CLR["tbl_head"])
    baslik.pack(fill="x", pady=(0,2))
    for txt, w in [("No",50),("Mükellef",160),("Dosya No",100),("Durum",90),("Uyarı Tarihi",100),("İşlem",120)]:
        tk.Label(baslik, text=txt, bg=CLR["tbl_head"], fg=CLR["accent"],
                 font=FB, width=w//7, anchor="center", pady=6).pack(side="left", padx=2)

    satir_widget_listesi = []

    def islem_yapildi(dosya_no, uyari_tarihi, satir_w, idx):
        """İşlem yapıldı olarak işaretle ve satırı gizle"""
        conn2 = get_conn()
        try:
            conn2.execute("INSERT OR REPLACE INTO uyari_tamamlandi(icmal_dosya_no, uyari_tarihi) VALUES(?,?)",
                         (dosya_no, uyari_tarihi))
            conn2.commit()
        except: pass
        conn2.close()
        satir_w.destroy()
        # Kalan satır sayısını güncelle
        kalan = sum(1 for w in satir_widget_listesi if w.winfo_exists())
        baslik_lbl.config(text=f"⚠  {kalan} adet işlem bekleyen kayıt var!" if kalan else "✅  Tüm işlemler tamamlandı!")
        if kalan == 0:
            win.after(1500, win.destroy)

    for i, row in enumerate(gecmis):
        mno, mukellef, dosya_no, durum, uyari_tarihi = row
        bg_renk = CLR["tbl_odd"] if i % 2 else CLR["tbl_even"]

        satir = tk.Frame(ic, bg=bg_renk, pady=2)
        satir.pack(fill="x", pady=1)
        satir_widget_listesi.append(satir)

        for txt, w in [(str(mno),50),(str(mukellef),160),(str(dosya_no),100),(str(durum),90),(str(uyari_tarihi),100)]:
            tk.Label(satir, text=txt, bg=bg_renk, fg=CLR["text"],
                     font=FS, width=w//7, anchor="center", pady=8).pack(side="left", padx=2)

        # İşlem yapıldı butonu
        btn_islem = tk.Button(satir, text="✅ İşlem Yapıldı",
                              bg=CLR["success"], fg="#ffffff",
                              font=("Segoe UI", 9, "bold"),
                              relief="flat", cursor="hand2", padx=8, pady=4,
                              command=lambda dn=dosya_no, ut=uyari_tarihi, sw=satir, idx=i:
                                  islem_yapildi(dn, ut, sw, idx))
        btn_islem.pack(side="left", padx=6)
        btn_islem.bind("<Enter>", lambda e, b=btn_islem: b.config(bg="#1e6b30"))
        btn_islem.bind("<Leave>", lambda e, b=btn_islem: b.config(bg=CLR["success"]))

    # Başlık label'ını dinamik yapmak için referans
    baslik_lbl = win.winfo_children()[0]

    # Kapat butonu
    tk.Frame(win, bg=CLR["border"], height=1).pack(fill="x", padx=12)
    bf = tk.Frame(win, bg=CLR["bg"], pady=10); bf.pack()
    styled_btn(bf, "Kapat", win.destroy, color=CLR["panel"]).pack()

    w_genislik = 660
    y_yukseklik = min(len(gecmis) * 46 + 200, 600)
    center_window(win, w_genislik, y_yukseklik)
    win.grab_set()

# ══════════════════════════════════════════════════════════════════
# ANA PENCERE — AÇILIR MENÜ SİSTEMİ
# ══════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__(); self.withdraw()
        self.title("İcra Yardım Programı v4.0 — Isparta İl Özel İdaresi Hukuk Müşavirliği")
        self.geometry("1280x760"); self.minsize(1000,650)
        self.configure(bg=CLR["bg"])
        self.aktif_kullanici=None
        self._unsaved=False
        init_db(); migrate_db()
        self.after(100,self._show_login)

    def _show_login(self): LoginEkrani(self)

    def after_login(self):
        self.deiconify(); self._build()
        self.after(600,lambda:uyari_kontrol(self))

    def set_unsaved(self, val=True): self._unsaved=val

    def _build(self):
        # ── Üst bar ──
        topbar=tk.Frame(self,bg=CLR["menubar"],height=54); topbar.pack(fill="x"); topbar.pack_propagate(False)
        try:
            import io; from PIL import Image, ImageTk
            img=Image.open(io.BytesIO(base64.b64decode(AMBLEM_B64))).resize((44,44),Image.LANCZOS)
            self._logo=ImageTk.PhotoImage(img)
            tk.Label(topbar,image=self._logo,bg=CLR["menubar"]).pack(side="left",padx=(10,8),pady=4)
        except:
            tk.Label(topbar,text="⚖",bg=CLR["menubar"],fg=CLR["red"],font=("Segoe UI",26)).pack(side="left",padx=(10,8))
        tk.Label(topbar,text="ISPARTA İL ÖZEL İDARESİ  —  HUKUK MÜŞAVİRLİĞİ",
            bg=CLR["menubar"],fg=CLR["red"],font=("Segoe UI",13,"bold")).pack(side="left")
        sag=tk.Frame(topbar,bg=CLR["menubar"]); sag.pack(side="right",padx=14)
        self.tarih_lbl=tk.Label(sag,text="",bg=CLR["menubar"],fg=CLR["subtext"],font=FS); self.tarih_lbl.pack(anchor="e")
        tk.Label(sag,text=f"👤 {self.aktif_kullanici['ad_soyad']}",bg=CLR["menubar"],fg=CLR["accent"],font=FS).pack(anchor="e")

        # Güncelle butonu (sağ üst)
        guncelle_topbar=tk.Button(topbar,text="🔄 Güncelle",
                                   command=lambda:guncelleme_kontrol_ve_goster(self),
                                   bg=CLR["menubar"],fg=CLR["subtext"],
                                   font=("Segoe UI",9),relief="flat",cursor="hand2",
                                   padx=8,pady=4)
        guncelle_topbar.pack(side="right",padx=(0,8))
        guncelle_topbar.bind("<Enter>",lambda e:guncelle_topbar.config(fg=CLR["accent"],bg=CLR["hover"]))
        guncelle_topbar.bind("<Leave>",lambda e:guncelle_topbar.config(fg=CLR["subtext"],bg=CLR["menubar"]))
        self._guncelle_tarih()

        # ── Açılır menü çubuğu ──
        menubar=tk.Frame(self,bg=CLR["panel"],height=46); menubar.pack(fill="x"); menubar.pack_propagate(False)
        tk.Frame(self,bg=CLR["red"],height=3).pack(fill="x")

        self._active="giris"; self.menu_btns={}

        # Menü tanımları: (başlık, key, alt_menüler)
        menu_tanim=[
            ("👤  Giriş","giris",[
                ("👤 Mükellef İşlemleri",None,[
                    ("＋ Yeni Kayıt","giris_yeni"),
                    ("✎ Güncelleme","giris_guncelle"),
                    ("✕ Kişi Sil","giris_sil"),
                ]),
                ("🔐 Kullanıcı Tanımlama","kullanici",None),
                ("⚙ Parametreler",None,[
                    ("💰 Faiz Oranı","oran_tanimlama"),
                    ("⚖ Vekalet Ücreti","vekalet_param"),
                ]),
            ]),
            ("📋  İcmal","icmal",[
                ("＋ Yeni Kayıt","icmal_yeni"),
                ("🔍 Ara","icmal_ara"),
            ]),
            ("🔀  Ayrıştırma","ayristirma",[
                ("＋ Yeni Kayıt","ayristirma_yeni"),
                ("✎ Düzenle","ayristirma_duzenle"),
                ("✕ Sil","ayristirma_sil"),
            ]),
            ("🧮  Hesaplama","hesaplama",None),
            ("📊  Raporlar","raporlar",None),
        ]

        for baslik,key,alt_menu in menu_tanim:
            btn=tk.Button(menubar,text=baslik,bg=CLR["panel"],fg=CLR["text"],
                font=("Segoe UI",11,"bold"),relief="flat",cursor="hand2",
                padx=20,pady=10,
                activebackground=CLR["hover"],activeforeground="#ffffff",
                command=lambda k=key,am=alt_menu:(self.navigate(k) if k and not am else (self._show_dropdown(am, None) if am else None)))
            btn.pack(side="left")
            self.menu_btns[key if key else baslik]=btn
            if alt_menu:
                # Hem tıklayınca hem hover'da açılsın
                btn.config(command=lambda k=key,am=alt_menu,b=btn:(
                    self.navigate(k) if k else None,
                    self._show_dropdown(am, b)
                ))
                btn.bind("<Enter>", lambda e, am=alt_menu, b=btn: self._show_dropdown(am, b))
            btn.bind("<Leave>",lambda e,b=btn,k=key:b.config(bg=CLR["red"] if self._active==k else CLR["panel"]))

        # ── İçerik ──
        self.content=tk.Frame(self,bg=CLR["bg"]); self.content.pack(fill="both",expand=True)
        self.screens={
            "giris":      GirisEkrani(self.content,self),
            "icmal":      IcmalEkrani(self.content,self),
            "ayristirma": AyristirmaEkrani(self.content,self),
            "hesaplama":  HesaplamaEkrani(self.content,self),
            "raporlar":   RaporlarEkrani(self.content),
            "kullanici":  KullaniciEkrani(self.content,self),
        }
        self.navigate("giris")
        self.protocol("WM_DELETE_WINDOW",self._on_close)

    def _menu_enter(self,btn,key):
        if self._active!=key: btn.config(bg=CLR["hover"])

    def _show_dropdown(self, items, parent_btn):
        """Özel şık dropdown menü — tüm item formatlarını destekler"""
        if not items: return

        # Varsa önceki dropdown'ı kapat
        if hasattr(self, '_aktif_dropdown') and self._aktif_dropdown:
            try: self._aktif_dropdown.destroy()
            except: pass
            self._aktif_dropdown = None

        popup = tk.Toplevel(self)
        popup.overrideredirect(True)
        popup.configure(bg=CLR["border"])
        popup.attributes("-topmost", True)
        self._aktif_dropdown = popup

        IKONLAR = {
            "Yeni Kayıt":           ("＋", CLR["success"]),
            "Güncelleme":           ("✎",  CLR["accent"]),
            "Kişi Sil":             ("✕",  CLR["danger"]),
            "Ara":                  ("🔍", CLR["accent2"]),
            "Düzenle":              ("✎",  CLR["accent"]),
            "Sil":                  ("✕",  CLR["danger"]),
            "Kullanıcı Tanımlama":  ("🔐", CLR["accent"]),
            "Parametreler":         ("⚙",  CLR["subtext"]),
            "Faiz Oranı":           ("💰", CLR["warning"]),
            "Vekalet Ücreti":       ("⚖",  CLR["accent"]),
            "Mükellef İşlemleri":   ("👤", CLR["accent2"]),
        }

        main_frame = tk.Frame(popup, bg="#1a2840",
                              highlightthickness=2,
                              highlightbackground=CLR["accent"])
        main_frame.pack(fill="both", expand=True)

        # Kırmızı üst çizgi
        tk.Frame(main_frame, bg=CLR["red"], height=3).pack(fill="x")

        def kapat(event=None):
            try: popup.destroy()
            except: pass
            self._aktif_dropdown = None

        def hover_on(widgets, bg="#2E5080"):
            for w in widgets:
                try: w.config(bg=bg)
                except: pass

        def hover_off(widgets, bg="#1a2840"):
            for w in widgets:
                try: w.config(bg=bg)
                except: pass

        # Alt menü container — sadece bir tane açık olur
        self._alt_container = tk.Frame(main_frame, bg="#1a2840")

        def ekle_satir(parent, lbl, action=None, sub_items=None, alt_bg="#1a2840", indent=0):
            temiz = lbl.strip()
            ikon, ikon_renk = IKONLAR.get(temiz, ("›", CLR["subtext"]))

            row_frame = tk.Frame(parent, bg=alt_bg, cursor="hand2")
            row_frame.pack(fill="x", pady=0)

            ikon_lbl = tk.Label(row_frame, text=ikon, bg=alt_bg, fg=ikon_renk,
                                font=("Segoe UI", 12), padx=10+indent, pady=10)
            ikon_lbl.pack(side="left")

            txt_lbl = tk.Label(row_frame, text=temiz, bg=alt_bg, fg="#FFFFFF",
                               font=("Segoe UI", 10, "bold"), anchor="w", pady=10)
            txt_lbl.pack(side="left", fill="x", expand=True)

            if sub_items:
                arr = tk.Label(row_frame, text=" ›", bg=alt_bg, fg=CLR["subtext"],
                               font=("Segoe UI", 13), padx=8)
                arr.pack(side="right")
                all_w = [row_frame, ikon_lbl, txt_lbl, arr]
            else:
                all_w = [row_frame, ikon_lbl, txt_lbl]

            # Hover
            row_frame.bind("<Enter>",  lambda e, ws=all_w: hover_on(ws))
            row_frame.bind("<Leave>",  lambda e, ws=all_w, bg=alt_bg: hover_off(ws, bg))
            for w in all_w[1:]:
                w.bind("<Enter>",  lambda e, ws=all_w: hover_on(ws))
                w.bind("<Leave>",  lambda e, ws=all_w, bg=alt_bg: hover_off(ws, bg))

            if sub_items:
                def ac_alt(e, si=sub_items, rf=row_frame):
                    # Yana açılan alt popup
                    if hasattr(self, '_aktif_alt_dropdown') and self._aktif_alt_dropdown:
                        try: self._aktif_alt_dropdown.destroy()
                        except: pass

                    alt_popup = tk.Toplevel(self)
                    alt_popup.overrideredirect(True)
                    alt_popup.configure(bg=CLR["border"])
                    alt_popup.attributes("-topmost", True)
                    self._aktif_alt_dropdown = alt_popup

                    alt_frame = tk.Frame(alt_popup, bg="#1C2B3A",
                                        highlightthickness=2,
                                        highlightbackground=CLR["accent"])
                    alt_frame.pack(fill="both", expand=True)
                    tk.Frame(alt_frame, bg=CLR["red"], height=3).pack(fill="x")

                    def kapat_alt(event=None):
                        try: alt_popup.destroy()
                        except: pass
                        self._aktif_alt_dropdown = None

                    for s_lbl, s_action in si:
                        s_temiz = s_lbl.strip()
                        s_ikon, s_renk = IKONLAR.get(s_temiz, ("›", CLR["subtext"]))
                        s_row = tk.Frame(alt_frame, bg="#1C2B3A", cursor="hand2")
                        s_row.pack(fill="x", pady=0)
                        tk.Label(s_row, text=s_ikon, bg="#1C2B3A", fg=s_renk,
                                 font=("Segoe UI",12), padx=12, pady=10).pack(side="left")
                        s_txt = tk.Label(s_row, text=s_temiz, bg="#1C2B3A", fg="#FFFFFF",
                                         font=("Segoe UI",10,"bold"), anchor="w", pady=10, padx=4)
                        s_txt.pack(side="left", fill="x", expand=True)

                        def s_hover_on(e, w=s_row, t=s_txt):
                            w.config(bg="#2E5080"); t.config(bg="#2E5080")
                            for c in w.winfo_children(): c.config(bg="#2E5080")
                        def s_hover_off(e, w=s_row, t=s_txt):
                            w.config(bg="#1C2B3A"); t.config(bg="#1C2B3A")
                            for c in w.winfo_children(): c.config(bg="#1C2B3A")

                        s_row.bind("<Enter>", s_hover_on)
                        s_row.bind("<Leave>", s_hover_off)
                        s_txt.bind("<Enter>", s_hover_on)
                        s_txt.bind("<Leave>", s_hover_off)

                        if s_action:
                            def s_tikla(e, a=s_action):
                                kapat_alt()
                                kapat()
                                self._menu_action(a)
                            s_row.bind("<Button-1>", s_tikla)
                            s_txt.bind("<Button-1>", s_tikla)
                            for c in s_row.winfo_children():
                                c.bind("<Button-1>", s_tikla)

                    # Konumu: ana popup'un sağına hizala
                    alt_popup.update_idletasks()
                    # Ana popup konumunu al
                    popup.update_idletasks()
                    px = popup.winfo_x() + popup.winfo_width()  # sağ kenar
                    # tıklanan satırın dikey konumu
                    rf.update_idletasks()
                    py = rf.winfo_rooty()
                    alt_w = 220
                    alt_h = alt_popup.winfo_reqheight()
                    # Ekran dışına taşmasın
                    sw = self.winfo_screenwidth()
                    if px + alt_w > sw:
                        px = popup.winfo_x() - alt_w
                    alt_popup.geometry(f"{alt_w}x{alt_h}+{px}+{py}")
                    alt_popup.bind("<FocusOut>", lambda e: self.after(150, kapat_alt))
                    alt_popup.focus_set()

                for w in all_w:
                    w.bind("<Button-1>", ac_alt)
            elif action:
                def tikla(e, a=action):
                    kapat()
                    self._menu_action(a)
                for w in all_w:
                    w.bind("<Button-1>", tikla)

        # Menü öğelerini çiz
        for item in items:
            lbl = item[0]
            # item formatı: (lbl, action_str) veya (lbl, None, [sub_items])
            action    = item[1] if len(item) > 1 and isinstance(item[1], str) else None
            sub_items = item[2] if len(item) > 2 and isinstance(item[2], list) else None

            # Kullanıcı Tanımlama gibi direct action (string key)
            if not action and not sub_items and len(item) > 1 and item[1]:
                action = item[1]

            ekle_satir(main_frame, lbl, action=action, sub_items=sub_items)

            # Ayırıcı çizgi
            if sub_items or (len(item) > 1 and item[1] and item[1] not in ("giris_yeni","giris_guncelle","giris_sil","icmal_yeni","icmal_ara","ayristirma_yeni","ayristirma_duzenle","ayristirma_sil","oran_tanimlama","vekalet_param")):
                tk.Frame(main_frame, bg=CLR["border"], height=1).pack(fill="x", padx=8)

        # Alt menü container (başlangıçta gizli)
        self._alt_container.pack_forget()

        # Konum
        if parent_btn:
            x = parent_btn.winfo_rootx()
            y = parent_btn.winfo_rooty() + parent_btn.winfo_height() + 1
        else:
            x = self.winfo_pointerx()
            y = self.winfo_pointery()

        popup.update_idletasks()
        popup.geometry(f"220x{popup.winfo_reqheight()}+{x}+{y}")

        # Dışarı tıklayınca kapat
        popup.bind("<FocusOut>", lambda e: self.after(100, kapat))
        popup.bind("<Escape>", kapat)
        popup.focus_set()

    def _bind_dropdown(self, parent_btn, items):
        parent_btn.bind("<Enter>", lambda e: self._show_dropdown(items, parent_btn))

    def _menu_action(self,action):
        if not action: return
        nav_map={"giris":"giris","icmal":"icmal","ayristirma":"ayristirma",
                 "hesaplama":"hesaplama","raporlar":"raporlar","kullanici":"kullanici"}
        if action in nav_map: self.navigate(nav_map[action]); return
        if action=="giris_yeni":
            self.navigate("giris"); self.after(50, self.screens["giris"].yeni_kayit)
        elif action=="giris_guncelle":
            self.navigate("giris"); self.after(50, self.screens["giris"].guncelleme)
        elif action=="giris_sil":
            self.navigate("giris"); self.after(50, self.screens["giris"].kisi_sil)
        elif action=="icmal_yeni":
            self.navigate("icmal"); self.after(50, self.screens["icmal"].yeni_kayit)
        elif action=="icmal_ara":
            self.navigate("icmal"); self.after(50, self.screens["icmal"].ara_pencere)
        elif action=="ayristirma_yeni":
            self.navigate("ayristirma"); self.after(50, self.screens["ayristirma"].yeni_kayit)
        elif action=="ayristirma_duzenle":
            self.navigate("ayristirma"); self.after(50, self.screens["ayristirma"].duzenle)
        elif action=="ayristirma_sil":
            self.navigate("ayristirma"); self.after(50, self.screens["ayristirma"].sil)
        elif action=="oran_tanimlama":
            ParametrelerPencere(self,"faiz")
        elif action=="vekalet_param":
            ParametrelerPencere(self,"vekalet")

    def navigate(self,key):
        if not key: return
        self._active=key
        for s in self.screens.values(): s.pack_forget()
        if key in self.screens: self.screens[key].pack(fill="both",expand=True)
        for k,btn in self.menu_btns.items():
            active=(k==key or (key=="giris" and k=="giris") or
                    (key=="kullanici" and k=="giris") or
                    (key=="hesaplama" and k=="hesaplama"))
            btn.config(bg=CLR["red"] if active else CLR["panel"],
                fg="#ffffff",
                font=("Segoe UI",10,"bold"))

    def _guncelle_tarih(self):
        self.tarih_lbl.config(text=datetime.now().strftime("%d.%m.%Y  %H:%M"))
        self.after(30000,self._guncelle_tarih)

    def _on_close(self):
        if messagebox.askyesno("Çıkış","Programdan çıkmak istediğinize emin misiniz?",
                               default=messagebox.NO):
            self.destroy()

if __name__=="__main__":
    app=App(); app.mainloop()
